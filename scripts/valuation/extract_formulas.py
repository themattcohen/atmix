"""Extract every formula from the source valuation calculator XLSX.

Writes a markdown report grouped by tab. For each formula-bearing cell we record:
  - cell address
  - formula text
  - cached value (the value Excel computed last time the file was saved)
  - the label cell to the left (column - 1) and above (row - 1), if any, for context

Usage:
    python scripts/valuation/extract_formulas.py <xlsx_path> <output_md_path>
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter


def cell_label(ws, cell: Cell) -> str:
    """Heuristic: nearest non-empty text cell to the left on the same row, then above in the same column."""
    row, col = cell.row, cell.column
    # Try same row, scanning left for the closest text label
    for c in range(col - 1, 0, -1):
        neighbor = ws.cell(row=row, column=c)
        if neighbor.value is None or isinstance(neighbor.value, (int, float)):
            continue
        if isinstance(neighbor.value, str) and neighbor.value.strip() and not neighbor.value.startswith("="):
            return neighbor.value.strip()[:80]
    # Try column header (row 1 or the closest text cell above)
    for r in range(row - 1, 0, -1):
        neighbor = ws.cell(row=r, column=col)
        if neighbor.value is None or isinstance(neighbor.value, (int, float)):
            continue
        if isinstance(neighbor.value, str) and neighbor.value.strip() and not neighbor.value.startswith("="):
            return f"(col header) {neighbor.value.strip()[:80]}"
    return ""


def extract_workbook(xlsx_path: Path, out_path: Path) -> None:
    # data_only=False: get formulas
    wb_f = load_workbook(xlsx_path, data_only=False)
    # data_only=True: get cached computed values
    wb_v = load_workbook(xlsx_path, data_only=True)

    lines: list[str] = []
    lines.append(f"# Source Calculator — Full Formula Extraction (openpyxl)\n")
    lines.append(f"**Source:** `{xlsx_path.name}`")
    lines.append(f"**Method:** openpyxl 3.1, formula + cached-value pull from the XLSX archive XML")
    lines.append("")
    lines.append("Every formula-bearing cell in every sheet. Cached value = what Excel computed last time the file was saved.")
    lines.append("")

    summary_rows: list[dict] = []

    for sheet_name in wb_f.sheetnames:
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]

        formula_cells: list[tuple[str, str, object, str]] = []  # (address, formula, value, label)
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # openpyxl exposes formula values as strings beginning with "="
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    addr = f"{get_column_letter(cell.column)}{cell.row}"
                    value = ws_v[addr].value
                    label = cell_label(ws_f, cell)
                    formula_cells.append((addr, cell.value, value, label))

        if not formula_cells:
            lines.append(f"## Sheet: `{sheet_name}` — no formulas")
            lines.append("")
            continue

        lines.append(f"## Sheet: `{sheet_name}` ({len(formula_cells)} formulas)")
        lines.append("")
        lines.append("| Cell | Label (heuristic) | Formula | Cached Value |")
        lines.append("|---|---|---|---|")
        for addr, formula, value, label in formula_cells:
            # Escape pipes inside formula text
            f_disp = formula.replace("|", "\\|").replace("\n", " ")
            label_disp = label.replace("|", "\\|") if label else ""
            v_disp = repr(value) if value is not None else ""
            v_disp = v_disp.replace("|", "\\|")
            lines.append(f"| `{addr}` | {label_disp} | `{f_disp}` | {v_disp} |")
        lines.append("")
        summary_rows.append({"sheet": sheet_name, "formula_count": len(formula_cells)})

    # Append a JSON summary block at the end for easy machine reuse
    lines.append("---")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append("```json")
    lines.append(json.dumps(summary_rows, indent=2))
    lines.append("```")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path} ({len(lines)} lines, {sum(r['formula_count'] for r in summary_rows)} total formulas)")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: extract_formulas.py <xlsx_path> <output_md_path>", file=sys.stderr)
        sys.exit(2)
    extract_workbook(Path(sys.argv[1]), Path(sys.argv[2]))
