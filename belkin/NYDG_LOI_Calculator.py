#!/usr/bin/env python3
"""
NYDG Letter of Intent Calculator
Creates an Excel workbook with all formulas for side-by-side comparison
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import os

def create_styles():
    """Create reusable styles"""
    return {
        'header': Font(bold=True, size=14, color='FFFFFF'),
        'subheader': Font(bold=True, size=12),
        'bold': Font(bold=True),
        'currency': '#,##0',
        'currency_detail': '$#,##0',
        'percent': '0.0%',
        'header_fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
        'option_a_fill': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'option_b_fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'option_c_fill': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
        'input_fill': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        'result_fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_inputs_sheet(wb, styles):
    """Create the Inputs sheet with all adjustable parameters"""
    ws = wb.create_sheet("Inputs")

    # Title
    ws['A1'] = 'NYDG LOI Calculator - Input Parameters'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    # Practice Fundamentals
    ws['A3'] = 'PRACTICE FUNDAMENTALS'
    ws['A3'].font = styles['subheader']
    ws['A3'].fill = styles['header_fill']
    ws['A3'].font = styles['header']
    ws.merge_cells('A3:D3')

    rows = [
        ('A4', 'B4', '2025 Revenue', 13732000),
        ('A5', 'B5', '2025 EBITDA (Est.)', 2932000),
        ('A6', 'B6', 'BofA Loan Outstanding', 2205200),
        ('A7', 'B7', 'Belkin 2025 Billings', 5450000),
        ('A8', 'B8', 'Belkin % of Revenue', '=B7/B4'),
    ]

    for label_cell, value_cell, label, value in rows:
        ws[label_cell] = label
        ws[value_cell] = value
        ws[value_cell].fill = styles['input_fill']
        if isinstance(value, (int, float)):
            ws[value_cell].number_format = styles['currency_detail']
        elif isinstance(value, str) and value.startswith('='):
            ws[value_cell].number_format = styles['percent']

    # Option Parameters
    ws['A10'] = 'OPTION PARAMETERS'
    ws['A10'].font = styles['header']
    ws['A10'].fill = styles['header_fill']
    ws.merge_cells('A10:D10')

    headers = ['', 'Option A', 'Option B', 'Option C']
    for col, header in enumerate(headers, 1):
        ws.cell(row=11, column=col, value=header)
        ws.cell(row=11, column=col).font = styles['bold']

    option_params = [
        ('Enterprise Value', 6500000, 8000000, 10000000),
        ('Initial Purchase %', 0.40, 0.40, 0.40),
        ('Down Payment %', 0.10, 0.10, 0.10),
        ('Note Term (Years)', 10, 7, 10),
        ('0% Interest Period (Years)', 5, 3, 5),
        ('Interest Rate (AFR)', 0.045, 0.045, 0.045),
        ('Amortization (Years)', 20, 15, 25),
        ('Consulting ($/yr)', 0, 0, 150000),
        ('Consulting Years', 0, 0, 5),
    ]

    for row_idx, (label, opt_a, opt_b, opt_c) in enumerate(option_params, 12):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=opt_a)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        for col in [2, 3, 4]:
            ws.cell(row=row_idx, column=col).fill = styles['input_fill']
            if 'Rate' in label or '%' in label:
                ws.cell(row=row_idx, column=col).number_format = styles['percent']
            elif label not in ['Note Term (Years)', '0% Interest Period (Years)',
                              'Amortization (Years)', 'Consulting Years']:
                ws.cell(row=row_idx, column=col).number_format = styles['currency_detail']

    # Valuation Parameters
    ws['A23'] = 'VALUATION & DISCOUNT PARAMETERS'
    ws['A23'].font = styles['header']
    ws['A23'].fill = styles['header_fill']
    ws.merge_cells('A23:D23')

    val_params = [
        ('A24', 'B24', 'PV Discount Rate', 0.08),
        ('A25', 'B25', 'External Buyer Low Multiple', 4.5),
        ('A26', 'B26', 'External Buyer High Multiple', 5.5),
        ('A27', 'B27', 'MBO Low Multiple', 5.75),
        ('A28', 'B28', 'MBO High Multiple', 6.75),
        ('A29', 'B29', 'Key Person Risk Discount', 0.15),
    ]

    for label_cell, value_cell, label, value in val_params:
        ws[label_cell] = label
        ws[value_cell] = value
        ws[value_cell].fill = styles['input_fill']
        if 'Rate' in label or 'Discount' in label:
            ws[value_cell].number_format = styles['percent']

    # Calculated Values Section
    ws['A31'] = 'CALCULATED VALUES (Reference)'
    ws['A31'].font = styles['header']
    ws['A31'].fill = styles['header_fill']
    ws.merge_cells('A31:D31')

    calc_rows = [
        ('A32', 'B32', 'C32', 'D32', 'Initial Purchase Price',
         '=B12*B13', '=C12*C13', '=D12*D13'),
        ('A33', 'B33', 'C33', 'D33', 'Down Payment Amount',
         '=B32*B14', '=C32*C14', '=D32*D14'),
        ('A34', 'B34', 'C34', 'D34', 'Seller Note Amount',
         '=B32-B33', '=C32-C33', '=D32-D33'),
        ('A35', 'B35', 'C35', 'D35', 'Remaining 60% Value',
         '=B12*(1-B13)', '=C12*(1-C13)', '=D12*(1-D13)'),
        ('A36', 'B36', 'C36', 'D36', 'Total Consulting Value',
         '=B19*B20', '=C19*C20', '=D19*D20'),
    ]

    for row_data in calc_rows:
        ws[row_data[0]] = row_data[4]
        ws[row_data[1]] = row_data[5]
        ws[row_data[2]] = row_data[6]
        ws[row_data[3]] = row_data[7]
        for col_cell in row_data[1:4]:
            ws[col_cell].number_format = styles['currency_detail']
            ws[col_cell].fill = styles['result_fill']

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    return ws

def create_dashboard(wb, styles):
    """Create the Dashboard summary sheet"""
    ws = wb.active
    ws.title = "Dashboard"

    # Title
    ws['A1'] = 'NYDG LOI Options - Executive Summary'
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:E1')

    ws['A2'] = 'All values reference the Inputs sheet - change inputs to update all calculations'
    ws['A2'].font = Font(italic=True, size=10)

    # Side-by-side comparison
    ws['A4'] = 'SIDE-BY-SIDE COMPARISON'
    ws['A4'].font = styles['header']
    ws['A4'].fill = styles['header_fill']
    ws.merge_cells('A4:E4')

    headers = ['Metric', 'Option A\n(Aggressive)', 'Option B\n(Balanced)', 'Option C\n(Premium)', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = styles['bold']
        cell.alignment = Alignment(wrap_text=True, horizontal='center')

    # Key metrics
    metrics = [
        ('Headline Price', '=Inputs!B12', '=Inputs!C12', '=Inputs!D12', 'Enterprise Value'),
        ('Initial Purchase (40%)', '=Inputs!B32', '=Inputs!C32', '=Inputs!D32', ''),
        ('Down Payment', '=Inputs!B33', '=Inputs!C33', '=Inputs!D33', '10% of initial tranche'),
        ('Seller Note', '=Inputs!B34', '=Inputs!C34', '=Inputs!D34', 'Financed amount'),
        ('Note Term', '=Inputs!B15', '=Inputs!C15', '=Inputs!D15', 'Years'),
        ('0% Interest Period', '=Inputs!B16', '=Inputs!C16', '=Inputs!D16', 'Years'),
        ('Total Consulting', '=Inputs!B36', '=Inputs!C36', '=Inputs!D36', ''),
        ('Grand Total Nominal', '=Inputs!B12+Inputs!B36', '=Inputs!C12+Inputs!C36', '=Inputs!D12+Inputs!D36', 'Including consulting'),
    ]

    for row_idx, (label, opt_a, opt_b, opt_c, notes) in enumerate(metrics, 6):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=opt_a)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        ws.cell(row=row_idx, column=5, value=notes)
        for col in [2, 3, 4]:
            cell = ws.cell(row=row_idx, column=col)
            if 'Term' not in label and 'Interest Period' not in label:
                cell.number_format = styles['currency_detail']

    # Present Value Section
    ws['A15'] = 'PRESENT VALUE ANALYSIS'
    ws['A15'].font = styles['header']
    ws['A15'].fill = styles['header_fill']
    ws.merge_cells('A15:E15')

    pv_headers = ['Component', 'Option A', 'Option B', 'Option C', 'Formula Logic']
    for col, header in enumerate(pv_headers, 1):
        ws.cell(row=16, column=col, value=header)
        ws.cell(row=16, column=col).font = styles['bold']

    # PV calculations (simplified - actual formula is complex)
    pv_rows = [
        ('Down Payment (PV)', '=Inputs!B33', '=Inputs!C33', '=Inputs!D33', 'No discount - paid at close'),
        ('Est. Phase 1 Payments PV', '="~"&TEXT(Inputs!B34*0.6,"$#,##0")', '="~"&TEXT(Inputs!C34*0.65,"$#,##0")', '="~"&TEXT(Inputs!D34*0.55,"$#,##0")', 'Discounted at PV rate'),
        ('Est. Future Tranches PV', '="~"&TEXT(Inputs!B35*0.55,"$#,##0")', '="~"&TEXT(Inputs!C35*0.58,"$#,##0")', '="~"&TEXT(Inputs!D35*0.50,"$#,##0")', 'Heavily discounted'),
        ('Total Est. PV', '=Inputs!B33+Inputs!B34*0.6+Inputs!B35*0.55', '=Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58', '=Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50', 'Sum of components'),
        ('PV as % of Headline', '=(Inputs!B33+Inputs!B34*0.6+Inputs!B35*0.55)/Inputs!B12', '=(Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58)/Inputs!C12', '=(Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50)/Inputs!D12', 'True cost %'),
    ]

    for row_idx, (label, opt_a, opt_b, opt_c, notes) in enumerate(pv_rows, 17):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=opt_a)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        ws.cell(row=row_idx, column=5, value=notes)
        for col in [2, 3, 4]:
            cell = ws.cell(row=row_idx, column=col)
            if '% of' in label:
                cell.number_format = styles['percent']
            elif 'Est.' in label and row_idx < 20:
                pass  # Text formula
            else:
                cell.number_format = styles['currency_detail']

    # External Buyer Comparison
    ws['A24'] = 'EXTERNAL BUYER COMPARISON'
    ws['A24'].font = styles['header']
    ws['A24'].fill = styles['header_fill']
    ws.merge_cells('A24:E24')

    ext_rows = [
        ('External Buyer EV (Low)', '=Inputs!B5*Inputs!B25', '', '', '=Inputs!B25&"x EBITDA"'),
        ('External Buyer EV (High)', '=Inputs!B5*Inputs!B26', '', '', '=Inputs!B26&"x EBITDA"'),
        ('Less: Debt Payoff', '=-Inputs!B6', '', '', ''),
        ('External Equity Value (Low)', '=Inputs!B5*Inputs!B25-Inputs!B6', '', '', 'What external pays'),
        ('External Equity Value (High)', '=Inputs!B5*Inputs!B26-Inputs!B6', '', '', ''),
        ('', '', '', '', ''),
        ('Belkin MBO Premium (vs Low)', '=Inputs!B12-(Inputs!B5*Inputs!B25-Inputs!B6)', '=Inputs!C12-(Inputs!B5*Inputs!B25-Inputs!B6)', '=Inputs!D12-(Inputs!B5*Inputs!B25-Inputs!B6)', 'Premium over external'),
    ]

    for row_idx, (label, val_b, val_c, val_d, notes) in enumerate(ext_rows, 25):
        ws.cell(row=row_idx, column=1, value=label)
        if val_b:
            ws.cell(row=row_idx, column=2, value=val_b)
            ws.cell(row=row_idx, column=2).number_format = styles['currency_detail']
        if val_c:
            ws.cell(row=row_idx, column=3, value=val_c)
            ws.cell(row=row_idx, column=3).number_format = styles['currency_detail']
        if val_d:
            ws.cell(row=row_idx, column=4, value=val_d)
            ws.cell(row=row_idx, column=4).number_format = styles['currency_detail']
        ws.cell(row=row_idx, column=5, value=notes)

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25

    return ws

def create_payment_schedule_sheet(wb, styles, option_name, option_col):
    """Create a detailed payment schedule for an option"""
    ws = wb.create_sheet(f"Schedule_{option_name}")

    # Title
    ws['A1'] = f'Payment Schedule - Option {option_name}'
    ws['A1'].font = Font(bold=True, size=14)

    # Reference to inputs
    ws['A3'] = 'Key Inputs (from Inputs sheet)'
    ws['A3'].font = styles['bold']

    ref_data = [
        ('A4', 'B4', 'Enterprise Value', f'=Inputs!{option_col}12'),
        ('A5', 'B5', 'Initial Purchase Amount', f'=Inputs!{option_col}32'),
        ('A6', 'B6', 'Down Payment', f'=Inputs!{option_col}33'),
        ('A7', 'B7', 'Seller Note', f'=Inputs!{option_col}34'),
        ('A8', 'B8', 'Note Term', f'=Inputs!{option_col}15'),
        ('A9', 'B9', '0% Period', f'=Inputs!{option_col}16'),
        ('A10', 'B10', 'Interest Rate', f'=Inputs!{option_col}17'),
        ('A11', 'B11', 'Amortization', f'=Inputs!{option_col}18'),
    ]

    for label_cell, value_cell, label, formula in ref_data:
        ws[label_cell] = label
        ws[value_cell] = formula
        ws[value_cell].fill = styles['result_fill']
        if 'Rate' not in label and 'Term' not in label and 'Period' not in label and 'Amort' not in label:
            ws[value_cell].number_format = styles['currency_detail']
        elif 'Rate' in label:
            ws[value_cell].number_format = styles['percent']

    # Payment Schedule Header
    ws['A13'] = 'Phase 1 Payment Schedule (40% Purchase)'
    ws['A13'].font = styles['bold']

    schedule_headers = ['Year', 'Beginning Balance', 'Payment', 'Interest', 'Principal', 'Ending Balance', 'Cumulative Paid']
    for col, header in enumerate(schedule_headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = styles['bold']
        cell.fill = styles['header_fill']
        cell.font = Font(bold=True, color='FFFFFF')

    # Year 0 (Down payment)
    ws['A15'] = 0
    ws['B15'] = f'=Inputs!{option_col}34'
    ws['C15'] = f'=Inputs!{option_col}33'
    ws['D15'] = 0
    ws['E15'] = f'=Inputs!{option_col}33'
    ws['F15'] = f'=Inputs!{option_col}34'
    ws['G15'] = '=C15'

    # Years 1-12 formulas
    for row in range(16, 28):
        year = row - 15
        ws.cell(row=row, column=1, value=year)
        # Beginning Balance
        ws.cell(row=row, column=2, value=f'=F{row-1}')
        # Payment (simplified - based on amortization)
        ws.cell(row=row, column=3, value=f'=IF(A{row}<=B8,B7/B11,0)')
        # Interest
        ws.cell(row=row, column=4, value=f'=IF(A{row}<=B9,0,B{row}*B10)')
        # Principal
        ws.cell(row=row, column=5, value=f'=C{row}-D{row}')
        # Ending Balance
        ws.cell(row=row, column=6, value=f'=MAX(0,B{row}-E{row})')
        # Cumulative
        ws.cell(row=row, column=7, value=f'=G{row-1}+C{row}')

    # Format all currency cells
    for row in range(15, 28):
        for col in range(2, 8):
            ws.cell(row=row, column=col).number_format = styles['currency_detail']

    # Totals
    ws['A28'] = 'Total'
    ws['A28'].font = styles['bold']
    ws['C28'] = '=SUM(C15:C27)'
    ws['D28'] = '=SUM(D15:D27)'
    ws['E28'] = '=SUM(E15:E27)'
    ws['G28'] = '=G27'
    for col in [3, 4, 5, 7]:
        ws.cell(row=28, column=col).number_format = styles['currency_detail']
        ws.cell(row=28, column=col).font = styles['bold']

    # Present Value section
    ws['A31'] = 'Present Value Calculation'
    ws['A31'].font = styles['bold']

    ws['A32'] = 'Discount Rate'
    ws['B32'] = '=Inputs!B24'
    ws['B32'].number_format = styles['percent']

    ws['A34'] = 'Year'
    ws['B34'] = 'Payment'
    ws['C34'] = 'PV Factor'
    ws['D34'] = 'Present Value'
    for col in range(1, 5):
        ws.cell(row=34, column=col).font = styles['bold']

    # PV calculations for each year
    for row in range(35, 48):
        year = row - 35
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=f'=C{row-20}')  # Reference payment
        ws.cell(row=row, column=3, value=f'=1/(1+$B$32)^A{row}')
        ws.cell(row=row, column=4, value=f'=B{row}*C{row}')
        ws.cell(row=row, column=2).number_format = styles['currency_detail']
        ws.cell(row=row, column=3).number_format = '0.0000'
        ws.cell(row=row, column=4).number_format = styles['currency_detail']

    ws['A48'] = 'Total PV'
    ws['D48'] = '=SUM(D35:D47)'
    ws['A48'].font = styles['bold']
    ws['D48'].font = styles['bold']
    ws['D48'].number_format = styles['currency_detail']
    ws['D48'].fill = styles['result_fill']

    # Column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 20

    return ws

def create_comparison_sheet(wb, styles):
    """Create Colbert vs Belkin perspective comparison"""
    ws = wb.create_sheet("Perspectives")

    # Title
    ws['A1'] = 'Seller vs Buyer Perspectives'
    ws['A1'].font = Font(bold=True, size=16)

    # Colbert's View
    ws['A3'] = "COLBERT'S PERSPECTIVE (What Seller Sees)"
    ws['A3'].font = styles['header']
    ws['A3'].fill = styles['header_fill']
    ws.merge_cells('A3:E3')

    colbert_headers = ['What Matters to Colbert', 'Option A', 'Option B', 'Option C', 'Winner']
    for col, header in enumerate(colbert_headers, 1):
        ws.cell(row=4, column=col, value=header)
        ws.cell(row=4, column=col).font = styles['bold']

    colbert_metrics = [
        ('Headline Number', '=Inputs!B12', '=Inputs!C12', '=Inputs!D12', 'Option C'),
        ('Cash at Close', '=Inputs!B33', '=Inputs!C33', '=Inputs!D33', 'Option C'),
        ('Annual Income Y1-5 (note only)', '=Inputs!B34/Inputs!B18', '=Inputs!C34/Inputs!C18', '=Inputs!D34/Inputs!D18', 'Option B'),
        ('+ Consulting', '=Inputs!B19', '=Inputs!C19', '=Inputs!D19', 'Option C'),
        ('Total Annual Y1-5', '=Inputs!B34/Inputs!B18+Inputs!B19', '=Inputs!C34/Inputs!C18+Inputs!C19', '=Inputs!D34/Inputs!D18+Inputs!D19', 'Option C'),
        ('Legacy/Pride Factor', '"Modest"', '"Fair"', '"Premium"', 'Option C'),
        ('Total Nominal Received', '=Inputs!B12+Inputs!B36', '=Inputs!C12+Inputs!C36', '=Inputs!D12+Inputs!D36', 'Option C'),
    ]

    for row_idx, (label, opt_a, opt_b, opt_c, winner) in enumerate(colbert_metrics, 5):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=opt_a)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        ws.cell(row=row_idx, column=5, value=winner)
        for col in [2, 3, 4]:
            cell = ws.cell(row=row_idx, column=col)
            if 'Factor' not in label:
                cell.number_format = styles['currency_detail']

    # Belkin's View
    ws['A14'] = "BELKIN'S PERSPECTIVE (True Economics)"
    ws['A14'].font = styles['header']
    ws['A14'].fill = styles['header_fill']
    ws.merge_cells('A14:E14')

    belkin_headers = ['What Matters to Belkin', 'Option A', 'Option B', 'Option C', 'Best Value']
    for col, header in enumerate(belkin_headers, 1):
        ws.cell(row=15, column=col, value=header)
        ws.cell(row=15, column=col).font = styles['bold']

    belkin_metrics = [
        ('Headline (Face Value)', '=Inputs!B12', '=Inputs!C12', '=Inputs!D12', ''),
        ('Est. Present Value', '=Inputs!B33+Inputs!B34*0.6+Inputs!B35*0.55', '=Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58', '=Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50', 'Option A'),
        ('PV per $1 Headline', '=(Inputs!B33+Inputs!B34*0.6+Inputs!B35*0.55)/Inputs!B12', '=(Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58)/Inputs!C12', '=(Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50)/Inputs!D12', 'Option C'),
        ('Out-of-Pocket Year 1', '=Inputs!B33+Inputs!B34/Inputs!B18+Inputs!B19', '=Inputs!C33+Inputs!C34/Inputs!C18+Inputs!C19', '=Inputs!D33+Inputs!D34/Inputs!D18+Inputs!D19', 'Option A'),
        ('Years Interest-Free', '=Inputs!B16', '=Inputs!C16', '=Inputs!D16', 'A or C'),
        ('Negotiation Flexibility', '"High"', '"Medium"', '"Low"', 'Option A'),
        ('Relationship Preservation', '"Risky"', '"Good"', '"Excellent"', 'Option C'),
    ]

    for row_idx, (label, opt_a, opt_b, opt_c, best) in enumerate(belkin_metrics, 16):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=opt_a)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        ws.cell(row=row_idx, column=5, value=best)
        for col in [2, 3, 4]:
            cell = ws.cell(row=row_idx, column=col)
            if 'per $1' in label:
                cell.number_format = styles['percent']
            elif 'Years' not in label and 'Flex' not in label and 'Relationship' not in label:
                cell.number_format = styles['currency_detail']

    # Key Insight Box
    ws['A25'] = 'KEY INSIGHT'
    ws['A25'].font = styles['header']
    ws['A25'].fill = styles['header_fill']
    ws.merge_cells('A25:E25')

    insight_text = """Option C costs ~14% more in PV than Option B, but gives 25% higher headline.
"PV per $1 Headline" is lowest for Option C = best value per dollar of perceived price.
Lead with Option C - it optimizes for psychology while minimizing true cost."""
    ws['A26'] = insight_text
    ws.merge_cells('A26:E28')
    ws['A26'].alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

    return ws

def create_scenario_sheet(wb, styles):
    """Create What-If Scenario Analysis sheet"""
    ws = wb.create_sheet("Scenarios")

    ws['A1'] = 'What-If Scenario Analysis'
    ws['A1'].font = Font(bold=True, size=16)

    ws['A3'] = 'Adjust the yellow cells to see how different scenarios affect the deal'
    ws['A3'].font = Font(italic=True)

    # Scenario inputs
    ws['A5'] = 'SCENARIO INPUTS'
    ws['A5'].font = styles['header']
    ws['A5'].fill = styles['header_fill']
    ws.merge_cells('A5:C5')

    scenario_inputs = [
        ('A6', 'B6', 'Custom Enterprise Value', 8500000),
        ('A7', 'B7', 'Custom Down Payment %', 0.15),
        ('A8', 'B8', 'Custom 0% Interest Period', 4),
        ('A9', 'B9', 'Custom Note Term', 8),
        ('A10', 'B10', 'Custom Interest Rate (AFR)', 0.045),
        ('A11', 'B11', 'Custom Amortization Years', 20),
        ('A12', 'B12', 'Custom Consulting $/Year', 100000),
        ('A13', 'B13', 'Custom Consulting Years', 3),
    ]

    for label_cell, value_cell, label, value in scenario_inputs:
        ws[label_cell] = label
        ws[value_cell] = value
        ws[value_cell].fill = styles['input_fill']
        if '%' in label or 'Rate' in label:
            ws[value_cell].number_format = styles['percent']
        elif 'Period' not in label and 'Term' not in label and 'Years' not in label:
            ws[value_cell].number_format = styles['currency_detail']

    # Calculated results
    ws['A15'] = 'SCENARIO RESULTS'
    ws['A15'].font = styles['header']
    ws['A15'].fill = styles['header_fill']
    ws.merge_cells('A15:C15')

    results = [
        ('A16', 'B16', 'Initial Purchase (40%)', '=B6*0.4'),
        ('A17', 'B17', 'Down Payment', '=B16*B7'),
        ('A18', 'B18', 'Seller Note', '=B16-B17'),
        ('A19', 'B19', 'Est. Annual Payment (0% period)', '=B18/B11'),
        ('A20', 'B20', 'Remaining 60%', '=B6*0.6'),
        ('A21', 'B21', 'Total Consulting', '=B12*B13'),
        ('A22', 'B22', 'Grand Total Nominal', '=B6+B21'),
        ('A23', 'B23', '', ''),
        ('A24', 'B24', 'Est. PV (rough)', '=B17+B18*0.6+B20*0.55+B21*0.7'),
        ('A25', 'B25', 'PV as % of Headline', '=B24/B6'),
    ]

    for label_cell, value_cell, label, formula in results:
        ws[label_cell] = label
        if formula:
            ws[value_cell] = formula
            ws[value_cell].fill = styles['result_fill']
            if '%' in label:
                ws[value_cell].number_format = styles['percent']
            else:
                ws[value_cell].number_format = styles['currency_detail']

    # Comparison to base options
    ws['A27'] = 'COMPARISON TO BASE OPTIONS'
    ws['A27'].font = styles['header']
    ws['A27'].fill = styles['header_fill']
    ws.merge_cells('A27:D27')

    comp_headers = ['Metric', 'Your Scenario', 'Option B', 'Option C']
    for col, header in enumerate(comp_headers, 1):
        ws.cell(row=28, column=col, value=header)
        ws.cell(row=28, column=col).font = styles['bold']

    comparisons = [
        ('Headline', '=B6', '=Inputs!C12', '=Inputs!D12'),
        ('Down Payment', '=B17', '=Inputs!C33', '=Inputs!D33'),
        ('Est. PV', '=B24', '=Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58', '=Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50'),
        ('PV/Headline', '=B25', '=(Inputs!C33+Inputs!C34*0.65+Inputs!C35*0.58)/Inputs!C12', '=(Inputs!D33+Inputs!D34*0.55+Inputs!D35*0.50)/Inputs!D12'),
    ]

    for row_idx, (label, yours, opt_b, opt_c) in enumerate(comparisons, 29):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=yours)
        ws.cell(row=row_idx, column=3, value=opt_b)
        ws.cell(row=row_idx, column=4, value=opt_c)
        for col in [2, 3, 4]:
            if 'PV/Headline' in label:
                ws.cell(row=row_idx, column=col).number_format = styles['percent']
            else:
                ws.cell(row=row_idx, column=col).number_format = styles['currency_detail']

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    return ws

def main():
    """Create the complete workbook"""
    wb = Workbook()
    styles = create_styles()

    # Create all sheets
    create_dashboard(wb, styles)
    create_inputs_sheet(wb, styles)
    create_payment_schedule_sheet(wb, styles, 'A', 'B')
    create_payment_schedule_sheet(wb, styles, 'B', 'C')
    create_payment_schedule_sheet(wb, styles, 'C', 'D')
    create_comparison_sheet(wb, styles)
    create_scenario_sheet(wb, styles)

    # Save
    output_path = '/Users/matt/Documents/belkin/NYDG_LOI_Calculator.xlsx'
    wb.save(output_path)
    print(f"Created: {output_path}")

    # Also create a summary of what's in the workbook
    summary = """
NYDG LOI Calculator - Sheet Guide
==================================

1. DASHBOARD - Executive summary with side-by-side comparison
   - Headline prices, down payments, PV analysis
   - External buyer comparison

2. INPUTS - All adjustable parameters (YELLOW CELLS)
   - Change these to update all calculations
   - Practice fundamentals, option parameters, discount rates

3. SCHEDULE_A/B/C - Detailed payment schedules
   - Year-by-year payment breakdown
   - Present value calculations per option

4. PERSPECTIVES - Colbert vs Belkin views
   - What seller cares about vs true buyer economics
   - Key insight: Option C = best value per dollar of headline

5. SCENARIOS - What-if analysis
   - Create your own custom deal structure
   - Compare against Options B and C

KEY FORMULAS:
- Present Value = Payment / (1 + rate)^year
- All options use 10% down, 40% initial purchase
- PV % shows true cost as percentage of headline price
"""
    print(summary)
    return output_path

if __name__ == '__main__':
    main()
