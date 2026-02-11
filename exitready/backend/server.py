from fastapi import FastAPI, APIRouter, File, UploadFile, HTTPException, Form, Request
from fastapi.responses import FileResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ValidationError
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import pandas as pd
import json
import base64
from io import BytesIO
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import stripe
import aiosmtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Stripe configuration
stripe.api_key = os.environ.get('STRIPE_SECRET_KEY')
stripe_webhook_secret = os.environ.get('STRIPE_WEBHOOK_SECRET')

# Email configuration
email_user = os.environ.get('GMAIL_USER')
email_password = os.environ.get('GMAIL_PASSWORD')

async def send_email(subject: str, body: str, to_email: str = None):
    """Send email notification via Gmail SMTP"""
    if not email_user or not email_password:
        logger.warning("Gmail credentials not configured, skipping email notification")
        return
    
    if not to_email:
        to_email = email_user  # Send to self if no recipient specified
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Add body to email
        msg.attach(MIMEText(body, 'plain'))
        
        # Gmail SMTP configuration
        await aiosmtplib.send(
            msg,
            hostname="smtp.gmail.com",
            port=587,
            start_tls=True,
            username=email_user,
            password=email_password,
        )
        
        logger.info(f"Email sent successfully to {to_email}: {subject}")
        
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class StatusCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class StatusCheckCreate(BaseModel):
    client_name: str

# QOE Models
class QOEProject(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    package_type: str  # "basic" or "comprehensive"
    client_name: str
    status: str = "started"  # started, uploaded, mapped, analyzed, completed
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    files_uploaded: Dict[str, Any] = {}
    account_mappings: Dict[str, Any] = {}
    addbacks: List[Dict[str, Any]] = []
    analysis_results: Dict[str, Any] = {}

class QOEProjectCreate(BaseModel):
    package_type: str
    client_name: str

class FileUploadResponse(BaseModel):
    id: str
    filename: str
    file_type: str
    rows_processed: int
    columns: List[str]
    sample_data: List[Dict[str, Any]]

class AccountMapping(BaseModel):
    original_account: str
    mapped_account: str
    account_category: str

class AddBack(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    amount: float
    category: str  # "owner_compensation", "personal_expenses", "one_time", "non_recurring"
    explanation: str
    frequency: str = "annual"

class QOEAnalysis(BaseModel):
    revenue_analysis: Dict[str, Any]
    expense_analysis: Dict[str, Any]
    addback_summary: Dict[str, Any]
    normalized_ebitda: float
    red_flags: List[Dict[str, Any]]
    recommendations: List[str]

# Value Calculator Models
class Question(BaseModel):
    id: int
    step: str
    excel_row: int
    question: str
    input_type: str
    weight: float
    scored: bool
    scoring_formula: str
    validation_source: str

class Option(BaseModel):
    question_id: int
    option_text: str

class TaxParam(BaseModel):
    param: str
    cell: str
    value: float

class NAICSModifier(BaseModel):
    industry_naics_codes: str
    naics_code: str
    industry_name: str

class Multiple(BaseModel):
    band: str
    multiple_min: float
    multiple_max: float

class ScoringRequest(BaseModel):
    run_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    answers: List[Dict[str, Any]]
    financial_inputs: Dict[str, float] = {}

class ScoringResponse(BaseModel):
    run_id: str
    sellability_score: float
    sellability_grade: str
    enterprise_value: float
    net_proceeds: float
    wealth_gap: float
    per_question: List[Dict[str, Any]]
    valuation_details: Dict[str, Any]
    assumptions_hash: str

# Consultation Models
class ConsultationBooking(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_info: Dict[str, Any]
    hours: float  # Number of hours (0.5, 1, 2, etc.)
    amount: int  # Amount in cents (calculated from hours * $250)
    stripe_payment_intent_id: str
    status: str = "pending"  # pending, processing, completed, failed
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class ConsultationBookingCreate(BaseModel):
    customer_info: Dict[str, Any]
    hours: float

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.dict()
    status_obj = StatusCheck(**status_dict)
    _ = await db.status_checks.insert_one(status_obj.dict())
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find().to_list(1000)
    return [StatusCheck(**status_check) for status_check in status_checks]

# Lead Magnet Models
class LeadMagnetSubmission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    company: str
    lead_magnet: str
    submitted_at: datetime = Field(default_factory=datetime.utcnow)

class LeadMagnetCreate(BaseModel):
    email: str
    company: str
    lead_magnet: str

# User Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    password: str  # In production, this should be hashed
    name: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    last_login: Optional[datetime] = None

class UserCreate(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

# User Endpoints
@api_router.post("/users/register")
async def register_user(user_data: UserCreate):
    """Register a new user"""
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists with this email")
    
    # Create new user
    user_dict = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "password": user_data.password,
        "name": user_data.name,
        "created_at": datetime.utcnow()
    }
    
    # Insert directly into database
    await db.users.insert_one(user_dict)
    
    # Return user without password
    response_user = {
        "id": user_dict["id"],
        "email": user_dict["email"],
        "name": user_dict["name"],
        "created_at": user_dict["created_at"]
    }
    
    return {"message": "User registered successfully", "user": response_user}

@api_router.post("/users/login")
async def login_user(login_data: UserLogin):
    """Login user"""
    # Find user by email only first
    user = await db.users.find_one({"email": login_data.email})
    
    # Debug log
    print(f"Login attempt for email: {login_data.email}")
    print(f"User found: {user is not None}")
    if user:
        print(f"Password in DB: {user.get('password')}")
        print(f"Password provided: {login_data.password}")
        print(f"Password match: {user.get('password') == login_data.password}")
    
    if not user or user.get("password") != login_data.password:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Update last login
    await db.users.update_one(
        {"_id": user["_id"]},
        {"$set": {"last_login": datetime.utcnow()}}
    )
    
    # Return user without password
    return {"message": "Login successful", "user": {"id": str(user["_id"]), "email": user["email"], "name": user["name"], "created_at": user["created_at"]}}

# Valuation Models
class ValuationSubmission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    valuation_type: str  # 'simple' or 'detailed'
    annual_revenue: Optional[float] = None
    ebitda: Optional[float] = None
    estimated_value: Optional[float] = None
    industry: Optional[str] = None
    firm_age: Optional[int] = None
    employee_count: Optional[int] = None
    growth_rate: Optional[float] = None
    submitted_at: datetime = Field(default_factory=datetime.utcnow)

class ValuationCreate(BaseModel):
    company_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    valuation_type: str
    annual_revenue: Optional[float] = None
    ebitda: Optional[float] = None
    estimated_value: Optional[float] = None
    industry: Optional[str] = None
    firm_age: Optional[int] = None
    employee_count: Optional[int] = None
    growth_rate: Optional[float] = None

# Lead Magnet Endpoints
@api_router.post("/lead-magnet/red-flags")
async def submit_red_flags_lead(submission_data: dict):
    """Submit lead magnet form and store lead info"""
    # Handle both lead_magnet and leadMagnet field names
    if "leadMagnet" in submission_data and "lead_magnet" not in submission_data:
        submission_data["lead_magnet"] = submission_data["leadMagnet"]
    
    # Create a LeadMagnetCreate object for validation
    try:
        # Extract only the fields we need for LeadMagnetCreate
        valid_data = {
            "email": submission_data.get("email", ""),
            "company": submission_data.get("company", ""),
            "lead_magnet": submission_data.get("lead_magnet", "")
        }
        submission = LeadMagnetCreate(**valid_data)
    except ValidationError as e:
        raise HTTPException(status_code=422, detail=str(e))
    
    # Validate business email (no Gmail)
    if '@gmail.com' in submission.email.lower():
        raise HTTPException(status_code=400, detail="Please use your business email address")
    
    # Store lead in database
    lead_dict = submission.dict()
    lead_obj = LeadMagnetSubmission(**lead_dict)
    await db.lead_magnets.insert_one(lead_obj.dict())
    
    return {"message": "Success", "id": lead_obj.id}

@api_router.get("/lead-magnet/red-flags-pdf")
async def download_red_flags_pdf():
    """Generate and return the 23 Red Flags PDF"""
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')  # Blue color
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.HexColor('#dc2626')  # Red color
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 11
    normal_style.spaceAfter = 12
    
    # Build PDF content
    story = []
    
    # Title and header
    story.append(Paragraph("THE 23 RED FLAGS", title_style))
    story.append(Paragraph("That Kill Accounting Firm Sales", title_style))
    story.append(Spacer(1, 20))
    
    # Introduction
    intro_text = """
    This comprehensive assessment reveals the hidden issues that derail accounting firm exits before they happen. 
    These red flags come from real experience - they're the actual problems that emerged during a 7-figure 
    accounting firm sale process.
    """
    story.append(Paragraph(intro_text, normal_style))
    story.append(Spacer(1, 20))
    
    # Financial Presentation Red Flags
    story.append(Paragraph("FINANCIAL PRESENTATION RED FLAGS", heading_style))
    
    financial_flags = [
        ("1. Inconsistent Monthly Financial Reporting", "Missing months, late closes, or different formats each period. Buyers need to see consistent, professional reporting."),
        ("2. Revenue Recognition Inconsistencies", "Mixing cash and accrual methods, or inconsistent treatment of retainers, work-in-progress, and seasonal adjustments."),
        ("3. Personal Expenses Mixed with Business", "Owner's personal costs running through the business without clear documentation of add-backs."),
        ("4. Unexplained Revenue Fluctuations", "Large month-to-month variations without clear seasonal patterns or business explanations."),
        ("5. Missing Supporting Documentation", "Bank reconciliations, trial balances, or source documents that can't be produced quickly during due diligence.")
    ]
    
    for flag_title, flag_desc in financial_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 15))
    
    # Client Concentration Red Flags
    story.append(Paragraph("CLIENT CONCENTRATION & RISK RED FLAGS", heading_style))
    
    client_flags = [
        ("6. Over-Concentration in Single Client", "Any client representing more than 15-20% of revenue creates buyer concern about sustainability."),
        ("7. Related Party Revenue", "Significant revenue from entities owned by the seller or family members without arm's length pricing."),
        ("8. Industry Over-Concentration", "More than 40% of revenue from a single industry creates sector risk concerns."),
        ("9. Aging Accounts Receivable", "Large amounts over 90 days or clients consistently paying late without collection efforts."),
        ("10. Client Retention Issues", "High client turnover, recent large client losses, or clients expressing dissatisfaction.")
    ]
    
    for flag_title, flag_desc in client_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(PageBreak())
    
    # Operational Dependency Red Flags
    story.append(Paragraph("OPERATIONAL DEPENDENCY RED FLAGS", heading_style))
    
    operational_flags = [
        ("11. Owner-Dependent Relationships", "Key clients who work exclusively with the owner and have no relationship with other staff."),
        ("12. Undocumented Processes", "Critical procedures that exist only in the owner's head with no written systems or training materials."),
        ("13. Single Points of Failure", "Key employees (other than owner) whose departure would significantly impact operations."),
        ("14. Manual, Non-Scalable Processes", "Heavy reliance on manual procedures that don't scale or integrate with modern systems."),
        ("15. No Cross-Training", "Staff who can only perform specific tasks with no backup coverage for key functions.")
    ]
    
    for flag_title, flag_desc in operational_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 15))
    
    # Technology & Systems Red Flags
    story.append(Paragraph("TECHNOLOGY & SYSTEMS RED FLAGS", heading_style))
    
    tech_flags = [
        ("16. Outdated Technology Stack", "Legacy software that's no longer supported or incompatible with modern accounting standards."),
        ("17. Data Security Vulnerabilities", "Lack of proper cybersecurity measures, outdated security protocols, or recent security incidents."),
        ("18. Poor Integration Between Systems", "Manual data entry between different software platforms creating inefficiency and error risk.")
    ]
    
    for flag_title, flag_desc in tech_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 15))
    
    # Legal & Compliance Red Flags
    story.append(Paragraph("LEGAL & COMPLIANCE RED FLAGS", heading_style))
    
    legal_flags = [
        ("19. Outstanding Legal Issues", "Pending litigation, regulatory investigations, or unresolved professional liability claims."),
        ("20. Compliance Gaps", "Missing professional licenses, lapsed insurance coverage, or failure to meet continuing education requirements."),
        ("21. Employment Law Issues", "Misclassified employees, unpaid overtime claims, or HR policy gaps that create liability.")
    ]
    
    for flag_title, flag_desc in legal_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 15))
    
    # Strategic & Positioning Red Flags
    story.append(Paragraph("STRATEGIC & POSITIONING RED FLAGS", heading_style))
    
    strategic_flags = [
        ("22. Declining or Flat Revenue Trends", "Multi-year patterns showing no growth or consistent decline without clear turnaround strategy."),
        ("23. Poor Service Mix", "Over-reliance on low-margin compliance work without higher-value advisory services.")
    ]
    
    for flag_title, flag_desc in strategic_flags:
        story.append(Paragraph(f"<b>{flag_title}</b>", normal_style))
        story.append(Paragraph(flag_desc, normal_style))
        story.append(Spacer(1, 8))
    
    story.append(PageBreak())
    
    # How to Use This List
    story.append(Paragraph("HOW TO USE THIS LIST", heading_style))
    
    usage_steps = [
        ("Step 1: Honest Assessment", "Go through each red flag and honestly evaluate whether it applies to your firm."),
        ("Step 2: Prioritize by Impact", "Some red flags are deal-killers, others just reduce valuation. Focus on the most critical issues first."),
        ("Step 3: Create Action Plan", "Develop specific steps to address each identified red flag with realistic timelines."),
        ("Step 4: Document Progress", "Track your improvements so you can demonstrate positive changes to potential buyers.")
    ]
    
    for step_title, step_desc in usage_steps:
        story.append(Paragraph(f"<b>{step_title}</b>", normal_style))
        story.append(Paragraph(step_desc, normal_style))
        story.append(Spacer(1, 12))
    
    story.append(Spacer(1, 20))
    
    # Conclusion
    conclusion_text = """
    <b>Remember:</b> The goal isn't perfection - it's preparation. Buyers expect some issues, but they want to see 
    that you've identified and addressed them proactively rather than having surprises emerge during due diligence.
    """
    story.append(Paragraph(conclusion_text, normal_style))
    
    story.append(Spacer(1, 30))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280')
    )
    
    story.append(Paragraph("Exit Ready | The Only Exit Prep System Built by Someone Who Actually Sold Their Accounting Firm", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Return PDF response
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=23-red-flags-assessment.pdf"}
    )

# Valuation Endpoints
@api_router.post("/valuation/submit")
async def submit_valuation(submission: ValuationCreate):
    """Submit valuation data for lead tracking"""
    # Store valuation in database
    valuation_dict = submission.dict()
    valuation_obj = ValuationSubmission(**valuation_dict)
    await db.valuations.insert_one(valuation_obj.dict())
    
    return {"message": "Valuation submitted successfully", "id": valuation_obj.id}

@api_router.get("/valuation/sample-simple-pdf")
async def get_simple_valuation_sample():
    """Get sample simple valuation PDF with proper branding and formatting"""
    # Create a properly formatted simple valuation PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#1e40af')
    )
    
    story = []
    
    # Header with Exit Ready branding
    story.append(Paragraph("EXIT READY", title_style))
    story.append(Paragraph("Quick Firm Valuation Report", title_style))
    story.append(Spacer(1, 30))
    
    # Introduction with proper formatting
    story.append(Paragraph("Introduction", heading_style))
    intro_text = """
    This sample represents the Quick Valuation Report generated by Exit Ready's simple calculator. 
    The simple valuation provides a reliable estimate using industry-standard revenue multiples, 
    perfect for firm owners who want a quick assessment of their practice's value.
    """
    story.append(Paragraph(intro_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # How it was calculated with better formatting
    story.append(Paragraph("How This Valuation Was Calculated", heading_style))
    calc_text = """
    Our simple valuation method uses the industry-standard revenue multiple approach:
    """
    story.append(Paragraph(calc_text, styles['Normal']))
    story.append(Spacer(1, 10))
    
    # Formatted calculation steps
    calc_steps = [
        "1. Base Multiple: We start with a 1.0x revenue multiple (standard for accounting firms)",
        "2. Quality Adjustments: We adjust this multiple based on your firm's characteristics",
        "3. Final Calculation: Adjusted multiple × Annual Revenue = Estimated Value"
    ]
    
    for step in calc_steps:
        story.append(Paragraph(step, styles['Normal']))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 15))
    story.append(Paragraph("This approach is used by business brokers and provides a reliable market-based estimate.", styles['Normal']))
    story.append(Spacer(1, 25))
    
    # Company details section with professional formatting
    story.append(Paragraph("Valuation Details", heading_style))
    
    # Sample data table with better formatting
    sample_data = [
        ['Company Name:', 'Sample Accounting Firm LLC'],
        ['Valuation Type:', 'Quick Valuation'],
        ['Annual Revenue:', '$1,250,000'],
        ['Industry:', 'Professional Services'],
        ['Base Multiple:', '1.0x'],
        ['Quality Adjustments:', '+0.5x'],
        ['Final Multiple:', '1.5x'],
        ['Estimated Value Range:', '$1,875,000'],
        ['Generated By:', 'Exit Ready - Professional Valuation System'],
        ['Report Date:', datetime.utcnow().strftime('%B %d, %Y')]
    ]
    
    table = Table(sample_data, colWidths=[2.5*inch, 3.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1e293b')),
        ('BACKGROUND', (0, 7), (1, 7), colors.HexColor('#dcfce7')),  # Highlight result
        ('FONTNAME', (0, 7), (1, 7), 'Helvetica-Bold'),  # Bold result
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 25))
    
    # Next steps section
    story.append(Paragraph("Next Steps", heading_style))
    next_steps_text = """
    For a more detailed analysis, consider our Comprehensive Valuation Report which includes:
    
    • Multiple valuation methodologies (Revenue, EBITDA, Asset-based)
    • 15-page detailed analysis with industry benchmarking
    • Quality score assessment and improvement recommendations
    • Risk factor analysis and mitigation strategies
    • Value enhancement opportunities and strategic positioning
    
    Contact Exit Ready to upgrade to our comprehensive analysis or discuss your exit preparation strategy.
    """
    story.append(Paragraph(next_steps_text, styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Professional footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280')
    )
    
    story.append(Paragraph("Exit Ready | The Only Exit Prep System Built by Someone Who Actually Sold Their Accounting Firm", footer_style))
    story.append(Paragraph("Professional Valuation Services | www.exitready.com", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sample-simple-valuation.pdf"}
    )

@api_router.get("/valuation/sample-detailed-pdf")
async def get_detailed_valuation_sample():
    """Get sample detailed valuation PDF - FULL 15 PAGES"""
    # Create a comprehensive 15-page sample PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#1e40af')
    )
    
    story = []
    
    # PAGE 1: Cover Page with Exit Ready Branding
    # (SimpleDocTemplate doesn't have setTitle method, so we'll add title in content)
    
    # Header with Exit Ready branding
    story.append(Paragraph("EXIT READY", title_style))
    story.append(Paragraph("Comprehensive Accounting Firm Valuation Report", title_style))
    story.append(Spacer(1, 40))
    
    # Company info
    story.append(Paragraph("Sample Accounting Firm LLC", heading_style))
    story.append(Paragraph("Professional Valuation Analysis", styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Report details
    report_details = [
        f"Report Date: {datetime.utcnow().strftime('%B %d, %Y')}",
        "Report Type: Comprehensive Analysis (15 Pages)",
        "Valuation Methods: Revenue Multiple, EBITDA Multiple, Asset-Based",
        "Prepared by: Exit Ready Professional Services"
    ]
    
    for detail in report_details:
        story.append(Paragraph(detail, styles['Normal']))
    
    story.append(Spacer(1, 40))
    
    # Executive summary box
    story.append(Paragraph("Executive Summary", heading_style))
    exec_summary = """
    This comprehensive 15-page valuation report provides an institutional-grade analysis of Sample 
    Accounting Firm LLC. Using multiple valuation methodologies and industry benchmarking, we estimate 
    the firm's value at $2,250,000 with a range of $1,875,000 to $2,625,000. This analysis includes 
    detailed recommendations for value enhancement and strategic positioning for a successful exit.
    """
    story.append(Paragraph(exec_summary, styles['Normal']))
    
    # PAGE 2: Table of Contents
    story.append(PageBreak())
    story.append(Paragraph("TABLE OF CONTENTS", title_style))
    story.append(Spacer(1, 20))
    
    toc_items = [
        ("1. Executive Summary", "3"),
        ("2. Firm Overview", "4"),
        ("3. Financial Analysis", "5"),
        ("4. Valuation Methodologies", "6"),
        ("5. Revenue Multiple Method", "7"),
        ("6. EBITDA Multiple Method", "8"),
        ("7. Asset-Based Method", "9"),
        ("8. Quality Score Analysis", "10"),
        ("9. Industry Benchmarking", "11"),
        ("10. Risk Factor Assessment", "12"),
        ("11. Value Enhancement Recommendations", "13"),
        ("12. Market Positioning Strategy", "14"),
        ("13. Exit Preparation Timeline", "15"),
        ("14. Next Steps", "16"),
        ("15. Appendices", "17")
    ]
    
    for item, page in toc_items:
        story.append(Paragraph(f"{item} ....................................... {page}", styles['Normal']))
    
    # PAGE 3: Executive Summary (Detailed)
    story.append(PageBreak())
    story.append(Paragraph("1. EXECUTIVE SUMMARY", heading_style))
    story.append(Spacer(1, 15))
    
    exec_content = """
    Sample Accounting Firm LLC represents a well-established professional services practice with strong 
    fundamentals and significant value enhancement potential. Our comprehensive analysis reveals:
    
    KEY FINDINGS:
    • Estimated Value: $2,250,000 (range: $1,875,000 - $2,625,000)
    • Revenue Multiple: 1.8x annual revenue
    • EBITDA Multiple: 4.1x normalized EBITDA  
    • Quality Score: 78/100 (above industry average)
    • Growth Potential: High with recommended improvements
    
    VALUATION SUMMARY:
    The firm demonstrates strong operational performance with annual revenue of $1,250,000 and normalized 
    EBITDA of $550,000, representing a healthy 44% margin. Client diversification is excellent with no 
    single client exceeding 12% of revenue. Technology adoption and process documentation present the 
    greatest opportunities for value enhancement.
    
    STRATEGIC RECOMMENDATIONS:
    Implementation of our recommended improvements could increase firm value by 25-35% within 18 months, 
    positioning the practice for a premium exit to strategic or financial buyers.
    """
    story.append(Paragraph(exec_content, styles['Normal']))
    
    # PAGE 4: Firm Overview
    story.append(PageBreak())
    story.append(Paragraph("2. FIRM OVERVIEW", heading_style))
    story.append(Spacer(1, 15))
    
    firm_overview = """
    COMPANY PROFILE:
    Sample Accounting Firm LLC is a full-service accounting practice established in 2010, serving a 
    diverse client base in the metropolitan area. The firm specializes in tax preparation, bookkeeping, 
    and advisory services for small to medium-sized businesses.
    
    OPERATIONAL METRICS:
    • Annual Revenue: $1,250,000
    • Client Base: 156 active clients
    • Staff Count: 8 employees (including owner)
    • Office Location: Professional building, 3,200 sq ft
    • Service Mix: 60% tax, 25% bookkeeping, 15% advisory
    
    COMPETITIVE POSITION:
    The firm maintains a strong market position through:
    • Long-term client relationships (average tenure 4.2 years)
    • Specialized industry expertise in professional services
    • Modern technology infrastructure
    • Experienced professional staff
    
    OWNERSHIP STRUCTURE:
    Single owner-operator with strong management team. Succession planning discussions initiated, 
    creating optimal timing for strategic exit.
    """
    story.append(Paragraph(firm_overview, styles['Normal']))
    
    # PAGE 5: Financial Analysis
    story.append(PageBreak())
    story.append(Paragraph("3. FINANCIAL ANALYSIS", heading_style))
    story.append(Spacer(1, 15))
    
    financial_analysis = """
    REVENUE ANALYSIS:
    Three-year revenue trend shows consistent growth:
    • 2022: $1,125,000 (baseline year)
    • 2023: $1,188,000 (+5.6% growth)
    • 2024: $1,250,000 (+5.2% growth)
    
    PROFITABILITY METRICS:
    • Gross Profit Margin: 85% (industry average: 75-80%)
    • EBITDA Margin: 44% (industry average: 25-35%)
    • Net Profit Margin: 38% (exceptional performance)
    
    NORMALIZED EBITDA CALCULATION:
    • Reported EBITDA: $425,000
    • Owner compensation adjustment: +$75,000
    • Personal expense add-backs: +$35,000
    • One-time expense adjustments: +$15,000
    • Normalized EBITDA: $550,000
    
    WORKING CAPITAL:
    • Accounts Receivable: $95,000 (30-day average)
    • Accounts Payable: $25,000
    • Net Working Capital: $85,000 (well-managed)
    
    FINANCIAL STRENGTHS:
    • Strong cash flow generation
    • Minimal debt obligations
    • Excellent collection practices
    • Conservative financial management
    """
    story.append(Paragraph(financial_analysis, styles['Normal']))
    
    # PAGE 6: Valuation Methodologies
    story.append(PageBreak())
    story.append(Paragraph("4. VALUATION METHODOLOGIES", heading_style))
    story.append(Spacer(1, 15))
    
    valuation_methods = """
    METHODOLOGY OVERVIEW:
    Professional business valuation requires multiple approaches to establish credible value ranges. 
    We employed three industry-standard methodologies:
    
    1. REVENUE MULTIPLE METHOD:
    Market-based approach using comparable transaction data from accounting firm sales. Multiples 
    adjusted for firm quality, size, and risk factors.
    
    2. EBITDA MULTIPLE METHOD:
    Earnings-based approach focusing on cash flow generation capacity. Considers normalized EBITDA 
    and applies industry-appropriate multiples based on firm characteristics.
    
    3. ASSET-BASED METHOD:
    Foundation approach combining tangible and intangible asset values. Particularly relevant for 
    service businesses where client relationships represent significant value.
    
    WEIGHTING METHODOLOGY:
    • Revenue Multiple: 40% weight (primary method for accounting firms)
    • EBITDA Multiple: 45% weight (strong earnings justify emphasis)
    • Asset-Based: 15% weight (provides floor value)
    
    FINAL VALUATION CALCULATION:
    Weighted average approach provides most reliable estimate while considering all value perspectives.
    """
    story.append(Paragraph(valuation_methods, styles['Normal']))
    
    # PAGE 7: Revenue Multiple Method
    story.append(PageBreak())
    story.append(Paragraph("5. REVENUE MULTIPLE METHOD", heading_style))
    story.append(Spacer(1, 15))
    
    revenue_method = """
    MARKET DATA ANALYSIS:
    Recent accounting firm transactions show revenue multiples ranging from 0.8x to 2.5x, with 
    quality firms achieving premium valuations.
    
    BASE MULTIPLE DETERMINATION:
    • Industry baseline: 1.0x revenue
    • Size adjustment: +0.2x (revenue >$1M)
    • Quality adjustment: +0.4x (based on quality score)
    • Market position: +0.2x (strong local presence)
    • Final multiple: 1.8x
    
    QUALITY FACTORS SUPPORTING PREMIUM:
    • Client retention rate: 89% (above average)
    • Revenue diversification: Well-balanced
    • Technology adoption: Modern systems
    • Staff leverage: Appropriate delegation
    • Process documentation: Good procedures
    
    CALCULATION:
    Annual Revenue: $1,250,000
    Applied Multiple: 1.8x
    Revenue Method Value: $2,250,000
    
    MARKET VALIDATION:
    This multiple aligns with recent transactions of similar-sized firms with comparable quality metrics. 
    Premium positioning reflects strong operational fundamentals and growth potential.
    """
    story.append(Paragraph(revenue_method, styles['Normal']))
    
    # PAGE 8: EBITDA Multiple Method
    story.append(PageBreak())
    story.append(Paragraph("6. EBITDA MULTIPLE METHOD", heading_style))
    story.append(Spacer(1, 15))
    
    ebitda_method = """
    EBITDA MULTIPLE ANALYSIS:
    Earnings-based valuation focuses on cash flow generation capacity, providing insight into 
    the firm's ability to service debt and generate returns for investors.
    
    NORMALIZED EBITDA: $550,000
    (See detailed calculation in Financial Analysis section)
    
    MULTIPLE DETERMINATION:
    • Base multiple (by size): 3.5x
    • Quality premium: +0.4x
    • Growth potential: +0.2x
    • Risk adjustment: +0.0x
    • Final EBITDA multiple: 4.1x
    
    SUPPORTING FACTORS:
    • Consistent EBITDA growth (3-year average +8%)
    • High-quality earnings (minimal one-time items)
    • Predictable revenue streams
    • Strong margin stability
    • Minimal capital requirements
    
    CALCULATION:
    Normalized EBITDA: $550,000
    Applied Multiple: 4.1x
    EBITDA Method Value: $2,255,000
    
    INDUSTRY COMPARISON:
    This multiple reflects above-average performance relative to industry benchmarks. Strong margins 
    and earnings quality justify premium valuation within the 3.5x-5.0x range typical for accounting firms.
    """
    story.append(Paragraph(ebitda_method, styles['Normal']))
    
    # Continue with remaining pages...
    # PAGE 9: Asset-Based Method
    story.append(PageBreak())
    story.append(Paragraph("7. ASSET-BASED METHOD", heading_style))
    story.append(Spacer(1, 15))
    
    asset_method = """
    ASSET VALUATION COMPONENTS:
    
    TANGIBLE ASSETS:
    • Office equipment and furniture: $45,000
    • Computer hardware and software: $25,000
    • Cash and receivables: $95,000
    • Total tangible assets: $165,000
    
    INTANGIBLE ASSETS:
    • Client relationships: $1,400,000
    • Brand value and reputation: $200,000
    • Intellectual property: $50,000
    • Total intangible assets: $1,650,000
    
    CLIENT RELATIONSHIP VALUATION:
    Valued using excess earnings method:
    • Annual client revenue: $1,250,000
    • Client retention rate: 89%
    • Average client tenure: 4.2 years
    • Calculated value: $1,400,000
    
    TOTAL ASSET-BASED VALUE: $1,815,000
    
    This method provides a conservative floor value, as it may not fully capture synergistic value 
    available to strategic buyers.
    """
    story.append(Paragraph(asset_method, styles['Normal']))
    
    # Continue building all 15 pages...
    # I'll add the remaining pages to complete the full 15-page report
    
    # PAGE 10: Quality Score Analysis
    story.append(PageBreak())
    story.append(Paragraph("8. QUALITY SCORE ANALYSIS", heading_style))
    story.append(Spacer(1, 15))
    
    quality_analysis = """
    COMPREHENSIVE QUALITY ASSESSMENT: 78/100
    
    The quality score directly impacts valuation multiples and buyer interest. Our assessment covers:
    
    CLIENT METRICS (25 points): 20/25
    • Retention rate: 89% (target: 90%+)
    • Average tenure: 4.2 years (excellent)
    • Concentration risk: Low (top client 12%)
    
    OPERATIONAL EFFICIENCY (20 points): 16/20
    • Technology adoption: Modern cloud systems
    • Process documentation: Good, needs enhancement
    • Staff leverage: Appropriate delegation
    
    FINANCIAL PERFORMANCE (25 points): 22/25
    • Profit margins: Exceptional (44% EBITDA)
    • Revenue growth: Steady 5%+ annually
    • Cash flow: Strong and predictable
    
    MARKET POSITION (15 points): 12/15
    • Service diversification: Good mix
    • Competitive advantages: Strong reputation
    • Geographic presence: Well-established
    
    MANAGEMENT DEPTH (15 points): 8/15
    • Owner dependency: Moderate concern
    • Succession planning: In development
    • Key employee retention: Good
    
    IMPROVEMENT OPPORTUNITIES:
    Focus on reducing owner dependency and enhancing process documentation could increase 
    score to 85+, adding $200K-$400K to firm value.
    """
    story.append(Paragraph(quality_analysis, styles['Normal']))
    
    # PAGE 11: Industry Benchmarking
    story.append(PageBreak())
    story.append(Paragraph("9. INDUSTRY BENCHMARKING", heading_style))
    story.append(Spacer(1, 15))
    
    benchmarking = """
    PERFORMANCE vs. INDUSTRY BENCHMARKS:
    
    FINANCIAL METRICS:
    Your Firm vs. Industry Average:
    • Revenue Multiple: 1.8x vs. 1.2x (50% premium)
    • EBITDA Margin: 44% vs. 28% (+57% above average)
    • Client Retention: 89% vs. 82% (+9% above average)
    • Revenue per Client: $8,013 vs. $5,200 (+54% premium)
    
    OPERATIONAL METRICS:
    • Revenue per Employee: $156,250 vs. $125,000 (+25%)
    • Technology Score: 8/10 vs. 6/10 (above average)
    • Service Diversification: Good vs. Average
    
    MARKET POSITION:
    Your firm ranks in the top 25% of accounting practices for:
    • Profitability metrics
    • Client relationship quality
    • Operational efficiency
    • Technology adoption
    
    AREAS FOR IMPROVEMENT:
    • Management depth (below average)
    • Process documentation (average)
    • Strategic planning (developing)
    
    VALUE IMPLICATIONS:
    Above-benchmark performance in key metrics supports premium valuation multiples and 
    positions the firm attractively for strategic buyers.
    """
    story.append(Paragraph(benchmarking, styles['Normal']))
    
    # PAGE 12: Risk Factor Assessment
    story.append(PageBreak())
    story.append(Paragraph("10. RISK FACTOR ASSESSMENT", heading_style))
    story.append(Spacer(1, 15))
    
    risk_assessment = """
    IDENTIFIED RISK FACTORS:
    
    MODERATE RISKS:
    
    1. Owner Dependency
    • Risk Level: Moderate
    • Impact: May reduce buyer universe
    • Mitigation: Develop management team, document relationships
    
    2. Technology Transition
    • Risk Level: Low-Moderate  
    • Impact: Potential disruption during integration
    • Mitigation: Strong existing systems reduce transition risk
    
    LOW RISKS:
    
    3. Client Concentration
    • Risk Level: Low
    • Impact: Minimal - well diversified
    • Mitigation: Continue relationship management
    
    4. Market Competition
    • Risk Level: Low
    • Impact: Strong competitive position
    • Mitigation: Maintain service quality and innovation
    
    RISK MITIGATION RECOMMENDATIONS:
    • Implement formal succession planning
    • Cross-train staff on key client relationships
    • Document all critical processes
    • Develop management incentive programs
    
    OVERALL RISK PROFILE: LOW-MODERATE
    Well-managed firm with minimal structural risks. Identified risks are addressable 
    through strategic planning and implementation.
    """
    story.append(Paragraph(risk_assessment, styles['Normal']))
    
    # PAGE 13: Value Enhancement Recommendations
    story.append(PageBreak())
    story.append(Paragraph("11. VALUE ENHANCEMENT RECOMMENDATIONS", heading_style))
    story.append(Spacer(1, 15))
    
    recommendations = """
    STRATEGIC IMPROVEMENT PLAN:
    
    IMMEDIATE ACTIONS (0-6 months):
    
    1. Management Development Program
    • Cross-train senior staff on client relationships
    • Implement formal management structure
    • Potential value impact: +$150K-$250K
    
    2. Process Documentation Enhancement
    • Document all key procedures
    • Create client service manuals
    • Standardize workflows
    • Potential value impact: +$100K-$200K
    
    MEDIUM-TERM INITIATIVES (6-18 months):
    
    3. Service Line Expansion
    • Develop advisory capabilities
    • Add specialized services
    • Potential value impact: +$200K-$400K
    
    4. Technology Integration
    • Implement advanced analytics
    • Enhance client portal functionality
    • Potential value impact: +$100K-$150K
    
    LONG-TERM POSITIONING (12-24 months):
    
    5. Strategic Partnerships
    • Develop referral networks
    • Consider merger opportunities
    • Potential value impact: +$300K-$500K
    
    TOTAL POTENTIAL VALUE ENHANCEMENT: $850K-$1.5M
    Implementation of these recommendations could increase firm value by 25-35%.
    """
    story.append(Paragraph(recommendations, styles['Normal']))
    
    # PAGE 14: Market Positioning Strategy
    story.append(PageBreak())
    story.append(Paragraph("12. MARKET POSITIONING STRATEGY", heading_style))
    story.append(Spacer(1, 15))
    
    positioning = """
    OPTIMAL BUYER POSITIONING:
    
    STRATEGIC BUYER ADVANTAGES:
    • Geographic expansion opportunity
    • Client base diversification
    • Technology platform integration
    • Management team retention
    
    FINANCIAL BUYER APPEAL:
    • Strong cash flow generation
    • Minimal capital requirements
    • Scalable operations
    • Professional management potential
    
    MARKETING STRATEGY:
    
    Phase 1: Preparation (Months 1-6)
    • Complete value enhancement initiatives
    • Prepare due diligence materials
    • Organize financial documentation
    
    Phase 2: Market Approach (Months 7-12)
    • Identify strategic buyers
    • Prepare confidential marketing materials
    • Initiate buyer conversations
    
    Phase 3: Transaction Execution (Months 13-18)
    • Formal sale process
    • Due diligence management
    • Negotiation and closing
    
    TARGET BUYER PROFILE:
    • Regional accounting firms seeking growth
    • Private equity-backed platforms
    • Entrepreneurs seeking established practices
    
    EXPECTED MARKET RECEPTION:
    Strong fundamentals and growth potential should generate significant buyer interest.
    """
    story.append(Paragraph(positioning, styles['Normal']))
    
    # PAGE 15: Exit Preparation Timeline & Next Steps
    story.append(PageBreak())
    story.append(Paragraph("13. EXIT PREPARATION TIMELINE", heading_style))
    story.append(Spacer(1, 15))
    
    timeline = """
    18-MONTH STRATEGIC EXIT PLAN:
    
    MONTHS 1-6: FOUNDATION BUILDING
    □ Complete management development program
    □ Enhance process documentation
    □ Implement technology improvements
    □ Strengthen financial reporting
    
    MONTHS 7-12: VALUE OPTIMIZATION  
    □ Execute service line expansion
    □ Develop strategic partnerships
    □ Complete succession planning
    □ Prepare marketing materials
    
    MONTHS 13-18: MARKET EXECUTION
    □ Launch confidential marketing process
    □ Conduct buyer meetings
    □ Manage due diligence process
    □ Negotiate and close transaction
    
    CRITICAL SUCCESS FACTORS:
    • Maintain business performance during process
    • Preserve confidentiality
    • Professional advisor engagement
    • Timeline flexibility for optimal outcomes
    
    NEXT STEPS:
    
    1. Review and prioritize recommendations
    2. Develop detailed implementation plan
    3. Engage professional advisors
    4. Begin value enhancement initiatives
    5. Schedule quarterly progress reviews
    
    Contact Exit Ready for ongoing support and guidance throughout your exit preparation journey.
    """
    story.append(Paragraph(timeline, styles['Normal']))
    
    # Build the PDF
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sample-comprehensive-valuation-15pages.pdf"}
    )

# User Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    company: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

# This section has been removed to avoid duplicate route definitions

# Admin Dashboard Models  
class AdminAuth(BaseModel):
    password: str

class AdminStats(BaseModel):
    total_lead_magnets: int
    total_qoe_projects: int
    total_valuations: int
    total_users: int
    recent_submissions: int

# Admin Dashboard Endpoints
@api_router.post("/admin/auth")
async def admin_login(auth: AdminAuth):
    """Authenticate admin access"""
    if auth.password != "exitready2024":
        raise HTTPException(status_code=401, detail="Invalid password")
    return {"message": "Authentication successful", "access": True}

@api_router.get("/admin/stats")
async def get_admin_stats():
    """Get overall statistics for admin dashboard"""
    # Count documents in each collection
    lead_magnets_count = await db.lead_magnets.count_documents({})
    qoe_projects_count = await db.qoe_projects.count_documents({})
    valuations_count = await db.valuations.count_documents({})
    users_count = await db.users.count_documents({}) if hasattr(db, 'users') else 0
    
    # Count recent submissions (last 30 days)
    from datetime import datetime, timedelta
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    recent_count = await db.lead_magnets.count_documents({
        "submitted_at": {"$gte": thirty_days_ago}
    }) + await db.valuations.count_documents({
        "submitted_at": {"$gte": thirty_days_ago}
    })
    
    return AdminStats(
        total_lead_magnets=lead_magnets_count,
        total_qoe_projects=qoe_projects_count,
        total_valuations=valuations_count,
        total_users=users_count,
        recent_submissions=recent_count
    )

@api_router.get("/admin/lead-magnets")
async def get_all_lead_magnets():
    """Get all lead magnet submissions"""
    submissions = await db.lead_magnets.find().sort("submitted_at", -1).to_list(1000)
    
    # Convert ObjectId to string for JSON serialization
    for submission in submissions:
        if '_id' in submission:
            submission['_id'] = str(submission['_id'])
    
    return submissions

@api_router.get("/admin/qoe-projects")
async def get_all_qoe_projects():
    """Get all QOE project submissions"""
    projects = await db.qoe_projects.find().sort("created_at", -1).to_list(1000)
    
    # Convert ObjectId to string for JSON serialization
    for project in projects:
        if '_id' in project:
            project['_id'] = str(project['_id'])
    
    return projects

@api_router.get("/admin/users")
async def get_admin_users():
    """Get all registered users for admin dashboard"""
    try:
        users = await db.users.find().to_list(1000)
        # Convert ObjectId to string and format for frontend
        formatted_users = []
        for user in users:
            formatted_user = {
                "id": user.get("id", str(user["_id"])),
                "name": user.get("name", "N/A"),
                "email": user.get("email", "N/A"),
                "created_at": user.get("created_at"),
                "last_login": user.get("last_login")
            }
            formatted_users.append(formatted_user)
        
        return formatted_users
    except Exception as e:
        # Return empty array if collection doesn't exist or other errors
        return []

@api_router.get("/admin/valuations")
async def get_all_valuations():
    """Get all valuation submissions"""
    valuations = await db.valuations.find().sort("submitted_at", -1).to_list(1000)
    
    # Convert ObjectId to string for JSON serialization
    for valuation in valuations:
        if '_id' in valuation:
            valuation['_id'] = str(valuation['_id'])
    
    return valuations

# Value Calculator Endpoints
@api_router.post("/value-calculator/score", response_model=ScoringResponse)
async def calculate_score(request: ScoringRequest):
    """Calculate sellability score and enterprise value"""
    try:
        # Get all questions and their weights
        questions = await db.value_questions.find().to_list(1000)
        options = await db.value_options.find().to_list(1000)
        tax_params = await db.value_tax_params.find().to_list(1000)
        multiples = await db.value_multiples.find().to_list(1000)
        
        # Create lookup dictionaries
        questions_dict = {q['id']: q for q in questions}
        options_dict = {}
        for opt in options:
            if opt['question_id'] not in options_dict:
                options_dict[opt['question_id']] = []
            options_dict[opt['question_id']].append(opt['option_text'])
        
        tax_params_dict = {tp['param']: tp['value'] for tp in tax_params}
        
        # Calculate per-question scores
        per_question = []
        total_weighted_score = 0
        total_weight = 0
        
        for answer in request.answers:
            question_id = answer.get('question_id')
            value = answer.get('value', '')
            
            if question_id in questions_dict:
                question = questions_dict[question_id]
                
                if question['scored']:
                    # Calculate raw score based on scoring formula
                    raw_score = calculate_question_score(question, value, options_dict.get(question_id, []))
                    weight = question['weight'] if question['weight'] != 'Off' else 0
                    weighted_score = raw_score * weight
                    
                    per_question.append({
                        'question_id': question_id,
                        'question': question['question'],
                        'value': value,
                        'raw_score': raw_score,
                        'weight': weight,
                        'weighted_score': weighted_score
                    })
                    
                    total_weighted_score += weighted_score
                    total_weight += weight
        
        # Calculate final sellability score (0-100)
        sellability_score = (total_weighted_score / total_weight * 20) if total_weight > 0 else 0
        
        # Determine grade
        if sellability_score >= 80:
            grade = "A"
        elif sellability_score >= 70:
            grade = "B" 
        elif sellability_score >= 60:
            grade = "C"
        elif sellability_score >= 50:
            grade = "D"
        else:
            grade = "F"
        
        # Calculate Enterprise Value
        revenue = request.financial_inputs.get('revenue', 0)
        ebitda = request.financial_inputs.get('ebitda', 0)
        
        # Determine multiple based on revenue band
        multiple = 1.0  # default
        for mult in multiples:
            if mult['band'] == '<1M' and revenue < 1000000:
                multiple = (mult['multiple_min'] + mult['multiple_max']) / 2
            elif mult['band'] == '1-10M' and 1000000 <= revenue <= 10000000:
                multiple = (mult['multiple_min'] + mult['multiple_max']) / 2
            elif mult['band'] == '10M+' and revenue > 10000000:
                multiple = (mult['multiple_min'] + mult['multiple_max']) / 2
        
        # Adjust multiple based on sellability score
        score_adjustment = (sellability_score - 50) / 100  # -0.5 to +0.5
        adjusted_multiple = multiple + score_adjustment
        
        enterprise_value = ebitda * adjusted_multiple
        
        # Calculate Net Proceeds (simplified)
        fees_rate = tax_params_dict.get('Estimated Professional Fees & Taxes as %', 0.1)
        net_proceeds = enterprise_value * (1 - fees_rate)
        
        # Calculate Wealth Gap (simplified)
        target_corpus = request.financial_inputs.get('target_corpus', 0)
        wealth_gap = target_corpus - net_proceeds
        
        # Create assumptions hash
        assumptions_hash = str(hash(str(tax_params_dict)))
        
        return ScoringResponse(
            run_id=request.run_id,
            sellability_score=round(sellability_score, 1),
            sellability_grade=grade,
            enterprise_value=round(enterprise_value, 0),
            net_proceeds=round(net_proceeds, 0),
            wealth_gap=round(wealth_gap, 0),
            per_question=per_question,
            valuation_details={
                'revenue': revenue,
                'ebitda': ebitda,
                'multiple': round(adjusted_multiple, 2),
                'fees_rate': fees_rate
            },
            assumptions_hash=assumptions_hash
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scoring calculation failed: {str(e)}")

def calculate_question_score(question, value, options):
    """Calculate score for a single question based on its scoring formula"""
    try:
        # Handle different input types
        if question['input_type'] == 'Dropdown':
            # Convert scoring formulas to Python logic
            formula = question['scoring_formula']
            
            # Simplified scoring based on common patterns
            if 'If(' in formula:
                # Handle Excel IF statements - convert to simple mapping
                if value in ["5+ Years", "Complete Alignment", "Yes"]:
                    return 5
                elif value in ["2 - 4 Years", "Partial Alignement", "A Large Amount"]:
                    return 4  
                elif value in ["1 - 2 Years", "A Medium Amount"]:
                    return 3
                elif value in ["< 12 Months", "A Small Amount"]:
                    return 2
                elif value in ["< 6 Months", "None", "No"]:
                    return 1
                else:
                    return 0
            else:
                # Default scoring
                if value and value != "Select":
                    return 3  # Average score
                return 0
        
        elif question['input_type'] == 'Free':
            # Handle numeric inputs
            try:
                numeric_value = float(value) if value else 0
                # Normalize to 0-5 scale based on typical ranges
                return min(5, max(0, numeric_value * 5))
            except:
                return 0
        
        return 0
        
    except Exception as e:
        print(f"Error calculating score for question {question['id']}: {e}")
        return 0

@api_router.get("/value-calculator/questions")
async def get_questions():
    """Get all questions for the value calculator"""
    questions = await db.value_questions.find().sort("id", 1).to_list(1000)
    options = await db.value_options.find().to_list(1000)
    
    # Group options by question_id
    options_dict = {}
    for opt in options:
        if opt['question_id'] not in options_dict:
            options_dict[opt['question_id']] = []
        options_dict[opt['question_id']].append(opt['option_text'])
    
    # Add options to questions
    for question in questions:
        question['options'] = options_dict.get(question['id'], [])
        if '_id' in question:
            del question['_id']
    
    return questions

@api_router.get("/value-calculator/init")
async def initialize_value_calculator():
    """Initialize the value calculator with seed data"""
    try:
        # Clear existing data
        await db.value_questions.delete_many({})
        await db.value_options.delete_many({})
        await db.value_tax_params.delete_many({})
        await db.value_naics_modifiers.delete_many({})
        await db.value_multiples.delete_many({})
        
        # Initialize with provided seed data
        questions_data = await import_questions_seed()
        options_data = await import_options_seed()
        tax_params_data = await import_tax_params_seed()
        naics_data = await import_naics_seed()
        multiples_data = await import_multiples_seed()
        
        # Insert data
        if questions_data:
            await db.value_questions.insert_many(questions_data)
        if options_data:
            await db.value_options.insert_many(options_data)
        if tax_params_data:
            await db.value_tax_params.insert_many(tax_params_data)
        if naics_data:
            await db.value_naics_modifiers.insert_many(naics_data)
        if multiples_data:
            await db.value_multiples.insert_many(multiples_data)
        
        return {
            "message": "Value calculator initialized successfully",
            "questions": len(questions_data) if questions_data else 0,
            "options": len(options_data) if options_data else 0,
            "tax_params": len(tax_params_data) if tax_params_data else 0,
            "naics": len(naics_data) if naics_data else 0,
            "multiples": len(multiples_data) if multiples_data else 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Initialization failed: {str(e)}")

# Admin endpoints for Value Calculator
@api_router.get("/value-questions/all")
async def get_all_value_questions():
    """Get all value calculator questions for admin"""
    questions = await db.value_questions.find().sort("id", 1).to_list(1000)
    for question in questions:
        if '_id' in question:
            del question['_id']
    return questions

@api_router.get("/value-options/all")
async def get_all_value_options():
    """Get all value calculator options for admin"""
    options = await db.value_options.find().sort("question_id", 1).to_list(1000)
    for option in options:
        if '_id' in option:
            del option['_id']
    return options

@api_router.get("/value-tax-params/all")
async def get_all_tax_params():
    """Get all tax parameters for admin"""
    params = await db.value_tax_params.find().to_list(1000)
    for param in params:
        if '_id' in param:
            del param['_id']
    return params

@api_router.get("/value-naics/all")
async def get_all_naics_modifiers():
    """Get all NAICS modifiers for admin"""
    naics = await db.value_naics_modifiers.find().to_list(1000)
    for n in naics:
        if '_id' in n:
            del n['_id']
    return naics

@api_router.get("/value-multiples/all")
async def get_all_multiples():
    """Get all multiples for admin"""
    multiples = await db.value_multiples.find().to_list(1000)
    for mult in multiples:
        if '_id' in mult:
            del mult['_id']
    return multiples

@api_router.post("/value-questions/update")
async def update_value_questions(questions: List[Dict]):
    """Update value calculator questions"""
    try:
        # Clear existing questions
        await db.value_questions.delete_many({})
        
        # Insert new questions
        if questions:
            await db.value_questions.insert_many(questions)
        
        return {"message": "Questions updated successfully", "count": len(questions)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")

@api_router.post("/value-options/update")
async def update_value_options(options: List[Dict]):
    """Update value calculator options"""
    try:
        # Clear existing options
        await db.value_options.delete_many({})
        
        # Insert new options
        if options:
            await db.value_options.insert_many(options)
        
        return {"message": "Options updated successfully", "count": len(options)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")

@api_router.post("/value-tax-params/update")
async def update_tax_params(params: List[Dict]):
    """Update tax parameters"""
    try:
        # Clear existing params
        await db.value_tax_params.delete_many({})
        
        # Insert new params
        if params:
            await db.value_tax_params.insert_many(params)
        
        return {"message": "Tax parameters updated successfully", "count": len(params)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")

@api_router.post("/value-multiples/update")
async def update_multiples(multiples: List[Dict]):
    """Update multiples"""
    try:
        # Clear existing multiples
        await db.value_multiples.delete_many({})
        
        # Insert new multiples
        if multiples:
            await db.value_multiples.insert_many(multiples)
        
        return {"message": "Multiples updated successfully", "count": len(multiples)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")

async def import_questions_seed():
    """Import questions from seed data"""
    questions = [
        {"id": 1, "step": "", "excel_row": 3, "question": "Company Information", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 2, "step": "", "excel_row": 4, "question": "Company", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 3, "step": "", "excel_row": 5, "question": "Industry", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 4, "step": "", "excel_row": 6, "question": "NAICS Code", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 5, "step": "", "excel_row": 8, "question": "STEP 1 - Exit & Owner Readiness", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 6, "step": "", "excel_row": 9, "question": "#", "input_type": "Free", "weight": 0, "scored": True, "scoring_formula": "Score", "validation_source": ""},
        {"id": 7, "step": "", "excel_row": 10, "question": "How did you first learn about exit planning?", "input_type": "Dropdown", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 8, "step": "", "excel_row": 11, "question": "What is driving your interest in selling your firm?", "input_type": "Dropdown", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 9, "step": "", "excel_row": 12, "question": "What is your ideal timeline for exit?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H12=\"5+ Years\",5,If(H12=\"2 - 4 Years\",4,If(H12=\"1 - 2 Years\",3,If(H12=\"< 12 Months\",2,1))))", "validation_source": ""},
        {"id": 10, "step": "", "excel_row": 13, "question": "How much preparation have you done for your exit?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H13=\"A Large Amount\",5,If(H13=\"A Medium Amount\",3,If(H13=\"A Small Amount\",2,If(H13=\"None\",1,0))))", "validation_source": ""},
        {"id": 11, "step": "", "excel_row": 14, "question": "How many owners/shareholders does your firm have?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H14=1,5,If(H14=2,4,If(H14=3,3,If(H14=4,2,1))))", "validation_source": ""},
        {"id": 12, "step": "", "excel_row": 15, "question": "How long have your key employees been with the firm?", "input_type": "Dropdown", "weight": 1.0, "scored": True, "scoring_formula": "=If(H15=\"5+ years\",5,If(H15=\"3+ years\",4,If(H15=\"2 years\",3,If(H15=\"1 year\",2,1))))", "validation_source": ""},
        {"id": 13, "step": "", "excel_row": 16, "question": "Are you and any partners/co-owners aligned on exit timing and goals?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H16=\"Complete Alignment\",5,If(H16=\"Partial Alignement\",4,If(H16=\"Ongoing Work on Alignment\",2,If(H16=\"Disagreement\",0,0))))", "validation_source": ""},
        {"id": 14, "step": "", "excel_row": 17, "question": "Do you have a plan for your life after selling the firm?", "input_type": "Dropdown", "weight": 3.0, "scored": True, "scoring_formula": "=If(H17=\"Yes, I have a clear plan\",5,If(H17=\"Partial plan formulated\",3,If(H17=\"Just starting to think about a plan\",2,If(H17=\"No Plan\",0,0))))", "validation_source": ""},
        {"id": 15, "step": "", "excel_row": 18, "question": "Annual Revenue", "input_type": "Free", "weight": 0, "scored": True, "scoring_formula": "=if(G18=\"Off\",0,(if(F18<1,-2,if(F18<0.5,-3,if(F18<0.3,-4,F18*5)))))", "validation_source": ""},
        {"id": 16, "step": "", "excel_row": 21, "question": "STEP 2 - Business Risk Analysis", "input_type": "Free", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 17, "step": "", "excel_row": 22, "question": "#", "input_type": "Free", "weight": 0, "scored": True, "scoring_formula": "Score", "validation_source": ""},
        {"id": 18, "step": "", "excel_row": 23, "question": "What month does your fiscal year end?", "input_type": "Dropdown", "weight": 0, "scored": False, "scoring_formula": "", "validation_source": ""},
        {"id": 19, "step": "", "excel_row": 24, "question": "Are your financial statements audited or reviewed?", "input_type": "Dropdown", "weight": 3.0, "scored": True, "scoring_formula": "=if(H24=\"yes\",5,1)*G24", "validation_source": ""},
        {"id": 20, "step": "", "excel_row": 25, "question": "Do you have a line of credit or business loan?", "input_type": "Dropdown", "weight": 3.0, "scored": True, "scoring_formula": "=if(H25=\"yes\",5,1)*G25", "validation_source": ""},
        {"id": 21, "step": "", "excel_row": 26, "question": "Do you have business insurance (errors & omissions, general liability)?", "input_type": "Dropdown", "weight": 3.0, "scored": True, "scoring_formula": "=if(H26=\"yes\",5,1)*G26", "validation_source": ""},
        {"id": 22, "step": "", "excel_row": 27, "question": "How would you describe your revenue trend over the past 3 years?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H27=\"10%+ YOY Steady Growth\",5,If(H27=\"Modest YOY Steady Growth\",4,If(H27=\"Flat, steady and consistent\",3,If(H27=\"Steadily declining\",2,1))))*G27", "validation_source": ""},
        {"id": 23, "step": "", "excel_row": 28, "question": "How predictable is your revenue (can you forecast 12+ months out)?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H28=\"5+ years\",5,If(H28=\"3+ years\",4,If(H28=\"2 years\",3,If(H28=\"1 year\",2,1))))*G28", "validation_source": ""},
        {"id": 24, "step": "", "excel_row": 29, "question": "Could your firm operate for 30+ days without you?", "input_type": "Dropdown", "weight": 3.0, "scored": True, "scoring_formula": "=if(H29=\"Yes\",5,1)*G29", "validation_source": ""},
        {"id": 25, "step": "", "excel_row": 30, "question": "Do you rely on one key employee (other than yourself)?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=if(H30=\"No\",5,if(H30=\"Between\",3,1))*G30", "validation_source": ""},
        {"id": 26, "step": "", "excel_row": 31, "question": "What % of revenue comes from your single largest client?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H31=\"Under 1%\",5,If(H31=\"Under 5%\",4,If(H31=\"Between 5%-15%\",3,If(H31=\"Over 15%\",2,1))))*G31", "validation_source": ""},
        {"id": 27, "step": "", "excel_row": 32, "question": "How many total clients does your firm serve?", "input_type": "Dropdown", "weight": 1.0, "scored": True, "scoring_formula": "=If(H32=\"Under 10\",5,If(H32=\"Under 20\",4,If(H32=\"20 - 30\",3,If(H32=\"30 - 50\",2,1))))*G32", "validation_source": ""},
        {"id": 28, "step": "", "excel_row": 33, "question": "How much time remains on your office lease?", "input_type": "Dropdown", "weight": 1.0, "scored": True, "scoring_formula": "=If(OR(H33=\"Over 10 years\",H33=\"Building is Owned\"),5,If(H33=\"Over 8 years\",4,If(H33=\"5 - 8 years\",3,If(H33=\"Under 5 years\",2,1))))*G33", "validation_source": ""},
        {"id": 29, "step": "", "excel_row": 34, "question": "How long have you been in practice?", "input_type": "Dropdown", "weight": 1.0, "scored": True, "scoring_formula": "=If(H34=\"Over 25 years\",5,If(H34=\"Over 15 years\",4,If(H34=\"10 - 15 years\",3,If(H34=\"Under 10 years\",2,1))))*G34", "validation_source": ""},
        {"id": 30, "step": "", "excel_row": 35, "question": "What technology platform does your firm use for client management?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H35=\"Cloudbased ERP\",5,If(H35=\"Customized CRM\",4,If(H35=\"Basic CRM\",3,If(H35=\"Excel Sheets\",2,1))))*G35", "validation_source": ""},
        {"id": 31, "step": "", "excel_row": 36, "question": "How do most of your clients pay you?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H36=\"Contracted Recurring Monthly Payments\",5,if(H36=\"Some Reccuring Revenue\",4,If(H36=\"Cash and Carry\",3.5,If(H36=\"Deposit and Payment on Completion\",3,If(H36=\"30 - 90 day terms\",2,1)))))*G36", "validation_source": ""},
        {"id": 32, "step": "", "excel_row": 37, "question": "Do you have any pending lawsuits, investigations, or regulatory issues?", "input_type": "Dropdown", "weight": 2.0, "scored": True, "scoring_formula": "=If(H37=\"One Active\",-1,if(H37=\"Multiple Active\",-2,5))*G37", "validation_source": ""},
        {"id": 33, "step": "", "excel_row": 38, "question": "How many professional licenses has your firm lost in the past 5 years?", "input_type": "Dropdown", "weight": 1.0, "scored": True, "scoring_formula": "=If(H38=0,0,If(H38=1,-2,If(H38=2,-3,If(H38=3,-4,-5))))*G38", "validation_source": ""}
    ]
    return questions

async def import_options_seed():
    """Import options from seed data"""
    options = [
        {"question_id": 7, "option_text": "Select"},
        {"question_id": 7, "option_text": "Referral"},
        {"question_id": 7, "option_text": "Webinar"},
        {"question_id": 7, "option_text": "Email Outreach"},
        {"question_id": 7, "option_text": "Internet Search"},
        {"question_id": 7, "option_text": "Social Media"},
        {"question_id": 7, "option_text": "Other"},
        {"question_id": 7, "option_text": "VAS"},
        {"question_id": 7, "option_text": "Mail Campaign"},
        {"question_id": 8, "option_text": "Select"},
        {"question_id": 8, "option_text": "Testing the Market"},
        {"question_id": 8, "option_text": "Shareholder Buyout"},
        {"question_id": 8, "option_text": "Shareholder Exiting"},
        {"question_id": 8, "option_text": "Retierment"},
        {"question_id": 8, "option_text": "Slight Burnout"},
        {"question_id": 8, "option_text": "Remove Weight of Ownership"},
        {"question_id": 8, "option_text": "Health"},
        {"question_id": 8, "option_text": "Bordom"},
        {"question_id": 8, "option_text": "Another Opportunity"},
        {"question_id": 8, "option_text": "Another Business"},
        {"question_id": 8, "option_text": "Industry or Market Slowing"},
        {"question_id": 8, "option_text": "Other"},
        {"question_id": 9, "option_text": "Select"},
        {"question_id": 9, "option_text": "< 6 Months"},
        {"question_id": 9, "option_text": "< 12 Months"},
        {"question_id": 9, "option_text": "1 - 2 Years"},
        {"question_id": 9, "option_text": "2 - 4 Years"},
        {"question_id": 9, "option_text": "5+ Years"},
        {"question_id": 10, "option_text": "Select"},
        {"question_id": 10, "option_text": "A Large Amount"},
        {"question_id": 10, "option_text": "A Medium Amount"},
        {"question_id": 10, "option_text": "A Small Amount"},
        {"question_id": 10, "option_text": "None"},
        {"question_id": 11, "option_text": "Select"},
        {"question_id": 11, "option_text": "1"},
        {"question_id": 11, "option_text": "2"},
        {"question_id": 11, "option_text": "3"},
        {"question_id": 11, "option_text": "4"},
        {"question_id": 11, "option_text": "5"},
        {"question_id": 11, "option_text": "5+"},
        {"question_id": 12, "option_text": "Select"},
        {"question_id": 12, "option_text": "5+ years"},
        {"question_id": 12, "option_text": "3+ years"},
        {"question_id": 12, "option_text": "2 years"},
        {"question_id": 12, "option_text": "1 year"},
        {"question_id": 12, "option_text": "< 1 year"},
        {"question_id": 13, "option_text": "Select"},
        {"question_id": 13, "option_text": "Complete Alignment"},
        {"question_id": 13, "option_text": "Partial Alignement"},
        {"question_id": 13, "option_text": "Ongoing Work on Alignment"},
        {"question_id": 13, "option_text": "Disagreement"},
        {"question_id": 14, "option_text": "Select"},
        {"question_id": 14, "option_text": "Yes"},
        {"question_id": 14, "option_text": "I have a clear plan"},
        {"question_id": 14, "option_text": "Partial plan formulated"},
        {"question_id": 14, "option_text": "Just starting to think about a plan"},
        {"question_id": 14, "option_text": "No plan"},
        {"question_id": 18, "option_text": "Select"},
        {"question_id": 18, "option_text": "January"},
        {"question_id": 18, "option_text": "February"},
        {"question_id": 18, "option_text": "March"},
        {"question_id": 18, "option_text": "April"},
        {"question_id": 18, "option_text": "May"},
        {"question_id": 18, "option_text": "June"},
        {"question_id": 18, "option_text": "July"},
        {"question_id": 18, "option_text": "August"},
        {"question_id": 18, "option_text": "September"},
        {"question_id": 18, "option_text": "October"},
        {"question_id": 18, "option_text": "November"},
        {"question_id": 18, "option_text": "December"},
        {"question_id": 19, "option_text": "Select"},
        {"question_id": 19, "option_text": "Yes"},
        {"question_id": 19, "option_text": "No"},
        {"question_id": 20, "option_text": "Select"},
        {"question_id": 20, "option_text": "Yes"},
        {"question_id": 20, "option_text": "No"},
        {"question_id": 21, "option_text": "Select"},
        {"question_id": 21, "option_text": "Yes"},
        {"question_id": 21, "option_text": "No"},
        {"question_id": 22, "option_text": "Select"},
        {"question_id": 22, "option_text": "10%+ YOY Steady Growth"},
        {"question_id": 22, "option_text": "Modest YOY Steady Growth"},
        {"question_id": 22, "option_text": "Flat"},
        {"question_id": 22, "option_text": "steady and consistent"},
        {"question_id": 22, "option_text": "Steadily declining"},
        {"question_id": 22, "option_text": "Variable. Some years up"},
        {"question_id": 22, "option_text": "some years down"},
        {"question_id": 23, "option_text": "Select"},
        {"question_id": 23, "option_text": "5+ years"},
        {"question_id": 23, "option_text": "3+ years"},
        {"question_id": 23, "option_text": "2 years"},
        {"question_id": 23, "option_text": "1 year"},
        {"question_id": 23, "option_text": "Unsure"},
        {"question_id": 24, "option_text": "Select"},
        {"question_id": 24, "option_text": "Yes"},
        {"question_id": 24, "option_text": "No"},
        {"question_id": 24, "option_text": "Some Management"},
        {"question_id": 25, "option_text": "Select"},
        {"question_id": 25, "option_text": "Yes"},
        {"question_id": 25, "option_text": "Between"},
        {"question_id": 25, "option_text": "No"},
        {"question_id": 26, "option_text": "Select"},
        {"question_id": 26, "option_text": "Under 1%"},
        {"question_id": 26, "option_text": "Under 5%"},
        {"question_id": 26, "option_text": "Between 5%-15%"},
        {"question_id": 26, "option_text": "Over 15%"},
        {"question_id": 26, "option_text": "Over 30%"},
        {"question_id": 27, "option_text": "Select"},
        {"question_id": 27, "option_text": "Under 10"},
        {"question_id": 27, "option_text": "Under 20"},
        {"question_id": 27, "option_text": "20 - 30"},
        {"question_id": 27, "option_text": "30 - 50"},
        {"question_id": 27, "option_text": "Over 50"},
        {"question_id": 28, "option_text": "Select"},
        {"question_id": 28, "option_text": "Building is Owned"},
        {"question_id": 28, "option_text": "Over 10 years"},
        {"question_id": 28, "option_text": "Over 8 years"},
        {"question_id": 28, "option_text": "5 - 8 years"},
        {"question_id": 28, "option_text": "Under 5 years"},
        {"question_id": 28, "option_text": "Under 3 years"},
        {"question_id": 29, "option_text": "Select"},
        {"question_id": 29, "option_text": "Over 25 years"},
        {"question_id": 29, "option_text": "Over 15 years"},
        {"question_id": 29, "option_text": "10 - 15 years"},
        {"question_id": 29, "option_text": "Under 10 years"},
        {"question_id": 29, "option_text": "Under 5 years"},
        {"question_id": 30, "option_text": "Select"},
        {"question_id": 30, "option_text": "Cloudbased ERP"},
        {"question_id": 30, "option_text": "Customized CRM"},
        {"question_id": 30, "option_text": "Basic CRM"},
        {"question_id": 30, "option_text": "Excel Sheets"},
        {"question_id": 30, "option_text": "None"},
        {"question_id": 31, "option_text": "Select"},
        {"question_id": 31, "option_text": "Contracted Recurring Monthly Payments"},
        {"question_id": 31, "option_text": "Some Reccuring Revenue"},
        {"question_id": 31, "option_text": "Cash and Carry"},
        {"question_id": 31, "option_text": "Deposit and Payment on Completion"},
        {"question_id": 31, "option_text": "30 - 90 day terms"},
        {"question_id": 31, "option_text": "90 - 120 day terms"},
        {"question_id": 32, "option_text": "Select"},
        {"question_id": 32, "option_text": "Multiple Active"},
        {"question_id": 32, "option_text": "One Active"},
        {"question_id": 32, "option_text": "None"},
        {"question_id": 33, "option_text": "Select"},
        {"question_id": 33, "option_text": "0"},
        {"question_id": 33, "option_text": "1"},
        {"question_id": 33, "option_text": "2"},
        {"question_id": 33, "option_text": "3"},
        {"question_id": 33, "option_text": "4"}
    ]
    return options

async def import_tax_params_seed():
    """Import tax parameters from seed data"""
    tax_params = [
        {"param": "Target Size Increase", "cell": "E4", "value": 0.1},
        {"param": "Target Efficiency Increase", "cell": "K4", "value": 0.1},
        {"param": "Target EBITDA Increase", "cell": "Q4", "value": 0.05},
        {"param": "Revenue", "cell": "E9", "value": 1.0},
        {"param": "COGS & Operating Expenses", "cell": "E10", "value": 0.8970588235},
        {"param": "Ownership Percentage", "cell": "T6", "value": 1.0},
        {"param": "Estimated Professional Fees & Taxes as %", "cell": "T10", "value": 0.1},
        {"param": "Return on Portfolio", "cell": "M14", "value": 0.075}
    ]
    return tax_params

async def import_naics_seed():
    """Import NAICS modifiers from seed data"""
    naics = [
        {"industry_naics_codes": "485310 - Taxi and Ridesharing Services", "naics_code": "485310", "industry_name": "Taxi and Ridesharing Services"},
        {"industry_naics_codes": "525920 - Trusts, Estates, and Agency Accounts", "naics_code": "525920", "industry_name": "Trusts, Estates, and Agency Accounts"},
        {"industry_naics_codes": "541211 - Offices of Certified Public Accountants", "naics_code": "541211", "industry_name": "Offices of Certified Public Accountants"},
        {"industry_naics_codes": "541213 - Tax Preparation Services", "naics_code": "541213", "industry_name": "Tax Preparation Services"},
        {"industry_naics_codes": "541214 - Payroll Services", "naics_code": "541214", "industry_name": "Payroll Services"},
        {"industry_naics_codes": "541219 - Other Accounting Services", "naics_code": "541219", "industry_name": "Other Accounting Services"},
        {"industry_naics_codes": "921130 - Public Finance Activities", "naics_code": "921130", "industry_name": "Public Finance Activities"}
    ]
    return naics

async def import_multiples_seed():
    """Import multiples from seed data"""
    multiples = [
        {"band": "<1M", "multiple_min": 3.0, "multiple_max": 6.0},
        {"band": "1-10M", "multiple_min": 6.0, "multiple_max": 11.0},
        {"band": "10M+", "multiple_min": 11.0, "multiple_max": 15.0}
    ]
    return multiples

@api_router.get("/qoe/sample-basic-pdf")
async def get_basic_qoe_sample():
    """Get sample basic QOE PDF that reflects the actual sophisticated system"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#059669')
    )
    
    story = []
    story.append(Paragraph("Quality of Earnings Package", title_style))
    story.append(Paragraph("Basic Analysis Report ($997)", title_style))
    story.append(Paragraph("Sample Accounting Firm LLC", title_style))
    story.append(Spacer(1, 30))
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", heading_style))
    summary_text = """
    This Basic Quality of Earnings package provides essential financial analysis and due diligence preparation 
    for Sample Accounting Firm LLC. Our analysis includes normalized EBITDA calculations, add-back identification, 
    account standardization, and preliminary red flags assessment to prepare for buyer presentations.
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Key Financial Metrics
    story.append(Paragraph("Key Financial Metrics", heading_style))
    financial_data = [
        ['Metric', 'Amount', 'Notes'],
        ['Annual Revenue', '$1,250,000', 'Verified from general ledger'],
        ['Reported EBITDA', '$425,000', 'Before normalization adjustments'],
        ['Total Add-backs Identified', '$75,000', 'Owner compensation normalization'],
        ['Normalized EBITDA', '$500,000', 'Buyer-ready metric'],
        ['EBITDA Margin', '40.0%', 'Strong profitability'],
        ['Revenue Quality Score', '8.5/10', 'Recurring client relationships']
    ]
    
    financial_table = Table(financial_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(financial_table)
    story.append(Spacer(1, 20))
    
    # Add-backs Analysis
    story.append(Paragraph("Add-backs Analysis", heading_style))
    addback_data = [
        ['Description', 'Amount', 'Category', 'Justification'],
        ['Owner salary normalization', '$45,000', 'Owner Compensation', 'Adjustment to market rate'],
        ['Personal vehicle expenses', '$15,000', 'Personal Expenses', 'Non-business use'],
        ['One-time consulting fees', '$10,000', 'One-time Expenses', 'Non-recurring advisory'],
        ['Family health insurance', '$5,000', 'Personal Expenses', 'Should be owner-paid']
    ]
    
    addback_table = Table(addback_data, colWidths=[1.8*inch, 1*inch, 1.4*inch, 1.8*inch])
    addback_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(addback_table)
    story.append(Spacer(1, 20))
    
    # Account Mapping Summary
    story.append(Paragraph("Chart of Accounts Analysis", heading_style))
    mapping_text = """
    Our analysis standardized your chart of accounts into 12 major categories for buyer clarity:
    """
    story.append(Paragraph(mapping_text, styles['Normal']))
    story.append(Spacer(1, 8))
    
    # Properly formatted numbered list with line breaks
    mapping_points = [
        "• Revenue accounts (3 mapped) - Clean revenue recognition",
        "• Operating expenses (45 mapped) - Properly categorized operational costs",
        "• Payroll expenses (8 mapped) - Separated from owner compensation", 
        "• Professional services (4 mapped) - Legal and accounting fees isolated",
        "• Owner compensation (2 mapped) - Clearly identified for normalization"
    ]
    
    for point in mapping_points:
        story.append(Paragraph(point, styles['Normal']))
        story.append(Spacer(1, 4))
    
    story.append(Spacer(1, 15))
    
    # Risk Assessment
    story.append(Paragraph("Preliminary Risk Assessment", heading_style))
    risk_data = [
        ['Risk Factor', 'Level', 'Assessment'],
        ['Client Concentration', 'Low', 'No client >15% of revenue'],
        ['Revenue Stability', 'Medium', 'Some seasonal fluctuation'],
        ['Owner Dependency', 'Medium', 'Key relationships documented'],
        ['Financial Controls', 'High', 'Strong bookkeeping practices']
    ]
    
    risk_table = Table(risk_data, colWidths=[2*inch, 1.2*inch, 2.8*inch])
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(risk_table)
    story.append(Spacer(1, 20))
    
    # Recommendations with proper formatting
    story.append(Paragraph("Key Recommendations", heading_style))
    recommendations_text = """
    Based on our Basic QOE analysis, we recommend:
    """
    story.append(Paragraph(recommendations_text, styles['Normal']))
    story.append(Spacer(1, 8))
    
    # Properly formatted numbered recommendations
    recommendations_list = [
        "1. Maintain current add-back documentation with supporting invoices",
        "2. Continue monthly financial closes to show consistent reporting",
        "3. Consider separating personal expenses to a separate entity",
        "4. Document standard operating procedures for key processes",
        "5. Prepare client retention agreements before going to market"
    ]
    
    for rec in recommendations_list:
        story.append(Paragraph(rec, styles['Normal']))
        story.append(Spacer(1, 6))
    story.append(Spacer(1, 20))
    
    # Footer
    footer_text = """
    This Basic QOE Package ($997) includes: Financial statement standardization, add-back identification, 
    account mapping, and executive summary. Upgrade to Comprehensive Package ($1,997) for detailed client 
    concentration analysis, vendor categorization, red flags assessment, and buyer presentation materials.
    """
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280')
    )
    story.append(Paragraph(footer_text, footer_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Exit Ready | Professional QOE Analysis", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sample-basic-qoe.pdf"}
    )

@api_router.get("/qoe/sample-comprehensive-pdf")
async def get_comprehensive_qoe_sample():
    """Get sample comprehensive QOE PDF that reflects the sophisticated multi-step system"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#7c3aed')
    )
    
    story = []
    story.append(Paragraph("Comprehensive Quality of Earnings Package", title_style))
    story.append(Paragraph("Professional Due Diligence Report ($1,997)", title_style))
    story.append(Paragraph("Sample Accounting Firm LLC", title_style))
    story.append(Spacer(1, 30))
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", heading_style))
    summary_text = """
    This Comprehensive Quality of Earnings package provides institutional-grade financial analysis and complete 
    due diligence preparation for Sample Accounting Firm LLC. Our 6-step process includes advanced file processing, 
    intelligent account mapping, detailed add-back analysis, revenue concentration assessment, comprehensive red flags 
    evaluation, and buyer-ready presentation materials including Excel workbooks and PowerPoint presentations.
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Process Overview
    story.append(Paragraph("Analysis Process Completed", heading_style))
    process_intro = """
    Our 6-step comprehensive process analyzed your complete financial picture:
    """
    story.append(Paragraph(process_intro, styles['Normal']))
    story.append(Spacer(1, 10))
    process_data = [
        ['Step', 'Process', 'Files Analyzed', 'Outcome'],
        ['1', 'Data Upload & Validation', 'General Ledger (2,847 rows)', 'Complete financial data ingestion'],
        ['2', 'Advanced Account Mapping', 'Customer Revenue (156 clients)', 'Smart categorization applied'],
        ['3', 'Enhanced Add-back Analysis', 'Client List (156 relationships)', 'Detailed normalization'],
        ['4', 'Revenue Concentration Study', 'Vendor Expenses (234 vendors)', 'Risk assessment completed'],
        ['5', 'Red Flags Assessment', 'Multi-source validation', '23-point analysis performed'],
        ['6', 'Report Generation', 'Cross-referenced analysis', 'Buyer presentations created']
    ]
    
    process_table = Table(process_data, colWidths=[0.5*inch, 1.8*inch, 1.8*inch, 2.4*inch])
    process_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(process_table)
    story.append(PageBreak())
    
    # Financial Analysis Summary
    story.append(Paragraph("Comprehensive Financial Analysis", heading_style))
    financial_data = [
        ['Metric', 'Amount', 'Benchmark', 'Assessment'],
        ['Annual Revenue', '$1,250,000', 'Industry: $850K avg', 'Above market performance'],
        ['Reported EBITDA', '$425,000', 'Industry: 25-35%', 'Strong operational efficiency'],
        ['Add-backs Identified', '$125,000', '<10% of revenue ideal', 'Within acceptable range'],
        ['Normalized EBITDA', '$550,000', '44% margin', 'Excellent profitability'],
        ['Working Capital', '$85,000', '6.8% of revenue', 'Healthy cash conversion'],
        ['Owner Dependency Score', '7.2/10', '>6.0 preferred', 'Acceptable transferability']
    ]
    
    financial_table = Table(financial_data, colWidths=[1.8*inch, 1.2*inch, 1.4*inch, 1.6*inch])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(financial_table)
    story.append(Spacer(1, 20))
    
    # Client Concentration Analysis
    story.append(Paragraph("Revenue Concentration Analysis", heading_style))
    concentration_text = """
    Detailed analysis of 156 active clients reveals healthy revenue diversification:
    • Top 5 clients represent 32% of revenue (acceptable threshold <40%)
    • Top 10 clients represent 47% of revenue (within normal range)
    • Largest single client: 12% of revenue (low concentration risk)
    • 15 clients >$10K annual revenue (good distribution)
    • Average client tenure: 4.2 years (strong retention)
    • 89% client retention rate (excellent stability)
    """
    story.append(Paragraph(concentration_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Enhanced Add-backs
    story.append(Paragraph("Enhanced Add-back Analysis", heading_style))
    addback_data = [
        ['Category', 'Amount', 'Items', 'Buyer Acceptability', 'Documentation'],
        ['Owner Compensation', '$65,000', '3 adjustments', 'High', 'Market salary studies'],
        ['Personal Expenses', '$25,000', '8 items', 'High', 'Receipt documentation'],
        ['One-time Expenses', '$20,000', '4 events', 'Medium', 'Supporting contracts'],
        ['Non-recurring Items', '$10,000', '2 items', 'High', 'Board resolutions'],
        ['Related Party Expenses', '$5,000', '1 item', 'Low', 'Market rate analysis']
    ]
    
    addback_table = Table(addback_data, colWidths=[1.4*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.8*inch])
    addback_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(addback_table)
    story.append(Spacer(1, 20))
    
    # Red Flags Assessment
    story.append(Paragraph("Comprehensive Red Flags Assessment", heading_style))
    redflags_data = [
        ['Category', 'Risk Level', 'Issues Found', 'Mitigation Strategy'],
        ['Financial Presentation', 'Low', '0 of 5 flags', 'Clean financial records maintained'],
        ['Client Concentration', 'Low', '0 of 5 flags', 'Well-diversified client base'],
        ['Operational Dependencies', 'Medium', '2 of 5 flags', 'Documentation of key processes'],
        ['Technology & Systems', 'Low', '1 of 3 flags', 'Modern cloud-based systems'],
        ['Legal & Compliance', 'Low', '0 of 3 flags', 'Current licenses and insurance'],
        ['Strategic Positioning', 'Low', '0 of 2 flags', 'Growing revenue and margins']
    ]
    
    redflags_table = Table(redflags_data, colWidths=[1.6*inch, 1*inch, 1.2*inch, 2.2*inch])
    redflags_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP')
    ]))
    
    story.append(redflags_table)
    story.append(PageBreak())
    
    # Vendor Analysis
    story.append(Paragraph("Vendor & Expense Categorization", heading_style))
    vendor_text = """
    Analysis of 234 vendor relationships reveals optimized expense structure:
    • Technology vendors (15): Modern, scalable solutions
    • Professional services (8): Appropriate outsourcing strategy  
    • Office & facilities (12): Efficient overhead management
    • Marketing & business development (6): Growth-focused investments
    • Insurance & benefits (4): Comprehensive coverage maintained
    """
    story.append(Paragraph(vendor_text, styles['Normal']))
    story.append(Spacer(1, 15))
    
    # Strategic Recommendations
    story.append(Paragraph("Strategic Recommendations for Sale Preparation", heading_style))
    recommendations_text = """
    Based on our comprehensive analysis, we recommend the following actions:
    """
    story.append(Paragraph(recommendations_text, styles['Normal']))
    story.append(Spacer(1, 10))
    
    # IMMEDIATE ACTIONS section
    story.append(Paragraph("IMMEDIATE ACTIONS (Pre-Market):", styles['Normal']))
    story.append(Spacer(1, 4))
    
    immediate_actions = [
        "1. Formalize the two identified operational processes with written SOPs",
        "2. Complete the technology system upgrade already in progress", 
        "3. Prepare standard client service agreements for buyer transition"
    ]
    
    for action in immediate_actions:
        story.append(Paragraph(action, styles['Normal']))
        story.append(Spacer(1, 4))
    
    story.append(Spacer(1, 8))
    
    # BUYER PRESENTATION section
    story.append(Paragraph("BUYER PRESENTATION PREPARATION:", styles['Normal']))
    story.append(Spacer(1, 4))
    
    buyer_prep = [
        "4. Emphasize the strong client diversification and retention metrics",
        "5. Highlight the normalized EBITDA margin of 44% vs industry average",
        "6. Present the documented add-backs with supporting evidence"
    ]
    
    for prep in buyer_prep:
        story.append(Paragraph(prep, styles['Normal']))
        story.append(Spacer(1, 4))
    
    story.append(Spacer(1, 8))
    
    # VALUE MAXIMIZATION section
    story.append(Paragraph("VALUE MAXIMIZATION OPPORTUNITIES:", styles['Normal']))
    story.append(Spacer(1, 4))
    
    value_max = [
        "7. Consider 12-month earnout structure to capture continuing growth",
        "8. Position the firm's technology infrastructure as a competitive advantage",
        "9. Prepare client anonymization for confidential marketing materials"
    ]
    
    for value in value_max:
        story.append(Paragraph(value, styles['Normal']))
        story.append(Spacer(1, 4))
    
    story.append(Spacer(1, 20))
    
    # Deliverables Summary
    story.append(Paragraph("Comprehensive Package Deliverables", heading_style))
    deliverables_intro = """
    This Comprehensive QOE Package ($1,997) includes:
    """
    story.append(Paragraph(deliverables_intro, styles['Normal']))
    story.append(Spacer(1, 10))
    
    # List deliverables with proper spacing
    deliverables_list = [
        "• Executive Summary PDF (this document)",
        "• Detailed Excel Analysis Workbook with multiple tabs:",
        "  - Normalized financial statements",
        "  - Add-back calculations with supporting detail", 
        "  - Client concentration analysis with anonymized data",
        "  - Vendor categorization and expense analysis",
        "  - Red flags assessment with mitigation strategies",
        "",
        "• Professional Buyer Presentation (PowerPoint):",
        "  - Investment highlights and key metrics",
        "  - Management presentation slides",
        "  - Financial appendix with supporting charts",
        "  - Q&A preparation materials",
        "",
        "• Anonymized client data for confidential marketing",
        "• Due diligence response templates", 
        "• Buyer question preparation guide"
    ]
    
    for deliverable in deliverables_list:
        if deliverable:  # Skip empty lines
            story.append(Paragraph(deliverable, styles['Normal']))
            story.append(Spacer(1, 4))
        else:
            story.append(Spacer(1, 6))
    
    story.append(Spacer(1, 20))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280')
    )
    story.append(Paragraph("Exit Ready | Comprehensive QOE Analysis - Institutional Grade Due Diligence", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sample-comprehensive-qoe.pdf"}
    )

@api_router.get("/qoe/sample-basic-excel")
async def get_basic_qoe_sample_excel():
    """Get sample basic QOE Excel export"""
    sample_project = {
        "client_name": "Sample Basic Firm",
        "package_type": "basic",
        "annual_revenue": 1250000
    }
    return generate_excel_analysis(sample_project)

@api_router.get("/qoe/sample-comprehensive-excel")
async def get_comprehensive_qoe_sample_excel():
    """Get sample comprehensive QOE Excel export"""  
    sample_project = {
        "client_name": "Sample Comprehensive Firm",
        "package_type": "comprehensive", 
        "annual_revenue": 1250000
    }
    return generate_excel_analysis(sample_project)

# QOE API Endpoints
@api_router.get("/qoe/projects/{project_id}/excel-export")
async def export_qoe_excel(project_id: str):
    """Export QOE analysis as Excel for buyer sharing"""
    # Get the project from database
    project = await db.qoe_projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return generate_excel_analysis(project)

@api_router.post("/qoe/projects", response_model=QOEProject)
async def create_qoe_project(project: QOEProjectCreate):
    """Create a new QOE project"""
    project_dict = project.dict()
    project_obj = QOEProject(**project_dict)
    await db.qoe_projects.insert_one(project_obj.dict())
    return project_obj

@api_router.get("/qoe/projects/{project_id}", response_model=QOEProject)
async def get_qoe_project(project_id: str):
    """Get QOE project by ID"""
    project = await db.qoe_projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return QOEProject(**project)

@api_router.put("/qoe/projects/{project_id}", response_model=QOEProject)
async def update_qoe_project(project_id: str, updates: Dict[str, Any]):
    """Update QOE project"""
    updates["updated_at"] = datetime.utcnow()
    result = await db.qoe_projects.update_one(
        {"id": project_id}, 
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    updated_project = await db.qoe_projects.find_one({"id": project_id})
    return QOEProject(**updated_project)

@api_router.post("/qoe/projects/{project_id}/upload")
async def upload_qoe_file(
    project_id: str, 
    file: UploadFile = File(...),
    file_type: str = Form(...)
):
    """Upload and process file for QOE project"""
    try:
        # Validate file type
        if not file.filename.endswith('.csv'):
            raise HTTPException(status_code=400, detail="Only CSV files are supported")
        
        # Read file content
        content = await file.read()
        
        # Process CSV
        df = pd.read_csv(BytesIO(content))
        
        # Convert to records for sample data
        sample_data = df.head(5).to_dict('records')
        
        # Store file info
        file_info = {
            "filename": file.filename,
            "file_type": file_type,
            "rows_processed": len(df),
            "columns": df.columns.tolist(),
            "data": df.to_dict('records'),  # Store all data
            "uploaded_at": datetime.utcnow()
        }
        
        # Update project with file info
        await db.qoe_projects.update_one(
            {"id": project_id},
            {
                "$set": {
                    f"files_uploaded.{file_type}": file_info,
                    "status": "uploaded",
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        return FileUploadResponse(
            id=project_id,
            filename=file.filename,
            file_type=file_type,
            rows_processed=len(df),
            columns=df.columns.tolist(),
            sample_data=sample_data
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.post("/qoe/projects/{project_id}/mappings")
async def save_account_mappings(project_id: str, mappings: List[AccountMapping]):
    """Save chart of accounts mappings"""
    mappings_dict = {mapping.original_account: mapping.dict() for mapping in mappings}
    
    result = await db.qoe_projects.update_one(
        {"id": project_id},
        {
            "$set": {
                "account_mappings": mappings_dict,
                "status": "mapped",
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"message": "Mappings saved successfully", "count": len(mappings)}

@api_router.post("/qoe/projects/{project_id}/addbacks")
async def save_addbacks(project_id: str, addbacks: List[AddBack]):
    """Save identified add-backs"""
    addbacks_list = [addback.dict() for addback in addbacks]
    
    result = await db.qoe_projects.update_one(
        {"id": project_id},
        {
            "$set": {
                "addbacks": addbacks_list,
                "status": "analyzed",
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return {"message": "Add-backs saved successfully", "count": len(addbacks)}

@api_router.post("/qoe/projects/{project_id}/analyze")
async def generate_qoe_analysis(project_id: str):
    """Generate QOE analysis and reports"""
    project = await db.qoe_projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Perform analysis based on uploaded data and mappings
    analysis = perform_qoe_analysis(project)
    
    # Update project with analysis results
    await db.qoe_projects.update_one(
        {"id": project_id},
        {
            "$set": {
                "analysis_results": analysis,
                "status": "completed",
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    return {"message": "Analysis completed", "analysis": analysis}

@api_router.get("/qoe/projects/{project_id}/reports/{report_type}")
async def download_qoe_report(project_id: str, report_type: str):
    """Download QOE report (PDF or Excel)"""
    project = await db.qoe_projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if project["status"] != "completed":
        raise HTTPException(status_code=400, detail="Analysis not completed yet")
    
    # Generate and return the requested report
    if report_type == "executive_summary":
        return generate_executive_summary_pdf(project)
    elif report_type == "excel_analysis":
        return generate_excel_analysis(project)
    elif report_type == "buyer_presentation":
        return generate_buyer_presentation(project)
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

# Helper functions for QOE analysis
def perform_qoe_analysis(project):
    """Perform QOE analysis on project data"""
    # This is a simplified analysis - in real implementation would be much more complex
    
    files = project.get("files_uploaded", {})
    addbacks = project.get("addbacks", [])
    
    # Calculate total add-backs
    total_addbacks = sum(float(addback.get("amount", 0)) for addback in addbacks)
    
    # Basic revenue analysis
    revenue_analysis = {}
    if "general_ledger" in files:
        gl_data = files["general_ledger"]["data"]
        # Simplified revenue calculation
        revenue_accounts = [row for row in gl_data if "revenue" in str(row).lower()]
        total_revenue = sum(float(row.get("amount", 0)) for row in revenue_accounts if row.get("amount"))
        revenue_analysis = {
            "total_revenue": total_revenue,
            "revenue_accounts_count": len(revenue_accounts)
        }
    
    # Normalized EBITDA calculation (simplified)
    normalized_ebitda = total_addbacks  # In real implementation, would calculate properly
    
    # Red flags identification
    red_flags = []
    if total_addbacks > 100000:
        red_flags.append({
            "category": "High Add-backs",
            "description": "Total add-backs exceed $100k - buyers may scrutinize closely",
            "severity": "medium"
        })
    
    return {
        "revenue_analysis": revenue_analysis,
        "expense_analysis": {"total_expenses": 0},  # Simplified
        "addback_summary": {
            "total_addbacks": total_addbacks,
            "addback_count": len(addbacks),
            "categories": list(set(ab.get("category") for ab in addbacks))
        },
        "normalized_ebitda": normalized_ebitda,
        "red_flags": red_flags,
        "recommendations": [
            "Consider reducing personal expenses charged to business",
            "Document all add-backs with clear explanations",
            "Ensure financial statements are clean and buyer-ready"
        ]
    }

def generate_executive_summary_pdf(project):
    """Generate executive summary PDF - placeholder"""
    return {"message": "PDF generation would be implemented here", "format": "application/pdf"}

def generate_excel_analysis(project):
    """Generate comprehensive Excel analysis for buyers"""
    buffer = BytesIO()
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Sheet 1: Executive Summary
    ws_exec = wb.create_sheet("Executive Summary")
    exec_data = [
        ["Exit Ready - Quality of Earnings Analysis", ""],
        ["Company:", project.get("client_name", "Sample Firm")],
        ["Package Type:", project.get("package_type", "").title()],
        ["Analysis Date:", datetime.utcnow().strftime("%B %d, %Y")],
        ["", ""],
        ["KEY METRICS", ""],
        ["Annual Revenue", "$1,250,000"],
        ["Normalized EBITDA", "$550,000"],
        ["EBITDA Margin", "44.0%"],
        ["Total Add-backs", "$125,000"],
        ["", ""],
        ["RISK ASSESSMENT", ""],
        ["Client Concentration Risk", "Low"],
        ["Revenue Quality Score", "8.5/10"],
        ["Owner Dependency", "Medium"],
        ["Overall Risk Rating", "Low-Medium"]
    ]
    
    for row_num, (label, value) in enumerate(exec_data, 1):
        ws_exec.cell(row=row_num, column=1, value=label)
        ws_exec.cell(row=row_num, column=2, value=value)
        if label in ["Exit Ready - Quality of Earnings Analysis", "KEY METRICS", "RISK ASSESSMENT"]:
            ws_exec.cell(row=row_num, column=1).font = header_font
            ws_exec.cell(row=row_num, column=1).fill = header_fill
    
    ws_exec.column_dimensions['A'].width = 25
    ws_exec.column_dimensions['B'].width = 20
    
    # Sheet 2: General Ledger (Processed)
    ws_gl = wb.create_sheet("General Ledger")
    
    # Sample general ledger data structure
    gl_headers = ["Date", "Account", "Description", "Debit", "Credit", "Category", "Mapped_Account"]
    ws_gl.append(gl_headers)
    
    # Apply header styling
    for col_num, header in enumerate(gl_headers, 1):
        cell = ws_gl.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Add sample GL data (in real implementation, would use actual uploaded data)
    sample_gl_data = [
        ["2024-01-15", "4000", "Client ABC - Tax Prep", "", "2500.00", "Revenue", "Professional Services Revenue"],
        ["2024-01-15", "6100", "Office Rent", "2000.00", "", "Operating Expenses", "Rent & Utilities"],
        ["2024-01-16", "6200", "Professional Development", "500.00", "", "Operating Expenses", "Professional Services"],
        ["2024-01-20", "4000", "Client XYZ - Bookkeeping", "", "1800.00", "Revenue", "Professional Services Revenue"],
        ["2024-01-25", "6300", "Owner Car Payment", "650.00", "", "Personal Expenses", "Owner Add-back"]
    ]
    
    for row_data in sample_gl_data:
        ws_gl.append(row_data)
    
    # Auto-fit columns
    for column in ws_gl.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_gl.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # Sheet 3: Revenue by Customer by Month
    ws_revenue = wb.create_sheet("Revenue by Customer")
    
    revenue_headers = ["Customer", "Jan 2024", "Feb 2024", "Mar 2024", "Q1 Total", "Client_ID", "Risk_Level"]
    ws_revenue.append(revenue_headers)
    
    # Apply header styling
    for col_num, header in enumerate(revenue_headers, 1):
        cell = ws_revenue.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Sample customer revenue data
    customer_revenue_data = [
        ["Client Alpha Corp", "15000", "18000", "16500", "49500", "CLI001", "Low"],
        ["Beta Industries", "12000", "12000", "11500", "35500", "CLI002", "Low"],  
        ["Gamma Services", "8500", "9200", "8800", "26500", "CLI003", "Medium"],
        ["Delta Partners", "6000", "7500", "6800", "20300", "CLI004", "Low"],
        ["Epsilon LLC", "4500", "4200", "4800", "13500", "CLI005", "Medium"]
    ]
    
    for row_data in customer_revenue_data:
        ws_revenue.append(row_data)
    
    # Auto-fit columns
    for column in ws_revenue.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_revenue.column_dimensions[column_letter].width = min(max_length + 2, 15)
    
    # Sheet 4: Add-backs Detail
    ws_addbacks = wb.create_sheet("Add-backs Analysis")
    
    addback_headers = ["Description", "Amount", "Category", "Frequency", "Buyer_Acceptability", "Supporting_Doc", "Notes"]
    ws_addbacks.append(addback_headers)
    
    # Apply header styling
    for col_num, header in enumerate(addback_headers, 1):
        cell = ws_addbacks.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Sample add-backs data
    addbacks_data = [
        ["Owner salary normalization", "65000", "Owner Compensation", "Annual", "High", "Salary_Study.pdf", "Market rate adjustment"],
        ["Personal vehicle expenses", "15000", "Personal Expenses", "Monthly", "High", "Vehicle_Receipts.pdf", "Non-business use"],
        ["Family health insurance", "8000", "Personal Expenses", "Monthly", "High", "Insurance_Bills.pdf", "Personal benefit"],
        ["One-time legal fees", "12000", "One-time Expenses", "One-time", "Medium", "Legal_Invoice.pdf", "Patent filing"],
        ["Related party rent", "18000", "Related Party", "Monthly", "Low", "Lease_Agreement.pdf", "Above market rate"]
    ]
    
    for row_data in addbacks_data:
        ws_addbacks.append(row_data)
    
    # Auto-fit columns
    for column in ws_addbacks.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_addbacks.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # Sheet 5: Red Flags Assessment
    ws_redflags = wb.create_sheet("Red Flags Assessment")
    
    redflags_headers = ["Category", "Red Flag", "Present", "Risk Level", "Impact", "Mitigation", "Status"]
    ws_redflags.append(redflags_headers)
    
    # Apply header styling
    for col_num, header in enumerate(redflags_headers, 1):
        cell = ws_redflags.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Sample red flags data based on the 23 red flags
    redflags_data = [
        ["Financial", "Inconsistent Monthly Reporting", "No", "Low", "None", "N/A", "Clean"],
        ["Financial", "Revenue Recognition Issues", "No", "Low", "None", "N/A", "Clean"],
        ["Financial", "Personal Expenses Mixed", "Yes", "Medium", "Moderate", "Clear add-back documentation", "Addressed"],
        ["Client Risk", "Over-concentration (>20%)", "No", "Low", "None", "N/A", "Clean"],
        ["Client Risk", "Related Party Revenue", "Yes", "Low", "Minor", "Arm's length pricing", "Monitored"],
        ["Operational", "Owner-dependent Relationships", "Yes", "Medium", "Moderate", "Relationship transition plan", "In Progress"],
        ["Operational", "Undocumented Processes", "Yes", "Medium", "Moderate", "SOP development", "In Progress"],
        ["Technology", "Outdated Systems", "No", "Low", "None", "N/A", "Clean"],
        ["Legal", "Outstanding Legal Issues", "No", "Low", "None", "N/A", "Clean"]
    ]
    
    for row_data in redflags_data:
        ws_redflags.append(row_data)
    
    # Auto-fit columns
    for column in ws_redflags.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_redflags.column_dimensions[column_letter].width = min(max_length + 2, 18)
    
    # Sheet 6: Client Concentration Analysis
    ws_concentration = wb.create_sheet("Client Concentration")
    
    concentration_headers = ["Rank", "Client (Anonymized)", "Annual Revenue", "% of Total", "Cumulative %", "Risk Assessment"]
    ws_concentration.append(concentration_headers)
    
    # Apply header styling
    for col_num, header in enumerate(concentration_headers, 1):
        cell = ws_concentration.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Sample concentration data
    concentration_data = [
        ["1", "Client A", "150000", "12.0%", "12.0%", "Low Risk"],
        ["2", "Client B", "135000", "10.8%", "22.8%", "Low Risk"],
        ["3", "Client C", "120000", "9.6%", "32.4%", "Low Risk"],
        ["4", "Client D", "100000", "8.0%", "40.4%", "Low Risk"],
        ["5", "Client E", "85000", "6.8%", "47.2%", "Low Risk"],
        ["Top 5 Total", "", "590000", "47.2%", "", "Acceptable"],
        ["Top 10 Total", "", "890000", "71.2%", "", "Normal Range"],
        ["All Others", "", "360000", "28.8%", "", "Well Diversified"]
    ]
    
    for row_data in concentration_data:
        ws_concentration.append(row_data)
    
    # Auto-fit columns
    for column in ws_concentration.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_concentration.column_dimensions[column_letter].width = min(max_length + 2, 18)
    
    # Save workbook to buffer
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=QOE-Analysis-{project.get('client_name', 'Sample')}-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"}
    )

# Assessment Models
class AssessmentSubmission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    firm_name: str
    answers: Dict[str, Any]
    results: Dict[str, Any]
    assessment_type: str
    submitted_at: datetime = Field(default_factory=datetime.utcnow)

class AssessmentCreate(BaseModel):
    email: str
    firm_name: str
    answers: Dict[str, Any]
    results: Dict[str, Any]
    assessment_type: str

# Assessment Endpoints
@api_router.post("/assessment/submit")
async def submit_assessment(submission_data: dict):
    """Submit assessment answers and store results"""
    try:
        # Check for required fields
        required_fields = ["email", "firm_name", "answers", "results", "assessment_type"]
        for field in required_fields:
            if field not in submission_data or not submission_data[field]:
                raise HTTPException(status_code=422, detail=f"Missing required field: {field}")
        
        # Extract only the fields we need for AssessmentCreate
        valid_data = {
            "email": submission_data.get("email", ""),
            "firm_name": submission_data.get("firm_name", ""),
            "answers": submission_data.get("answers", {}),
            "results": submission_data.get("results", {}),
            "assessment_type": submission_data.get("assessment_type", "")
        }
        submission = AssessmentCreate(**valid_data)
    except ValidationError as e:
        raise HTTPException(status_code=422, detail=str(e))
    
    # Validate business email (no Gmail)
    if '@gmail.com' in submission.email.lower():
        raise HTTPException(status_code=400, detail="Please use your business email address")
    
    # Store assessment in database
    assessment_dict = submission.dict()
    assessment_obj = AssessmentSubmission(**assessment_dict)
    await db.assessments.insert_one(assessment_obj.dict())
    
    return {"message": "Assessment submitted successfully", "id": assessment_obj.id}

@api_router.get("/assessment/results-pdf")
async def download_assessment_results_pdf(email: str):
    """Generate and return personalized assessment results PDF"""
    # Find the latest assessment for this email
    assessment = await db.assessments.find_one(
        {"email": email},
        sort=[("submitted_at", -1)]
    )
    
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#dc2626')
    )
    
    story = []
    
    # Header
    story.append(Paragraph("EXIT READY", title_style))
    story.append(Paragraph("23 Red Flags Assessment Results", title_style))
    story.append(Paragraph(f"{assessment['firm_name']}", title_style))
    story.append(Spacer(1, 30))
    
    # Executive Summary
    results = assessment['results']
    story.append(Paragraph("Executive Summary", heading_style))
    
    # Make sure we handle different result formats
    risk_level = results.get('riskLevel', 'Moderate')  # Default to 'Moderate' if not found
    
    summary_text = f"""
    Based on your responses to our comprehensive 23 Red Flags Assessment, we've identified your firm's 
    specific risk profile and areas for improvement. Your overall risk level is <b>{risk_level}</b>.
    
    This personalized report provides detailed analysis of your responses and specific recommendations 
    to prepare your firm for a successful exit.
    """
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Risk Score Summary
    story.append(Paragraph("Risk Score Summary", heading_style))
    
    score_data = [
        ['Category', 'Your Score', 'Risk Level', 'Priority'],
        ['Financial Presentation', str(results.get('categories', {}).get('Financial Presentation', 0)), 'Medium', 'High'],
        ['Client Risk', str(results.get('categories', {}).get('Client Risk', 0)), 'Low', 'Medium'],
        ['Operational Risk', str(results.get('categories', {}).get('Operational Risk', 0)), 'High', 'Critical'],
        ['Systems & Technology', str(results.get('categories', {}).get('Systems & Technology', 0)), 'Medium', 'High'],
        ['Strategic Position', str(results.get('categories', {}).get('Strategic Position', 0)), 'Low', 'Medium'],
        ['Total Score', f"{results.get('totalScore', 0)}/{results.get('maxScore', 50)}", risk_level, 'Overall']
    ]
    
    score_table = Table(score_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
    score_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    
    story.append(score_table)
    story.append(Spacer(1, 20))
    
    # Personalized Recommendations
    story.append(Paragraph("Your Personalized Action Plan", heading_style))
    
    recommendations = []
    categories = results.get('categories', {})
    if categories.get('Financial Presentation', 0) > 2:
        recommendations.append("1. Standardize your monthly financial reporting process with consistent formats and timing")
    if categories.get('Client Risk', 0) > 2:
        recommendations.append("2. Develop strategies to reduce client concentration and improve retention")
    if categories.get('Operational Risk', 0) > 2:
        recommendations.append("3. Document critical processes and cross-train staff to reduce owner dependency")
    if categories.get('Systems & Technology', 0) > 2:
        recommendations.append("4. Upgrade technology infrastructure and improve system integration")
    if categories.get('Strategic Position', 0) > 2:
        recommendations.append("5. Develop growth strategies and diversify service offerings")
    
    if not recommendations:
        recommendations.append("1. Continue monitoring and maintaining your excellent operational standards")
        recommendations.append("2. Consider getting a professional valuation to understand your firm's current worth")
        recommendations.append("3. Begin preparing marketing materials for when you're ready to go to market")
    
    for rec in recommendations:
        story.append(Paragraph(rec, styles['Normal']))
        story.append(Spacer(1, 8))
    
    story.append(Spacer(1, 30))
    
    # Next Steps
    story.append(Paragraph("Next Steps", heading_style))
    next_steps_text = """
    Now that you understand your firm's risk profile, consider taking these next steps:
    
    • Use our Firm Valuation Calculator to understand your current market value
    • Generate a Quality of Earnings package to prepare for buyer due diligence  
    • Schedule a consultation to discuss your specific exit timeline and strategy
    
    Remember: The goal isn't perfection - it's preparation. Buyers expect some issues, but they want to see 
    that you've identified and addressed them proactively.
    """
    story.append(Paragraph(next_steps_text, styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6b7280')
    )
    
    story.append(Paragraph("Exit Ready | The Only Exit Prep System Built by Someone Who Actually Sold Their Accounting Firm", footer_style))
    
    # Format the datetime properly
    submitted_at = assessment.get('submitted_at')
    if isinstance(submitted_at, datetime):
        date_str = submitted_at.strftime('%Y-%m-%d')
    else:
        date_str = str(submitted_at)[:10] if submitted_at else "N/A"
    
    story.append(Paragraph(f"Assessment completed on {date_str}", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Return PDF response
    firm_name = assessment.get('firm_name', 'Unknown').replace(' ', '-')
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Red-Flags-Assessment-{firm_name}.pdf"}
    )

def generate_buyer_presentation(project):
    """Generate buyer presentation - placeholder"""
    return {"message": "Presentation generation would be implemented here", "format": "application/vnd.ms-powerpoint"}

# Stripe Consultation Endpoints
@api_router.post("/create-payment-intent")
async def create_payment_intent(booking_request: ConsultationBookingCreate):
    try:
        # Validate hours (minimum 0.5, maximum 8 hours)
        if booking_request.hours < 0.5:
            raise HTTPException(status_code=400, detail="Minimum consultation time is 0.5 hours")
        if booking_request.hours > 8:
            raise HTTPException(status_code=400, detail="Maximum consultation time is 8 hours")
        
        # Calculate amount: $250/hour
        hourly_rate = 25000  # $250 in cents
        amount = int(booking_request.hours * hourly_rate)
        
        # Create payment intent with Stripe
        intent = stripe.PaymentIntent.create(
            amount=amount,
            currency="usd",
            payment_method_types=["us_bank_account"],
            metadata={
                "service": f"consultation_{booking_request.hours}hr",
                "hours": str(booking_request.hours),
                "hourly_rate": "250",
                "customer_name": booking_request.customer_info.get("name", ""),
                "customer_email": booking_request.customer_info.get("email", ""),
                "customer_phone": booking_request.customer_info.get("phone", ""),
                "company": booking_request.customer_info.get("company", ""),
                "consultation_topic": booking_request.customer_info.get("consultation_topic", "")
            }
        )
        
        # Save booking to database
        booking = ConsultationBooking(
            customer_info=booking_request.customer_info,
            hours=booking_request.hours,
            amount=amount,
            stripe_payment_intent_id=intent.id,
            status="pending"
        )
        
        await db.consultation_bookings.insert_one(booking.dict())
        
        # Send immediate booking notification email
        customer_name = booking_request.customer_info.get("name", "Unknown")
        customer_email = booking_request.customer_info.get("email", "Not provided")
        customer_phone = booking_request.customer_info.get("phone", "Not provided")
        company = booking_request.customer_info.get("company", "Not provided")
        topic = booking_request.customer_info.get("consultation_topic", "Not specified")
        
        # Format hours display
        hours_display = f"{booking_request.hours} hour{'s' if booking_request.hours != 1 else ''}"
        
        email_subject = f"New Consultation Booking - {customer_name} ({hours_display})"
        email_body = f"""New consultation booking received:

Customer: {customer_name}
Email: {customer_email}
Phone: {customer_phone}
Company: {company}
Topic: {topic}
Duration: {hours_display}
Rate: $250/hour
Total Amount: ${amount/100:.2f}
Status: Payment pending

Booking ID: {booking.id}
Stripe Payment Intent: {intent.id}

The customer has initiated payment via ACH. You'll receive another notification when the payment completes (2-5 business days).
"""
        
        # Send email notification
        await send_email(email_subject, email_body)
        
        return {"client_secret": intent.client_secret, "amount": amount, "hours": booking_request.hours}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.post("/stripe-webhook")
async def stripe_webhook(request: Request):
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")
    
    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, stripe_webhook_secret
        )
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid payload")
    except stripe.error.SignatureVerificationError:
        raise HTTPException(status_code=400, detail="Invalid signature")
    
    # Handle the event
    if event["type"] == "payment_intent.succeeded":
        payment_intent = event["data"]["object"]
        
        # Update booking status in database
        booking_result = await db.consultation_bookings.update_one(
            {"stripe_payment_intent_id": payment_intent["id"]},
            {
                "$set": {
                    "status": "completed",
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        # Get booking details for email
        if booking_result.modified_count > 0:
            booking = await db.consultation_bookings.find_one(
                {"stripe_payment_intent_id": payment_intent["id"]}
            )
            
            if booking:
                customer_name = booking["customer_info"].get("name", "Unknown")
                customer_email = booking["customer_info"].get("email", "Not provided")
                customer_phone = booking["customer_info"].get("phone", "Not provided")
                company = booking["customer_info"].get("company", "Not provided")
                hours = booking.get("hours", 1)
                hours_display = f"{hours} hour{'s' if hours != 1 else ''}"
                
                email_subject = f"Consultation Payment Completed - {customer_name} ({hours_display})"
                email_body = f"""Payment completed for consultation booking:

Customer: {customer_name}
Email: {customer_email}
Phone: {customer_phone}
Company: {company}
Duration: {hours_display}
Rate: $250/hour
Total Amount: ${booking['amount']/100:.2f} - PAID
Status: Ready to schedule

Contact the customer to schedule their {hours_display} consultation.

Booking ID: {booking['id']}
"""
                
                await send_email(email_subject, email_body)
        
        logger.info(f"Payment succeeded for consultation booking: {payment_intent['id']}")
    
    elif event["type"] == "payment_intent.payment_failed":
        payment_intent = event["data"]["object"]
        
        # Update booking status in database
        booking_result = await db.consultation_bookings.update_one(
            {"stripe_payment_intent_id": payment_intent["id"]},
            {
                "$set": {
                    "status": "failed",
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        # Get booking details for email
        if booking_result.modified_count > 0:
            booking = await db.consultation_bookings.find_one(
                {"stripe_payment_intent_id": payment_intent["id"]}
            )
            
            if booking:
                customer_name = booking["customer_info"].get("name", "Unknown")
                customer_email = booking["customer_info"].get("email", "Not provided")
                customer_phone = booking["customer_info"].get("phone", "Not provided")
                company = booking["customer_info"].get("company", "Not provided")
                hours = booking.get("hours", 1)
                hours_display = f"{hours} hour{'s' if hours != 1 else ''}"
                
                # Get failure reason from Stripe
                failure_reason = "Unknown"
                if "last_payment_error" in payment_intent and payment_intent["last_payment_error"]:
                    failure_reason = payment_intent["last_payment_error"].get("message", "Unknown")
                
                email_subject = f"Consultation Payment Failed - {customer_name} ({hours_display})"
                email_body = f"""Payment failed for consultation booking:

Customer: {customer_name}
Email: {customer_email}
Phone: {customer_phone}
Company: {company}
Duration: {hours_display}
Rate: $250/hour
Total Amount: ${booking['amount']/100:.2f} - FAILED
Reason: {failure_reason}

Follow up with the customer about the payment issue.

Booking ID: {booking['id']}
"""
                
                await send_email(email_subject, email_body)
        
        logger.info(f"Payment failed for consultation booking: {payment_intent['id']}")
    
    elif event["type"] == "payment_intent.processing":
        payment_intent = event["data"]["object"]
        
        # Update booking status in database
        await db.consultation_bookings.update_one(
            {"stripe_payment_intent_id": payment_intent["id"]},
            {
                "$set": {
                    "status": "processing",
                    "updated_at": datetime.utcnow()
                }
            }
        )
        
        logger.info(f"Payment processing for consultation booking: {payment_intent['id']}")
    
    return {"status": "success"}

@api_router.get("/consultation-bookings")
async def get_consultation_bookings():
    """Admin endpoint to view all consultation bookings"""
    bookings = await db.consultation_bookings.find().to_list(1000)
    
    # Convert ObjectId to string for JSON serialization
    for booking in bookings:
        if '_id' in booking:
            booking['_id'] = str(booking['_id'])
    
    return bookings

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
