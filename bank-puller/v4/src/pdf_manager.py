"""Shared PDF rename/validate module.

Centralized utility for all bank scripts. Not per-bank — every bank calls this.

After a bank run downloads PDFs to ~/Downloads, this module:
1. Validates each PDF (magic bytes, size)
2. Extracts account number from PDF content (last 4-5 digits)
3. Looks up client_name from clients.xlsx (by username + bank_name)
4. Renames to: ClientName__BankName #xxxx yyyy-mm.pdf
5. Moves to output/{yyyy-mm}/
"""

from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Optional

import openpyxl
import pypdf


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_pdf(file_path: Path) -> tuple[bool, str]:
    """Check %PDF magic bytes and file size (10KB - 50MB).

    Returns (True, "245KB") on success or (False, reason) on failure.
    """
    if not file_path.exists():
        return False, f"file not found: {file_path}"

    size = file_path.stat().st_size
    if size < 10_240:
        return False, f"too small: {size} bytes (min 10KB)"
    if size > 50 * 1024 * 1024:
        return False, f"too large: {size // 1024 // 1024}MB (max 50MB)"

    with open(file_path, "rb") as f:
        header = f.read(5)

    if not header.startswith(b"%PDF"):
        preview = header[:20].decode("ascii", errors="replace")
        return False, f"not a PDF: starts with '{preview}'"

    return True, f"{size // 1024}KB"


# ---------------------------------------------------------------------------
# Account number extraction
# ---------------------------------------------------------------------------

def extract_account_number(file_path: Path) -> Optional[str]:
    """Open a PDF and extract the account number (last 4-5 digits).

    Reads text from the first page and searches for common patterns:
    - "Account Number: XXXX XXXX XXXX 3623" (credit cards)
    - "Account Number: ...9866" or full numeric like "000000733659866"
    - Fallback: match the digits from the original filename (e.g., "9866" from "statements-9866-.pdf")

    Returns the last 4-5 digits as a string, or None if not found.
    """
    # Pattern 1 (most reliable): Extract from the original filename.
    # Chase: "20260331-statements-9866-.pdf"
    # This is always correct — the bank puts the right last4 in the filename.
    fname_match = re.search(r"statements?-(\d{4,5})-?", file_path.name)
    if fname_match:
        return fname_match.group(1)

    # If filename doesn't have it, read the PDF content
    try:
        reader = pypdf.PdfReader(str(file_path))
        if not reader.pages:
            return None
        text = reader.pages[0].extract_text() or ""
    except Exception:
        return None

    # Pattern 2: "Account Number: XXXX XXXX XXXX 3623" (credit cards)
    match = re.search(r"Account\s*(?:Number|#)\s*:\s*(?:X+\s*)+(\d{4,5})\b", text, re.IGNORECASE)
    if match:
        return match.group(1)

    # Pattern 3: "Account Number:" followed by digits
    match = re.search(r"Account\s*(?:Number|#)\s*:\s*\D*(\d{4,5})\b", text, re.IGNORECASE)
    if match:
        return match.group(1)

    # Pattern 4: Long account number like "000000733659866" — take last 4
    match = re.search(r"\b0{4,}\d+?(\d{4})\b", text)
    if match:
        return match.group(1)

    return None


# ---------------------------------------------------------------------------
# Canonical filename
# ---------------------------------------------------------------------------

def build_canonical_filename(
    client_name: str, bank_name: str, account_digits: str, month: str
) -> str:
    """Build: ClientName__BankName #xxxx yyyy-mm.pdf

    - Spaces in client_name become underscores
    - bank_name is title-cased
    - account_digits: 4-5 digits as-is
    - month: yyyy-mm format
    """
    clean_client = client_name.replace(" ", "_")
    clean_bank = bank_name.strip().title()
    return f"{clean_client}__{clean_bank} #{account_digits} {month}.pdf"


# ---------------------------------------------------------------------------
# Client name lookup
# ---------------------------------------------------------------------------

def lookup_client_name(
    username: str, bank_name: str, excel_path: Path
) -> Optional[str]:
    """Match username + bank_name in clients.xlsx to get client_name.

    Searches columns case-insensitively. Returns None if not found.
    """
    if not excel_path.exists():
        return None

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    ws = wb.active

    # Find column indices by header
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=0):
        if cell.value:
            key = str(cell.value).strip().lower().replace(" ", "_")
            headers[key] = col_idx

    # Map possible header names
    username_col = headers.get("username") or headers.get("user")
    bank_col = headers.get("bank_name") or headers.get("bank")
    client_col = headers.get("client_name") or headers.get("client_name") or headers.get("client")

    if username_col is None or client_col is None:
        wb.close()
        return None

    username_lower = username.strip().lower()
    bank_lower = bank_name.strip().lower()

    for row in ws.iter_rows(min_row=2):
        row_user = str(row[username_col].value or "").strip().lower()
        row_bank = str(row[bank_col].value or "").strip().lower() if bank_col is not None else ""

        if row_user == username_lower:
            # If bank column exists, also match on bank
            if bank_col is not None and bank_lower and row_bank and bank_lower not in row_bank and row_bank not in bank_lower:
                continue
            client = str(row[client_col].value or "").strip()
            if client:
                wb.close()
                return client

    wb.close()
    return None


# ---------------------------------------------------------------------------
# Batch rename
# ---------------------------------------------------------------------------

def rename_batch(
    download_dir: Path,
    username: str,
    bank_name: str,
    month: str,
    excel_path: Path,
    output_base: Path = Path("output"),
    max_age_sec: int = 600,
) -> list[dict]:
    """Batch rename all new PDFs from a bank run.

    Scans download_dir for PDF files modified within max_age_sec.
    For each:
      1. Validate (magic bytes + size)
      2. Extract account number from PDF content
      3. Look up client_name from clients.xlsx
      4. Rename and move to output/{month}/

    Returns list of result dicts:
      {"account": "9866", "status": "ok", "source": Path, "dest": Path}
      {"account": None, "status": "error", "source": Path, "reason": "..."}
    """
    import time

    now = time.time()
    output_dir = output_base / month
    output_dir.mkdir(parents=True, exist_ok=True)

    # Find recent PDFs (exclude .crdownload, .tmp)
    recent_pdfs = []
    for f in download_dir.glob("*.pdf"):
        if f.stat().st_mtime > now - max_age_sec:
            recent_pdfs.append(f)

    recent_pdfs.sort(key=lambda f: f.stat().st_mtime)

    # Lookup client name once
    client_name = lookup_client_name(username, bank_name, excel_path)

    results = []
    for pdf in recent_pdfs:
        # Validate
        valid, detail = validate_pdf(pdf)
        if not valid:
            results.append({
                "account": None,
                "status": "error",
                "source": pdf,
                "reason": f"validation failed: {detail}",
            })
            continue

        # Extract account number
        account = extract_account_number(pdf)
        if not account:
            results.append({
                "account": None,
                "status": "error",
                "source": pdf,
                "reason": "could not extract account number from PDF",
            })
            continue

        # Build canonical filename
        if client_name:
            canonical = build_canonical_filename(client_name, bank_name, account, month)
        else:
            # Fallback: use username if client_name not found in Excel
            canonical = build_canonical_filename(username, bank_name, account, month)

        dest = output_dir / canonical

        # Handle duplicates (don't overwrite)
        if dest.exists():
            results.append({
                "account": account,
                "status": "skipped",
                "source": pdf,
                "dest": dest,
                "reason": "file already exists",
            })
            continue

        # Move
        try:
            shutil.move(str(pdf), str(dest))
            results.append({
                "account": account,
                "status": "ok",
                "source": pdf,
                "dest": dest,
            })
        except OSError as e:
            results.append({
                "account": account,
                "status": "error",
                "source": pdf,
                "reason": f"move failed: {e}",
            })

    return results
