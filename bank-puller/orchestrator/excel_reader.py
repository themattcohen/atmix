"""Excel reader/writer for bank-puller client workbook.

Manages clients.xlsx with three sheets: Accounts, Run Log, Retry Queue.
Provides file-level locking to prevent concurrent access corruption.
"""

from __future__ import annotations

import os
import re
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from config import CLIENTS_XLSX, MAX_RETRY_COUNT, RETRY_BACKOFF
from utils.logger import log

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
LOCK_TIMEOUT_SECONDS = 3600  # 1 hour — locks older than this are stale
SHEET_ACCOUNTS = "Accounts"
SHEET_RUN_LOG = "Run Log"
SHEET_RETRY_QUEUE = "Retry Queue"

ACCOUNTS_COLUMNS = [
    "client_name", "file_safe_name", "bank_name", "login_url",
    "url_verified_date", "username", "password", "account_last4",
    "statement_available_date", "2fa", "2fa_target", "2fa_method",
    "2fa_detail", "2fa_preference_order", "2fa_sender",
    "sq1_keyword", "sq1_answer", "sq2_keyword", "sq2_answer",
    "sq3_keyword", "sq3_answer", "sq4_keyword", "sq4_answer",
    "sq5_keyword", "sq5_answer",
    "last_successful_login", "status", "notes",
]

RUN_LOG_COLUMNS = [
    "timestamp", "client_name", "bank_name", "account_last4",
    "target_month", "status", "filename", "notes",
    "ai_calls_count", "ai_cost_usd", "duration_seconds",
]

RETRY_QUEUE_COLUMNS = [
    "timestamp", "client_name", "bank_name", "account_last4",
    "target_month", "failure_reason", "retry_count",
    "next_retry_date", "status",
]


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------
@dataclass
class AccountJob:
    """A single bank account to download statements from."""

    client_name: str
    file_safe_name: str
    bank_name: str
    login_url: str
    url_verified_date: Optional[date]
    username: str
    password: str
    account_last4: str
    statement_available_date: int  # day of month
    has_2fa: bool
    tfa_target: str  # "asc" or "client"
    tfa_method: str  # "email", "sms", "totp", "push"
    tfa_detail: str  # email addr, TOTP secret, or "dialpad"
    tfa_preference_order: list[str]  # ["1992", "2212"]
    tfa_sender: str  # auto-learned sender ID
    security_questions: list[tuple[str, str]]  # [(keyword, answer), ...]
    last_successful_login: Optional[date]
    status: str  # "active", "locked", "password_expired", "disabled"
    notes: str


@dataclass
class RunResult:
    """Result of a single statement download attempt."""

    timestamp: datetime
    client_name: str
    bank_name: str
    account_last4: str
    target_month: str  # "yyyy-mm"
    status: str  # "success", "failed", "skipped"
    filename: str
    notes: str
    ai_calls_count: int
    ai_cost_usd: float
    duration_seconds: int

    def to_row(self) -> list:
        """Convert to list for Excel append."""
        return [
            self.timestamp, self.client_name, self.bank_name,
            self.account_last4, self.target_month, self.status,
            self.filename, self.notes, self.ai_calls_count,
            self.ai_cost_usd, self.duration_seconds,
        ]


@dataclass
class RetryItem:
    """An entry in the retry queue."""

    timestamp: datetime
    client_name: str
    bank_name: str
    account_last4: str
    target_month: str
    failure_reason: str
    retry_count: int
    next_retry_date: date
    status: str  # "pending", "resolved", "abandoned"


# ---------------------------------------------------------------------------
# File locking
# ---------------------------------------------------------------------------
class FileLockError(Exception):
    """Raised when the workbook is locked by another process."""


def _lock_path(xlsx_path: Path) -> Path:
    return xlsx_path.with_suffix(".xlsx.lock")


def _acquire_lock(xlsx_path: Path) -> None:
    """Create a .lock file with the current PID.

    If a lock file exists and is stale (>1 hour old), remove it and proceed.
    If the lock is held by a live process, raise ``FileLockError``.
    """
    lock_file = _lock_path(xlsx_path)

    if lock_file.exists():
        lock_age = time.time() - lock_file.stat().st_mtime
        try:
            lock_pid = int(lock_file.read_text().strip())
        except (ValueError, OSError):
            lock_pid = -1

        if lock_age > LOCK_TIMEOUT_SECONDS:
            lock_file.unlink(missing_ok=True)
        elif lock_pid == os.getpid():
            # Re-entrant — same process already holds the lock.
            return
        else:
            raise FileLockError(
                f"clients.xlsx is locked by PID {lock_pid} "
                f"(age {lock_age:.0f}s). Remove {lock_file} if the process "
                "is no longer running."
            )

    lock_file.write_text(str(os.getpid()))


def _release_lock(xlsx_path: Path) -> None:
    """Remove the .lock file if it belongs to this process."""
    lock_file = _lock_path(xlsx_path)
    if lock_file.exists():
        try:
            lock_pid = int(lock_file.read_text().strip())
        except (ValueError, OSError):
            lock_pid = -1
        if lock_pid == os.getpid():
            lock_file.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Low-level workbook open/close
# ---------------------------------------------------------------------------
def open_workbook(path: Path) -> Workbook:
    """Open the workbook and acquire the file lock."""
    _acquire_lock(path)
    return load_workbook(path)


def close_workbook(wb: Workbook, path: Path) -> None:
    """Save, close the workbook and release the file lock."""
    try:
        wb.save(path)
        wb.close()
    finally:
        _release_lock(path)


# ---------------------------------------------------------------------------
# Context-manager wrapper
# ---------------------------------------------------------------------------
class ExcelManager:
    """Context manager for safe workbook access with file locking.

    Usage::

        with ExcelManager(CLIENTS_XLSX) as mgr:
            accounts = mgr.read_accounts()
            mgr.log_result(result)
    """

    def __init__(self, path: Path | None = None) -> None:
        self.path = Path(path) if path else CLIENTS_XLSX
        self._wb: Workbook | None = None

    # -- context protocol ---------------------------------------------------
    def __enter__(self) -> "ExcelManager":
        self._wb = open_workbook(self.path)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:  # noqa: ANN001
        if self._wb is not None:
            close_workbook(self._wb, self.path)
            self._wb = None

    # -- helpers ------------------------------------------------------------
    @property
    def wb(self) -> Workbook:
        if self._wb is None:
            raise RuntimeError("Workbook not open. Use ExcelManager as a context manager.")
        return self._wb

    def _sheet(self, name: str) -> Worksheet:
        if name not in self.wb.sheetnames:
            raise KeyError(f"Sheet '{name}' not found in workbook.")
        return self.wb[name]

    def _header_map(self, ws: Worksheet) -> dict[str, int]:
        """Return {column_name: col_index (1-based)} from the first row."""
        headers: dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip().lower()] = col_idx
        return headers

    # -- parsing helpers ----------------------------------------------------
    @staticmethod
    def _make_file_safe_name(client_name: str) -> str:
        """Generate a filesystem-safe name from the client name."""
        name = client_name.replace("'", "").replace("&", "and")
        name = re.sub(r"[^a-zA-Z0-9 _-]", "", name)
        name = re.sub(r"\s+", "_", name.strip())
        return name

    @staticmethod
    def _parse_bool(value: object) -> bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in ("y", "yes", "true", "1")

    @staticmethod
    def _parse_date(value: object) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _parse_preference_order(value: object) -> list[str]:
        if not value:
            return []
        return [s.strip() for s in str(value).split(",") if s.strip()]

    @staticmethod
    def _cell_str(value: object) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_last4(value: object) -> str:
        """Normalize account_last4: strip .0 from floats, remove # and non-alphanumeric."""
        if value is None:
            return ""
        s = str(value).strip()
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        s = s.lstrip("#").strip()
        return re.sub(r"[^a-zA-Z0-9]", "", s)

    @staticmethod
    def _normalize_tfa_config(has_2fa, tfa_method, tfa_detail, client_name, bank_name):
        """Validate and self-heal 2FA config. Returns (method, detail, changed)."""
        if not has_2fa:
            return tfa_method, tfa_detail, False
        method = tfa_method.lower().strip()
        detail = tfa_detail.strip()
        tag = f"[2FA] {client_name}/{bank_name}:"
        if not method:
            log.warning(f"{tag} 2fa=y but method is blank")
            return "", detail, False
        if method == "sms":
            digits_only = re.sub(r"\D", "", detail)
            if len(digits_only) >= 10:
                log.warning(f"{tag} phone number '{detail}' -> normalizing to 'dialpad'")
                return method, "dialpad", True
            if detail.lower() != "dialpad" and detail:
                log.warning(f"{tag} sms detail '{detail}' -- expected 'dialpad'")
        if method == "totp":
            clean = re.sub(r"[^A-Za-z2-7=]", "", detail)
            if len(clean) < 16:
                log.warning(f"{tag} totp detail '{detail}' doesn't look like base32")
        return method, detail, False

    @staticmethod
    def _cell_int(value: object, default: int = 0) -> int:
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default

    # -- public API ---------------------------------------------------------
    def read_accounts(self) -> list[AccountJob]:
        """Read all rows from the Accounts sheet into AccountJob objects."""
        ws = self._sheet(SHEET_ACCOUNTS)
        headers = self._header_map(ws)
        jobs: list[AccountJob] = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            values = {col_name: row[idx - 1].value for col_name, idx in headers.items()}

            client_name = self._cell_str(values.get("client_name"))
            if not client_name:
                continue  # skip empty rows

            raw_fsn = self._cell_str(values.get("file_safe_name"))
            file_safe_name = raw_fsn if raw_fsn else self._make_file_safe_name(client_name)

            # Security questions — up to 5 pairs
            sq: list[tuple[str, str]] = []
            for i in range(1, 6):
                kw = self._cell_str(values.get(f"sq{i}_keyword"))
                ans = self._cell_str(values.get(f"sq{i}_answer"))
                if kw and ans:
                    sq.append((kw, ans))

            last4 = self._normalize_last4(values.get("account_last4"))
            raw_method = self._cell_str(values.get("2fa_method"))
            raw_detail = self._cell_str(values.get("2fa_detail"))
            has_2fa = self._parse_bool(values.get("2fa"))
            bank_name = self._cell_str(values.get("bank_name"))
            norm_method, norm_detail, tfa_changed = self._normalize_tfa_config(
                has_2fa, raw_method, raw_detail, client_name, bank_name,
            )

            job = AccountJob(
                client_name=client_name,
                file_safe_name=file_safe_name,
                bank_name=bank_name,
                login_url=self._cell_str(values.get("login_url")),
                url_verified_date=self._parse_date(values.get("url_verified_date")),
                username=self._cell_str(values.get("username")),
                password=self._cell_str(values.get("password")),
                account_last4=last4,
                statement_available_date=self._cell_int(values.get("statement_available_date"), 1),
                has_2fa=has_2fa,
                tfa_target=self._cell_str(values.get("2fa_target")),
                tfa_method=norm_method,
                tfa_detail=norm_detail,
                tfa_preference_order=self._parse_preference_order(values.get("2fa_preference_order")),
                tfa_sender=self._cell_str(values.get("2fa_sender")),
                security_questions=sq,
                last_successful_login=self._parse_date(values.get("last_successful_login")),
                status=self._cell_str(values.get("status")) or "active",
                notes=self._cell_str(values.get("notes")),
            )
            jobs.append(job)
            if tfa_changed:
                if "2fa_detail" in headers:
                    row[headers["2fa_detail"] - 1].value = norm_detail
                    log.info(f"[2FA] {client_name}/{bank_name}: wrote normalized 2fa_detail back to Excel")
                else:
                    log.warning(f"[2FA] {client_name}/{bank_name}: 2fa_detail column missing, cannot persist normalization")
            if job.statement_available_date <= 1:
                log.warning(f"[Data] {client_name}/{bank_name} #{last4}: statement_available_date not set -- will auto-learn on first download")

        return jobs

    def log_result(self, result: RunResult) -> None:
        """Append a run result row to the Run Log sheet."""
        ws = self._sheet(SHEET_RUN_LOG)
        ws.append(result.to_row())

    def add_to_retry_queue(self, job: AccountJob, failure_reason: str) -> None:
        """Add a failed job to the Retry Queue with backoff scheduling."""
        ws = self._sheet(SHEET_RETRY_QUEUE)
        headers = self._header_map(ws)

        # Count existing retries for this account + month combo
        existing_count = 0
        target_month = datetime.now().strftime("%Y-%m")
        for row in ws.iter_rows(min_row=2, values_only=False):
            values = {col_name: row[idx - 1].value for col_name, idx in headers.items()}
            r_client = self._cell_str(values.get("client_name"))
            r_bank = self._cell_str(values.get("bank_name"))
            r_last4 = self._normalize_last4(values.get("account_last4"))
            r_month = self._cell_str(values.get("target_month"))
            if (r_client == job.client_name and r_bank == job.bank_name
                    and r_last4 == job.account_last4 and r_month == target_month):
                existing_count = max(existing_count, self._cell_int(values.get("retry_count")))

        retry_count = existing_count + 1
        status = "pending" if retry_count <= MAX_RETRY_COUNT else "abandoned"
        backoff_days = RETRY_BACKOFF.get(retry_count, max(RETRY_BACKOFF.values()))
        next_retry = date.today() + timedelta(days=backoff_days)

        ws.append([
            datetime.now(),
            job.client_name,
            job.bank_name,
            job.account_last4,
            target_month,
            failure_reason,
            retry_count,
            next_retry,
            status,
        ])

    def get_retry_queue(self) -> list[RetryItem]:
        """Read all pending retry items from the Retry Queue sheet."""
        ws = self._sheet(SHEET_RETRY_QUEUE)
        headers = self._header_map(ws)
        items: list[RetryItem] = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            values = {col_name: row[idx - 1].value for col_name, idx in headers.items()}
            status = self._cell_str(values.get("status"))
            if status != "pending":
                continue

            ts = values.get("timestamp")
            if isinstance(ts, datetime):
                timestamp = ts
            elif ts is not None:
                try:
                    timestamp = datetime.fromisoformat(str(ts))
                except ValueError:
                    timestamp = datetime.now()
            else:
                timestamp = datetime.now()

            items.append(RetryItem(
                timestamp=timestamp,
                client_name=self._cell_str(values.get("client_name")),
                bank_name=self._cell_str(values.get("bank_name")),
                account_last4=self._cell_str(values.get("account_last4")),
                target_month=self._cell_str(values.get("target_month")),
                failure_reason=self._cell_str(values.get("failure_reason")),
                retry_count=self._cell_int(values.get("retry_count")),
                next_retry_date=self._parse_date(values.get("next_retry_date")) or date.today(),
                status=status,
            ))

        return items

    def has_success(
        self, client: str, bank: str, last4: str, target_month: str,
    ) -> bool:
        """Check if Run Log already has a 'success' entry for this combo."""
        ws = self._sheet(SHEET_RUN_LOG)
        headers = self._header_map(ws)

        for row in ws.iter_rows(min_row=2, values_only=False):
            values = {col_name: row[idx - 1].value for col_name, idx in headers.items()}
            if (self._cell_str(values.get("client_name")) == client
                    and self._cell_str(values.get("bank_name")) == bank
                    and self._normalize_last4(values.get("account_last4")) == last4
                    and self._cell_str(values.get("target_month")) == target_month
                    and self._cell_str(values.get("status")) == "success"):
                return True
        return False

    def update_account_field(
        self, client: str, bank: str, last4: str, field_name: str, value: object,
    ) -> bool:
        """Update a single field in the Accounts sheet for a matching row.

        Supported fields: login_url, last_successful_login, status,
        2fa_sender, 2fa_method, 2fa_detail, url_verified_date,
        statement_available_date.

        Returns True if a matching row was found and updated.
        """
        allowed_fields = {
            "login_url", "last_successful_login", "status",
            "2fa_sender", "url_verified_date", "statement_available_date",
            "2fa_method", "2fa_detail",
        }
        if field_name not in allowed_fields:
            raise ValueError(
                f"Field '{field_name}' not in allowed update fields: {allowed_fields}"
            )

        ws = self._sheet(SHEET_ACCOUNTS)
        headers = self._header_map(ws)

        if field_name not in headers:
            raise KeyError(f"Column '{field_name}' not found in Accounts sheet.")

        target_col = headers[field_name]
        client_col = headers.get("client_name")
        bank_col = headers.get("bank_name")
        last4_col = headers.get("account_last4")

        if not all([client_col, bank_col, last4_col]):
            raise KeyError("Missing required lookup columns in Accounts sheet.")

        for row in ws.iter_rows(min_row=2):
            r_client = self._cell_str(row[client_col - 1].value)
            r_bank = self._cell_str(row[bank_col - 1].value)
            r_last4 = self._normalize_last4(row[last4_col - 1].value)

            if r_client == client and r_bank == bank and r_last4 == last4:
                row[target_col - 1].value = value
                return True

        return False


# ---------------------------------------------------------------------------
# Template workbook generator
# ---------------------------------------------------------------------------
def create_template_workbook(output_path: Path) -> None:
    """Create a blank clients.xlsx template with headers, validation, and
    one example row on the Accounts sheet."""
    wb = Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)

    def _style_headers(ws: Worksheet, columns: list[str]) -> None:
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            # Auto-width estimate
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 14)

    # -- Accounts sheet -----------------------------------------------------
    ws_accounts = wb.active
    ws_accounts.title = SHEET_ACCOUNTS
    _style_headers(ws_accounts, ACCOUNTS_COLUMNS)

    # Data validations (applied to rows 2-500)
    dv_2fa = DataValidation(type="list", formula1='"y,n"', allow_blank=True)
    dv_2fa.error = "Enter y or n"
    dv_2fa.errorTitle = "Invalid 2FA value"
    ws_accounts.add_data_validation(dv_2fa)

    dv_target = DataValidation(type="list", formula1='"asc,client"', allow_blank=True)
    dv_target.error = "Enter asc or client"
    dv_target.errorTitle = "Invalid 2FA target"
    ws_accounts.add_data_validation(dv_target)

    dv_method = DataValidation(type="list", formula1='"email,sms,totp,push"', allow_blank=True)
    dv_method.error = "Enter email, sms, totp, or push"
    dv_method.errorTitle = "Invalid 2FA method"
    ws_accounts.add_data_validation(dv_method)

    dv_status = DataValidation(
        type="list", formula1='"active,locked,password_expired,disabled"', allow_blank=True,
    )
    dv_status.error = "Enter active, locked, password_expired, or disabled"
    dv_status.errorTitle = "Invalid status"
    ws_accounts.add_data_validation(dv_status)

    # Map column names to 1-based indices
    col_map = {name: idx for idx, name in enumerate(ACCOUNTS_COLUMNS, start=1)}

    # Apply validations to column ranges
    last_row = 500
    dv_2fa.add(f"{get_column_letter(col_map['2fa'])}2:{get_column_letter(col_map['2fa'])}{last_row}")
    dv_target.add(f"{get_column_letter(col_map['2fa_target'])}2:{get_column_letter(col_map['2fa_target'])}{last_row}")
    dv_method.add(f"{get_column_letter(col_map['2fa_method'])}2:{get_column_letter(col_map['2fa_method'])}{last_row}")
    dv_status.add(f"{get_column_letter(col_map['status'])}2:{get_column_letter(col_map['status'])}{last_row}")

    # Example row
    example = {
        "client_name": "Acme Corp",
        "file_safe_name": "Acme_Corp",
        "bank_name": "Chase Business",
        "login_url": "https://www.chase.com/business/login",
        "url_verified_date": date(2026, 3, 1),
        "username": "acme_user",
        "password": "s3cure!Pass",
        "account_last4": "7890",
        "statement_available_date": 3,
        "2fa": "y",
        "2fa_target": "asc",
        "2fa_method": "sms",
        "2fa_detail": "+13035551234",
        "2fa_preference_order": "1992,2212",
        "2fa_sender": "",
        "sq1_keyword": "pet",
        "sq1_answer": "Rover",
        "sq2_keyword": "city",
        "sq2_answer": "Denver",
        "sq3_keyword": "",
        "sq3_answer": "",
        "sq4_keyword": "",
        "sq4_answer": "",
        "sq5_keyword": "",
        "sq5_answer": "",
        "last_successful_login": None,
        "status": "active",
        "notes": "Example row — delete before use",
    }
    for col_name in ACCOUNTS_COLUMNS:
        ws_accounts.cell(row=2, column=col_map[col_name], value=example.get(col_name, ""))

    # Freeze top row
    ws_accounts.freeze_panes = "A2"

    # -- Run Log sheet ------------------------------------------------------
    ws_log = wb.create_sheet(SHEET_RUN_LOG)
    _style_headers(ws_log, RUN_LOG_COLUMNS)
    ws_log.freeze_panes = "A2"

    # -- Retry Queue sheet --------------------------------------------------
    ws_retry = wb.create_sheet(SHEET_RETRY_QUEUE)
    _style_headers(ws_retry, RETRY_QUEUE_COLUMNS)
    ws_retry.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()
    print(f"Template workbook created: {output_path}")


# ---------------------------------------------------------------------------
# CLI entry point — generate template
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    create_template_workbook(CLIENTS_XLSX)
