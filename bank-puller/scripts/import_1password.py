"""Import bank credentials from 1Password CSV export into clients.xlsx.

Usage:
    python scripts/import_1password.py --csv PATH --dry-run   # preview changes
    python scripts/import_1password.py --csv PATH --apply      # apply changes
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Resolve project root so we can import from bank-puller
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from config import CLIENTS_XLSX  # noqa: E402
from orchestrator.excel_reader import ACCOUNTS_COLUMNS, ExcelManager  # noqa: E402

# ---------------------------------------------------------------------------
# Bank URL → bank_name mapping
# ---------------------------------------------------------------------------
BANK_DOMAIN_MAP: list[tuple[str, str]] = [
    # Order matters — first match wins. More specific patterns first.
    ("secure01a.chase.com", "Chase Business"),
    ("secure02ea.chase.com", "Chase Business"),
    ("secure03ea.chase.com", "Chase Business"),
    ("secure05b.chase.com", "Chase Business"),
    ("secure05c.chase.com", "Chase Business"),
    ("secure05ea.chase.com", "Chase Business"),
    ("secure.chase.com", "Chase Business"),
    ("chase.com", "Chase Business"),
    ("secmail.bankofamerica.com", "Bank of America"),
    ("secure.bankofamerica.com", "Bank of America"),
    ("bankofamerica.com", "Bank of America"),
    ("connect.secure.wellsfargo.com", "Wells Fargo"),
    ("wellsoffice.ceo.wellsfargo.com", "Wells Fargo"),
    ("wellsfargo.com", "Wells Fargo"),
    ("businessaccess.citibank.citigroup.com", "Citibank"),
    ("citiretailservices.citibankonline.com", "Citibank"),
    ("online.citi.com", "Citibank"),
    ("citi.com", "Citibank"),
    ("americanexpress.com", "American Express"),
    ("myaccounts.capitalone.com", "Capital One"),
    ("verified.capitalone.com", "Capital One"),
    ("capitalone.com", "Capital One"),
    ("login.morganstanleyclientserv.com", "Morgan Stanley"),
    ("morganstanley.loanadministration.com", "Morgan Stanley"),
    ("psrs.morganstanley.com", "Morgan Stanley"),
    ("morganstanleyclientserv.com", "Morgan Stanley"),
    ("morganstanley.com", "Morgan Stanley"),
    ("onlineservices.ubs.com", "UBS"),
    ("eastwest.bankonline.com", "Eastwest Bank"),
    ("app.mercury.com", "Mercury"),
    ("app.relayfi.com", "Relay"),
    ("ibx.key.com", "KeyBank"),
    ("key.com", "KeyBank"),
    ("olb-ebanking.com", "American Riviera Bank"),
    ("secure.americafirst.com", "America First CU"),
    ("onlinebanking.avidbank.com", "AvidBank"),
    ("columbiabank.ebanking-services.com", "Pacific Premier Bank"),
    ("1cbank.ebanking-services.com", "1st Century Bank"),
    ("secure.online.boh.com", "Bank of Hawaii"),
    ("cbc.comerica.com", "Comerica"),
    ("onlinebanking.syb.com", "Stock Yards Bank"),
    ("secure.rbcwm-usa.com", "RBC Wealth Management"),
    ("usbank.com", "US Bank"),
    ("secureonline.pnc.com", "PNC Bank"),
    ("servicecu.org", "Service Credit Union"),
    ("secure.myvirtualbranch.com", "Sound Federal CU"),
    ("citizensbankonline.com", "Citizens Bank"),
    ("home.penfed.org", "PenFed"),
    ("olui2.fs.ml.com", "Merrill Lynch"),
    ("clientaccess.rjf.com", "Raymond James"),
    ("accountview.lpl.com", "LPL Financial"),
    ("secure.paymentech.com", "Chase Paymentech"),
    ("web17.secureinternetbank.com", "KeyBank"),
    ("login.capchase.com", "Capchase"),
]

# Non-bank domains to skip
SKIP_DOMAINS: set[str] = {
    "keepersecurity.com", "www.keeper.app", "qbo.intuit.com", "accounts.intuit.com",
    "payroll.intuit.com", "www.eftps.gov", "onlineservices.cdtfa.ca.gov",
    "apps.sandiego.gov", "id.northwestregisteredagent.com",
    "www.northwestregisteredagent.com", "ofcpa.xyz", "www.amazon.com",
    "sellercentral.amazon.com", "www.facebook.com", "www.instagram.com",
    "instagram.com", "twitter.com", "slack.com", "discord.com", "trello.com",
    "zoom.us", "www.grammarly.com", "www.notion.so", "bsky.app",
    "onlyfans.com", "fansly.com", "www.reddit.com", "venmo.com",
    "www.paypal.com", "squareup.com", "dashboard.stripe.com",
    "app.hubspot.com", "app.clickup.com", "www.loom.com", "www.make.com",
    "linktr.ee", "carrd.co", "sso.godaddy.com", "www.namecheap.com",
    "auth.hostinger.com", "accounts.google.com", "google.com",
    "idmsa.apple.com", "appstoreconnect.apple.com", "identity.att.com",
    "www.cox.com", "geico.com", "www.blueshieldca.com", "guardianlife.com",
    "www.healthcare.gov", "www.discover.com", "www.barclaycardus.com",
    "app.gusto.com", "login.gusto.com", "online.adp.com", "runpayroll.adp.com",
    "rippling.com", "login.justworks.com", "www.justworks.com",
    "services.primepay.com", "login.flex.paychex.com", "benefits.paychex.com",
    "app.bill.com", "app02.us.bill.com", "cdn.spend.bill.com",
    "app.meliopayments.com", "settle.com", "app.settle.co",
    "my.waveapps.com", "go.xero.com", "www.liveflow.com",
    "transactions.saasant.com", "auth.numeric.io", "auth.subscript.com",
    "app.contentsnare.com", "connect.qualia.com", "dialpad.com",
    "app.pandadoc.com", "login.huddle.com", "login.electric.ai",
    "login.zenbusiness.com", "www.legalzoom.com", "filedba.com",
    "www.parasec.com", "www.businesslicenses.com", "www.nvsilverflume.gov",
    "www.myfloridalicense.com", "ecorp.azcc.gov", "idm.sos.ca.gov",
    "cofs.lara.state.mi.us", "ccfs.sos.wa.gov", "onestop.delaware.gov",
    "onestop.portal.ky.gov", "secure.login.gov",
    # Tax/government
    "webapp.ftb.ca.gov", "www.aztaxes.gov", "aztaxes.gov",
    "mytax.illinois.gov", "mytax.mo.gov", "www.myvtax.vermont.gov",
    "www7b.tax.ny.gov", "my.ny.gov", "mypath.pa.gov", "uctax.pa.gov",
    "drs.ct.gov", "security.app.cpa.state.tx.us", "oktap.tax.ok.gov",
    "idahotap.gentax.com", "mdtaxconnect.gov", "revenue.maine.gov",
    "hitax.hawaii.gov", "login.ehawaii.gov", "mynvtax.nv.gov",
    "www.nevadatax.nv.gov", "gtc.dor.ga.gov", "www.lasvegasnevada.gov",
    "dor.wa.gov", "secure.dor.wa.gov", "secureaccess.wa.gov",
    "mytaxes.sbtaxes.org", "thousandoaks.hdlgov.com",
    "shelbycountyky.taxandrevenue.opengov.com",
    # Unemployment/labor
    "edd.ca.gov", "eddservices.edd.ca.gov", "myedd.edd.ca.gov",
    "apps.twc.texas.gov", "secure.dol.state.nj.us", "www1.state.nj.us",
    "nui.nv.gov", "eresponse.gdol.ga.gov", "labor.alabama.gov",
    "huiclaims3.hawaii.gov", "secure.lni.wa.gov", "employer.calsavers.com",
    # Insurance/utilities/misc
    "cp.mercuryinsurance.com", "mybusiness.libertymutual.com",
    "account.thehartford.com", "cs-eb.pacificguardian.com",
    "www.surety-cloud.com", "www.securitymetrics.com",
    "www.myprepaidcenter.com", "www.expresstoll.com", "e-zpassny.com",
    "www.dmv.ca.gov", "www.fedex.com", "www.pitneybowes.us",
    "ipostal1.com", "www.oru.com", "duquesnelight.com",
    "columbiagaspa.com", "www.sce.com", "myaccount.directenergy.com",
    "myaccount.socalgas.com", "utilitiesonline.santabarbaraca.gov",
    "santabarbara.watersmart.com", "santabarbara.bizlicenseonline.com",
    "slcgov.my.site.com", "engage.huntingtonbeachca.gov",
    "cityofroswellga-energovweb.tylerhost.net",
    "ct.wolterskluwer.com", "www.1099online.com",
    # SaaS/industry-specific
    "login.bigcommerce.com", "avalara.com", "avalara-app.shopifysvc.com",
    "identity.avalara.com", "accounts.shopify.com",
    "www.braintreegateway.com", "login.authorize.net",
    "dashboard.maverickpayments.com", "dashboard.payarc.net",
    "dashboard.sezzle.com", "cardpointe.cardconnect.com",
    "signup.payoneer.com", "wise.com",
    "app.ramp.com", "dashboard.brex.com", "app.joinarc.com",
    "login.app.carta.com", "www.realpage.com",
    "signin.mindbodyonline.com", "fama.thoughtspot.cloud",
    "identity.thoughtspotlogin.cloud", "looker.fama.io",
    "artist.atvenu.com", "venue.atvenu.com",
    "portal.fuel.me", "erp.mojaverepeater.com",
    "www.healthwavehq.com", "member.everbridge.net",
    "accounts.timeclockwizard.com", "accounts.toggl.com",
    "signin.inspirato.com", "www.servicechannel.com",
    "www.sponsorinsight.com", "mydbr.socialintel.com",
    "app.dripify.io", "app.doublehq.com", "app.kajabi.com",
    "tmr.mcasuite.com", "app.conversionpro.io",
    "access.enom.com", "support.enom.com",
    "accounts.sanity.io", "app.surferseo.com",
    "alpinetax.taxdome.com", "tagpro.liscio.me",
    "lending.sba.gov", "id-provider.tco.census.gov",
    "registration.tco.census.gov",
    # Loan admin / special
    "loanview.berkadia.com", "ubs.loanadministration.com",
    "morganstanley.loanadministration.com",
    # Ariba/supplier portals
    "mu.ariba.com", "service.ariba.com", "portal.us.bn.cloud.ariba.com",
    "buyer.na1prd.taulia.com", "supplier.coupahost.com",
    "suppliers.tipalti.com",
    # Payroll/HR platforms
    "coadquantum.coadvantage.com", "coadazureprod.b2clogin.com",
    "login-esdi-saasfaprod1.fa.ocs.oraclecloud.com",
    "payroll.razorpay.com",
    # Document/reporting
    "dzhphillips.sharefile.com", "ereports.merchantfactors.com",
    "ereports.merchantfinancial.com", "www.depositorcontrol.com",
    "portal2.ftnirdc.com", "uwp22.duckcreekondemand.com",
    "ux-ctadmin.ctadvantage.com", "opus.olbanking.com",
    "www.specialfinancingco.com", "www.spservicing.com",
    "web1.zixmail.net", "api.id.me",
    "login.capchase.com",
    "awalkonwater.org",
}


# ---------------------------------------------------------------------------
# Dataclass for parsed 1Password entry
# ---------------------------------------------------------------------------
@dataclass
class ParsedEntry:
    """A single parsed 1Password row ready for Excel insertion."""
    client_name: str = ""
    file_safe_name: str = ""
    bank_name: str = ""
    login_url: str = ""
    username: str = ""
    password: str = ""
    account_last4: str = "???"
    has_2fa: bool = False
    tfa_target: str = ""
    tfa_method: str = ""
    tfa_detail: str = ""
    status: str = "active"
    notes_parts: list[str] = field(default_factory=list)
    security_questions: list[tuple[str, str]] = field(default_factory=list)
    # Metadata for dedup
    title: str = ""
    archived: bool = False
    has_use_this: bool = False
    is_client_login: bool = False
    is_do_not_use: bool = False
    is_asc_username: bool = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def classify_bank(url: str) -> str | None:
    """Return bank_name if the URL matches a known bank domain, else None."""
    if not url:
        return None
    try:
        netloc = urlparse(url).netloc.lower()
    except Exception:
        return None
    if not netloc:
        return None

    # Skip non-bank domains
    if netloc in SKIP_DOMAINS:
        return None

    for pattern, bank_name in BANK_DOMAIN_MAP:
        if pattern in netloc:
            return bank_name
    return None


def make_file_safe_name(client_name: str) -> str:
    """Generate a filesystem-safe name from the client name."""
    name = client_name.replace("'", "").replace("&", "and")
    name = re.sub(r"[^a-zA-Z0-9 _-]", "", name)
    name = re.sub(r"\s+", "_", name.strip())
    return name


def extract_client_name(title: str, bank_name: str) -> str:
    """Extract client name from 1Password title.

    Patterns:
      "BankName - ClientName"
      "BankName - ClientName (xxxx)"
      "ClientName bankname.com"
    """
    # Clean title of markers
    clean = title.strip()
    for marker in ["(USE THIS)", "(DO NOT USE)", "(CLIENT LOGIN)",
                    "CLIENT LOGIN", "LOGIN DOES NOT WORK",
                    "LOGIN DOES NOT CURRENTLY WORK", "ACCOUNT CLOSED",
                    "[USE THIS]", "[ACCOUNT CLOSED]"]:
        clean = clean.replace(marker, "").strip()

    # Remove trailing account numbers like (6752, 0592, 3811) or (5911)
    clean = re.sub(r"\s*\([0-9, ]+\)\s*$", "", clean).strip()

    # Try "Bank - Client" pattern
    if " - " in clean:
        parts = clean.split(" - ", 1)
        client = parts[1].strip()
        if client:
            return client

    # Try "Client bankdomain.com" pattern
    domain_patterns = [
        "chase.com", "bankofamerica.com", "wellsfargo.com", "citi.com",
        "bofa", "wells fargo", "wellsfargo", "penfed.org",
    ]
    lower = clean.lower()
    for dp in domain_patterns:
        idx = lower.find(dp)
        if idx > 0:
            candidate = clean[:idx].strip()
            if candidate:
                return candidate

    # Fallback: use the whole cleaned title
    return clean


# Security question keyword normalization
SQ_KEYWORD_MAP = {
    "pet": "pet", "first pet": "pet",
    "street": "street",
    "school": "school", "high school": "school",
    "friend": "friend", "best friend": "friend", "boyfriend": "friend",
    "bestfriend's last name": "friend",
    "city": "city", "in which city": "city",
    "model car": "car", "car": "car",
    "game": "game",
    "boss": "boss", "boss name": "boss",
    "company": "company",
    "subject": "subject",
    "uncle": "uncle",
    "first pay": "first_pay",
    "age": "age",
    "girlfriend": "girlfriend",
    "mother": "mother", "mother's maiden name": "mother",
    "vacation": "vacation",
    "sibling": "sibling",
    "manager": "manager",
    "movie": "movie",
    "hero": "hero",
    "teacher's name": "teacher",
    "teacher": "teacher",
    "prom": "prom",
    "composer": "composer",
    "sports movie": "sports_movie",
    "number": "number",
    "cartoon": "cartoon",
    "graduated hs": "grad_hs",
    "graduated college": "grad_college",
}


def parse_notes(notes: str, title: str) -> dict:
    """Parse 1Password notes field into structured data.

    Returns dict with keys:
        security_code, company_id, client_id, business_code,
        accounts (list of last4), has_2fa, tfa_target, tfa_method, tfa_detail,
        security_questions (list of (keyword, answer)),
        notes_parts (list of extra note strings),
        status_override
    """
    result = {
        "security_code": "",
        "company_id": "",
        "client_id": "",
        "business_code": "",
        "accounts": [],
        "has_2fa": False,
        "tfa_target": "",
        "tfa_method": "",
        "tfa_detail": "",
        "security_questions": [],
        "notes_parts": [],
        "status_override": "",
    }

    if not notes:
        return result

    title_upper = title.upper()

    # Check title for status markers
    if "LOGIN DOES NOT WORK" in title_upper or "DO NOT USE" in title_upper:
        result["status_override"] = "disabled"
    if "CLOSED" in title_upper or "ACCOUNT CLOSED" in title_upper:
        result["status_override"] = "disabled"
    if "CLIENT LOGIN" in title_upper:
        result["tfa_target"] = "client"

    lines = notes.replace("\r\n", "\n").split("\n")
    # Also split on | which 1Password uses as line separator in CSV
    expanded: list[str] = []
    for line in lines:
        expanded.extend(line.split("|"))

    for raw_line in expanded:
        line = raw_line.strip()
        if not line or line.startswith("--") or line.startswith("----"):
            continue
        lower = line.lower()

        # Security code
        m = re.match(r"(?:security\s+)?code[:\s]+(.+)", line, re.IGNORECASE)
        if m and "business code" not in lower and "token" not in lower:
            val = m.group(1).strip()
            # Skip if it's a 2FA-related code reference
            if val.lower() in ("client's phone", "client's chase app",
                               "client's phone/email", "matt's phone",
                               "matt's phone or matt's email",
                               "client's phone or matt's email"):
                # This is a 2FA target indicator
                if "client" in val.lower():
                    result["tfa_target"] = "client"
                    result["has_2fa"] = True
                else:
                    result["tfa_target"] = "asc"
                    result["has_2fa"] = True
                    result["tfa_method"] = "sms"
                continue
            result["security_code"] = val
            continue

        # Company ID
        m = re.match(r"company\s*id[:\s]+(.+)", line, re.IGNORECASE)
        if m:
            result["company_id"] = m.group(1).strip()
            continue

        # Client ID
        m = re.match(r"client\s*id[:\s]+(.+)", line, re.IGNORECASE)
        if m:
            result["client_id"] = m.group(1).strip()
            continue

        # Business Code
        m = re.match(r"business\s*code[:\s]+(.+)", line, re.IGNORECASE)
        if m:
            result["business_code"] = m.group(1).strip()
            continue

        # Accounts (multi or single)
        m = re.match(r"accounts?[:\s]+(.+)", line, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            # Extract 4-digit numbers
            accts = re.findall(r"\b(\d{4})\b", val)
            if accts:
                result["accounts"].extend(accts)
                # Check if any are marked closed
                if "closed" in val.lower():
                    result["notes_parts"].append(f"account_note={val}")
            continue

        # 2FA indicators
        if re.search(r"requires?\s+2fa", lower) or "2fa" in lower:
            result["has_2fa"] = True
            if "token" in lower or "authenticator" in lower:
                result["tfa_method"] = "totp"
            if "dialpad" in lower:
                result["tfa_target"] = "asc"
                result["tfa_method"] = "sms"
                result["tfa_detail"] = "dialpad"
                # Extract dialpad number if present
                dm = re.search(r"dialpad\s+(\d{4})", lower)
                if dm:
                    result["tfa_detail"] = f"dialpad {dm.group(1)}"
            if "austin" in lower:
                result["tfa_target"] = "asc"
            if "client" in lower:
                result["tfa_target"] = "client"
            continue

        if "requires token" in lower or "use token" in lower or "requires 2fa token" in lower:
            result["has_2fa"] = True
            result["tfa_method"] = "totp"
            result["tfa_target"] = "asc"
            continue

        if "send" in lower and "2fa" in lower and "dialpad" in lower:
            result["has_2fa"] = True
            result["tfa_target"] = "asc"
            result["tfa_method"] = "sms"
            dm = re.search(r"dialpad\s+(\d{4})", lower)
            result["tfa_detail"] = f"dialpad {dm.group(1)}" if dm else "dialpad"
            continue

        if re.search(r"send.*code.*dialpad", lower) or re.search(r"2fa.*dialpad", lower):
            result["has_2fa"] = True
            result["tfa_target"] = "asc"
            result["tfa_method"] = "sms"
            dm = re.search(r"dialpad\s+(\d{4})", lower)
            result["tfa_detail"] = f"dialpad {dm.group(1)}" if dm else "dialpad"
            continue

        if "code: client's phone" in lower or "code: client" in lower:
            result["has_2fa"] = True
            result["tfa_target"] = "client"
            continue

        if "code to matt's phone" in lower or "code: matt's phone" in lower:
            result["has_2fa"] = True
            result["tfa_target"] = "asc"
            result["tfa_method"] = "sms"
            continue

        if "client's chase app" in lower or "client's amex" in lower:
            result["has_2fa"] = True
            result["tfa_target"] = "client"
            result["tfa_method"] = "push"
            continue

        if "verification from the client" in lower:
            result["has_2fa"] = True
            result["tfa_target"] = "client"
            result["tfa_method"] = "push"
            continue

        # Security questions: "key: value" or "key = value" patterns
        # Match patterns like "pet: zoe", "first pay: 18", "street: whale"
        sq_match = re.match(r"([a-z][a-z '\-]+?)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if sq_match:
            raw_key = sq_match.group(1).strip().lower()
            raw_val = sq_match.group(2).strip()
            # Check if this looks like a security question
            normalized_key = SQ_KEYWORD_MAP.get(raw_key)
            if normalized_key:
                result["security_questions"].append((normalized_key, raw_val))
                continue

        # Check for numbered security questions: "1 - In which city..."
        sq_numbered = re.match(r"\d+\s*[-–]\s*(.+?)\?\s*(.+)", line)
        if sq_numbered:
            q = sq_numbered.group(1).strip().lower()
            a = sq_numbered.group(2).strip()
            # Try to extract a keyword
            for kw_pattern, kw_norm in SQ_KEYWORD_MAP.items():
                if kw_pattern in q:
                    result["security_questions"].append((kw_norm, a))
                    break
            else:
                result["security_questions"].append((q[:20], a))
            continue

        # Specific patterns in the data
        if lower.startswith("social "):
            result["notes_parts"].append(f"ssn_last4={line.split()[-1]}")
            continue

        if lower.startswith("only track"):
            m = re.search(r"#?(\d{4})", line)
            if m:
                result["accounts"] = [m.group(1)]
            continue

        # Skip noise lines
        noise_patterns = [
            "questions:", "security questions:", "note:", "notes:",
            "log in not working", "login not working", "error login",
            "---", "----", "clients:", "this access includes",
            "this is for", "all of", "the full account",
            "will prompt", "when asked for",
        ]
        if any(p in lower for p in noise_patterns):
            # Some of these have useful info, capture selectively
            if "this is for" in lower or "all of" in lower:
                result["notes_parts"].append(line)
            continue

        # Credit card info — skip (CVV, exp dates)
        if re.search(r"\b\d{13,16}\b", line) or re.search(r"cvv\s+\d{3}", lower):
            continue

        # Remaining non-empty lines go to notes
        if len(line) > 2 and not line.startswith("--"):
            # Skip email lines and keeper references
            if "keeper security" in lower or "email:" in lower:
                continue
            result["notes_parts"].append(line)

    return result


def is_bank_url(url: str) -> bool:
    """Check if URL matches any known bank domain."""
    return classify_bank(url) is not None


def is_asc_username(username: str) -> bool:
    """Check if username follows ASC/OFCPA/Austin naming pattern."""
    u = username.lower()
    return any(u.startswith(p) for p in [
        "asc", "ofcpa", "austin", "matt", "pm@allsolutions",
        "pm@compound", "austin@", "matt@allsolutions",
    ])


# ---------------------------------------------------------------------------
# Main import logic
# ---------------------------------------------------------------------------
def read_1password_csv(csv_path: Path) -> list[dict]:
    """Read the 1Password CSV export."""
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


def parse_entries(rows: list[dict]) -> list[ParsedEntry]:
    """Parse 1Password rows into ParsedEntry objects, filtering to bank entries only."""
    entries: list[ParsedEntry] = []

    for row in rows:
        title = row.get("Title", "").strip()
        url = row.get("Url", "").strip()
        username = row.get("Username", "").strip()
        password = row.get("Password", "").strip()
        otp = row.get("OTPAuth", "").strip()
        archived = row.get("Archived", "").strip().lower() == "true"
        notes_raw = row.get("Notes", "").strip()

        # Classify bank
        bank_name = classify_bank(url)
        if not bank_name:
            continue

        # Parse notes
        parsed = parse_notes(notes_raw, title)

        # Extract client name
        client_name = extract_client_name(title, bank_name)

        # Title markers
        title_upper = title.upper()
        has_use_this = "(USE THIS)" in title_upper or "[USE THIS]" in title_upper
        is_client_login = "CLIENT LOGIN" in title_upper
        is_do_not_use = "DO NOT USE" in title_upper

        # Determine status
        status = "active"
        if parsed["status_override"]:
            status = parsed["status_override"]
        elif is_do_not_use:
            status = "disabled"

        # Build notes field
        notes_parts: list[str] = []
        if parsed["security_code"]:
            notes_parts.append(f"security_code={parsed['security_code']}")
        if parsed["company_id"]:
            notes_parts.append(f"company_id={parsed['company_id']}")
        if parsed["client_id"]:
            notes_parts.append(f"client_id={parsed['client_id']}")
        if parsed["business_code"]:
            notes_parts.append(f"business_code={parsed['business_code']}")
        notes_parts.extend(parsed["notes_parts"])

        # 2FA from OTPAuth field
        has_2fa = parsed["has_2fa"]
        tfa_method = parsed["tfa_method"]
        tfa_target = parsed["tfa_target"]
        tfa_detail = parsed["tfa_detail"]
        if otp:
            has_2fa = True
            tfa_method = tfa_method or "totp"
            tfa_target = tfa_target or "asc"

        # Chase ALWAYS requires 2FA on fresh browser sessions — set defaults
        # so that Dialpad is launched even if 1Password has no OTPAuth field.
        if bank_name == "Chase Business":
            has_2fa = True
            tfa_target = tfa_target or "asc"
            tfa_method = tfa_method or "sms"
            tfa_detail = tfa_detail or "dialpad"

        # If client login, set target
        if is_client_login:
            tfa_target = "client"

        # Determine accounts
        accounts = parsed["accounts"] if parsed["accounts"] else ["???"]

        # Create one entry per account
        for acct in accounts:
            entry = ParsedEntry(
                client_name=client_name,
                file_safe_name=make_file_safe_name(client_name),
                bank_name=bank_name,
                login_url=url,
                username=username,
                password=password,
                account_last4=acct,
                has_2fa=has_2fa,
                tfa_target=tfa_target,
                tfa_method=tfa_method,
                tfa_detail=tfa_detail,
                status=status,
                notes_parts=notes_parts.copy(),
                security_questions=parsed["security_questions"][:5],
                title=title,
                archived=archived,
                has_use_this=has_use_this,
                is_client_login=is_client_login,
                is_do_not_use=is_do_not_use,
                is_asc_username=is_asc_username(username),
            )
            entries.append(entry)

    return entries


def dedup_entries(entries: list[ParsedEntry]) -> list[ParsedEntry]:
    """Apply deduplication strategy.

    Pass 1: For each (bank, client), if a non-archived entry exists, drop
            archived entries with different usernames (they're old credentials).
            Keep archived entries that are CLIENT LOGINs (different user = separate row).
    Pass 2: Within (bank, client, username, account) groups, pick the best entry.
    """
    # Pass 1: Drop archived entries when non-archived exists for same bank+client
    by_bank_client: dict[tuple[str, str], list[ParsedEntry]] = {}
    for e in entries:
        key = (e.bank_name.lower(), e.client_name.lower())
        by_bank_client.setdefault(key, []).append(e)

    filtered: list[ParsedEntry] = []
    for key, group in by_bank_client.items():
        has_non_archived = any(not e.archived for e in group)
        for e in group:
            if e.archived and has_non_archived and not e.is_client_login:
                # Drop archived entry — a newer non-archived version exists
                continue
            filtered.append(e)

    # Pass 2: Dedup within (bank, client, username, account)
    groups: dict[tuple, list[ParsedEntry]] = {}
    for e in filtered:
        key = (e.bank_name.lower(), e.client_name.lower(), e.username.lower(), e.account_last4)
        groups.setdefault(key, []).append(e)

    result: list[ParsedEntry] = []
    for key, group in groups.items():
        if len(group) == 1:
            result.append(group[0])
            continue

        # Sort by priority: non-archived > USE THIS > ASC username > not DO NOT USE
        def sort_key(e: ParsedEntry) -> tuple:
            return (
                not e.archived,       # non-archived first
                e.has_use_this,       # USE THIS first
                e.is_asc_username,    # ASC usernames preferred
                not e.is_do_not_use,  # not DO NOT USE first
                len(e.password),      # longer password = more likely current
            )

        group.sort(key=sort_key, reverse=True)
        best = group[0]

        # Merge notes and security questions from other entries
        seen_notes = set(best.notes_parts)
        seen_sq = set(best.security_questions)
        for other in group[1:]:
            for n in other.notes_parts:
                if n not in seen_notes:
                    best.notes_parts.append(n)
                    seen_notes.add(n)
            for sq in other.security_questions:
                if sq not in seen_sq:
                    best.security_questions.append(sq)
                    seen_sq.add(sq)
            # If one of the others is a client login and best is not, keep client login too
            if other.is_client_login and not best.is_client_login:
                client_entry = ParsedEntry(
                    client_name=other.client_name,
                    file_safe_name=other.file_safe_name,
                    bank_name=other.bank_name,
                    login_url=other.login_url,
                    username=other.username,
                    password=other.password,
                    account_last4=other.account_last4,
                    has_2fa=other.has_2fa,
                    tfa_target="client",
                    tfa_method=other.tfa_method,
                    tfa_detail=other.tfa_detail,
                    status=other.status,
                    notes_parts=other.notes_parts.copy(),
                    security_questions=other.security_questions[:5],
                    title=other.title,
                    archived=other.archived,
                    is_client_login=True,
                    is_asc_username=other.is_asc_username,
                )
                result.append(client_entry)

        result.append(best)

    return result


# ---------------------------------------------------------------------------
# Excel read/write
# ---------------------------------------------------------------------------
def read_existing_accounts(wb: Workbook) -> list[dict]:
    """Read existing Accounts rows as dicts keyed by column name."""
    ws = wb["Accounts"]
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = col_idx

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = {}
        for col_name, col_idx in headers.items():
            values[col_name] = row[col_idx - 1].value
        # Skip empty rows
        if not values.get("client_name"):
            continue
        values["_row_num"] = row[0].row
        rows.append(values)

    return rows


def match_existing(entry: ParsedEntry, existing: list[dict]) -> dict | None:
    """Find an existing row matching by (username, bank_name) or (username, bank_name, account_last4)."""
    for row in existing:
        row_username = str(row.get("username", "") or "").strip().lower()
        row_bank = str(row.get("bank_name", "") or "").strip().lower()
        row_last4 = str(row.get("account_last4", "") or "").strip()

        if (row_username == entry.username.lower() and
                row_bank == entry.bank_name.lower()):
            # If both have specific account numbers, they must match
            if (entry.account_last4 != "???" and row_last4 != "???" and
                    row_last4 and entry.account_last4 != row_last4):
                continue
            return row
    return None


@dataclass
class DiffEntry:
    action: str  # "add", "update", "skip"
    client_name: str
    bank_name: str
    username: str
    account_last4: str
    reason: str = ""
    changes: list[str] = field(default_factory=list)


def compute_diff(
    entries: list[ParsedEntry],
    existing: list[dict],
) -> tuple[list[DiffEntry], list[ParsedEntry], list[tuple[ParsedEntry, dict]]]:
    """Compute what changes need to be made.

    Returns (diff_entries, new_rows, update_pairs).
    """
    diffs: list[DiffEntry] = []
    new_rows: list[ParsedEntry] = []
    update_pairs: list[tuple[ParsedEntry, dict]] = []

    for entry in entries:
        match = match_existing(entry, existing)

        if match is None:
            new_rows.append(entry)
            diffs.append(DiffEntry(
                action="add",
                client_name=entry.client_name,
                bank_name=entry.bank_name,
                username=entry.username,
                account_last4=entry.account_last4,
            ))
        else:
            # Check what would change
            changes: list[str] = []
            old_pw = str(match.get("password", "") or "").strip()
            # Only fill password if existing is blank
            if not old_pw and entry.password:
                changes.append(f"password: (empty) -> '{entry.password[:3]}...'")

            old_notes = str(match.get("notes", "") or "").strip()
            new_notes = "; ".join(entry.notes_parts)
            if new_notes and new_notes != old_notes:
                merged = old_notes
                for part in entry.notes_parts:
                    if part not in merged:
                        merged = f"{merged}; {part}" if merged else part
                if merged != old_notes:
                    changes.append(f"notes: merged new metadata")

            if changes:
                update_pairs.append((entry, match))
                diffs.append(DiffEntry(
                    action="update",
                    client_name=entry.client_name,
                    bank_name=entry.bank_name,
                    username=entry.username,
                    account_last4=entry.account_last4,
                    changes=changes,
                ))
            else:
                diffs.append(DiffEntry(
                    action="skip",
                    client_name=entry.client_name,
                    bank_name=entry.bank_name,
                    username=entry.username,
                    account_last4=entry.account_last4,
                    reason="no changes needed",
                ))

    return diffs, new_rows, update_pairs


def entry_to_row_values(entry: ParsedEntry) -> dict:
    """Convert ParsedEntry to a dict of column values for Excel."""
    notes = "; ".join(entry.notes_parts) if entry.notes_parts else ""
    row = {
        "client_name": entry.client_name,
        "file_safe_name": entry.file_safe_name,
        "bank_name": entry.bank_name,
        "login_url": entry.login_url,
        "url_verified_date": None,
        "username": entry.username,
        "password": entry.password,
        "account_last4": entry.account_last4,
        "statement_available_date": None,
        "2fa": "y" if entry.has_2fa else "",
        "2fa_target": entry.tfa_target,
        "2fa_method": entry.tfa_method,
        "2fa_detail": entry.tfa_detail,
        "2fa_preference_order": "",
        "2fa_sender": "",
        "last_successful_login": None,
        "status": entry.status,
        "notes": notes,
    }

    # Security questions
    for i in range(1, 6):
        if i <= len(entry.security_questions):
            kw, ans = entry.security_questions[i - 1]
            row[f"sq{i}_keyword"] = kw
            row[f"sq{i}_answer"] = ans
        else:
            row[f"sq{i}_keyword"] = ""
            row[f"sq{i}_answer"] = ""

    return row


def apply_changes(
    wb: Workbook,
    new_rows: list[ParsedEntry],
    update_pairs: list[tuple[ParsedEntry, dict]],
) -> None:
    """Write changes to the workbook."""
    ws = wb["Accounts"]
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = col_idx

    # Apply updates
    for entry, match in update_pairs:
        row_num = match["_row_num"]

        # Only fill password if existing is blank (Excel is source of truth)
        old_pw = str(match.get("password", "") or "").strip()
        if not old_pw and entry.password:
            col = headers.get("password")
            if col:
                ws.cell(row=row_num, column=col, value=entry.password)

        # Merge notes
        old_notes = str(match.get("notes", "") or "").strip()
        for part in entry.notes_parts:
            if part and part not in old_notes:
                old_notes = f"{old_notes}; {part}" if old_notes else part
        col = headers.get("notes")
        if col:
            ws.cell(row=row_num, column=col, value=old_notes)

        # Update 2FA fields if entry has them and existing doesn't
        for field_name in ["2fa", "2fa_target", "2fa_method", "2fa_detail"]:
            col = headers.get(field_name)
            if not col:
                continue
            existing_val = str(match.get(field_name, "") or "").strip()
            if field_name == "2fa":
                new_val = "y" if entry.has_2fa else ""
            elif field_name == "2fa_target":
                new_val = entry.tfa_target
            elif field_name == "2fa_method":
                new_val = entry.tfa_method
            elif field_name == "2fa_detail":
                new_val = entry.tfa_detail
            else:
                new_val = ""
            if new_val and not existing_val:
                ws.cell(row=row_num, column=col, value=new_val)

        # Update security questions if entry has them and existing doesn't
        for i in range(1, 6):
            if i <= len(entry.security_questions):
                kw, ans = entry.security_questions[i - 1]
                kw_col = headers.get(f"sq{i}_keyword")
                ans_col = headers.get(f"sq{i}_answer")
                if kw_col and not str(match.get(f"sq{i}_keyword", "") or "").strip():
                    ws.cell(row=row_num, column=kw_col, value=kw)
                if ans_col and not str(match.get(f"sq{i}_answer", "") or "").strip():
                    ws.cell(row=row_num, column=ans_col, value=ans)

    # Append new rows
    for entry in new_rows:
        values = entry_to_row_values(entry)
        row_data = []
        for col_name in ACCOUNTS_COLUMNS:
            row_data.append(values.get(col_name, ""))
        ws.append(row_data)


def print_diff(diffs: list[DiffEntry]) -> None:
    """Print a human-readable diff report."""
    adds = [d for d in diffs if d.action == "add"]
    updates = [d for d in diffs if d.action == "update"]
    skips = [d for d in diffs if d.action == "skip"]

    print(f"\n{'='*70}")
    print(f"  IMPORT SUMMARY")
    print(f"{'='*70}")
    print(f"  New rows to add:    {len(adds)}")
    print(f"  Existing to update: {len(updates)}")
    print(f"  Skipped (no change):{len(skips)}")
    print(f"{'='*70}\n")

    if adds:
        print("--- NEW ROWS ---")
        for d in sorted(adds, key=lambda x: (x.client_name, x.bank_name)):
            print(f"  + {d.client_name:<30s} {d.bank_name:<25s} #{d.account_last4:<6s} user={d.username}")
        print()

    if updates:
        print("--- UPDATES ---")
        for d in sorted(updates, key=lambda x: (x.client_name, x.bank_name)):
            print(f"  ~ {d.client_name:<30s} {d.bank_name:<25s} #{d.account_last4:<6s}")
            for c in d.changes:
                print(f"      {c}")
        print()

    if skips:
        print(f"--- SKIPPED ({len(skips)} entries, no changes needed) ---")
        for d in sorted(skips, key=lambda x: (x.client_name, x.bank_name)):
            print(f"  = {d.client_name:<30s} {d.bank_name:<25s} #{d.account_last4:<6s}")
        print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Import 1Password bank credentials into clients.xlsx")
    parser.add_argument("--csv", required=True, help="Path to 1Password CSV export")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    group.add_argument("--apply", action="store_true", help="Apply changes to clients.xlsx")
    parser.add_argument("--xlsx", default=str(CLIENTS_XLSX), help="Path to clients.xlsx (default: bank-puller/clients.xlsx)")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    xlsx_path = Path(args.xlsx)

    if not csv_path.exists():
        print(f"ERROR: CSV not found: {csv_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"ERROR: Excel workbook not found: {xlsx_path}")
        sys.exit(1)

    # Step 1: Read 1Password CSV
    print(f"Reading 1Password CSV: {csv_path}")
    raw_rows = read_1password_csv(csv_path)
    print(f"  Total entries: {len(raw_rows)}")

    # Step 2: Parse and filter to bank entries
    entries = parse_entries(raw_rows)
    print(f"  Bank entries (before dedup): {len(entries)}")

    # Step 3: Dedup
    entries = dedup_entries(entries)
    print(f"  Bank entries (after dedup): {len(entries)}")

    # Step 4: Read existing workbook
    print(f"\nReading existing workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    existing = read_existing_accounts(wb)
    print(f"  Existing rows: {len(existing)}")

    # Step 5: Compute diff
    diffs, new_rows, update_pairs = compute_diff(entries, existing)
    print_diff(diffs)

    if args.dry_run:
        print("DRY RUN — no changes written.")
        wb.close()
        return

    # Step 6: Apply changes
    if not new_rows and not update_pairs:
        print("No changes to apply.")
        wb.close()
        return

    # Backup first
    backup_path = xlsx_path.with_suffix(".xlsx.bak")
    shutil.copy2(xlsx_path, backup_path)
    print(f"Backup saved: {backup_path}")

    apply_changes(wb, new_rows, update_pairs)
    wb.save(xlsx_path)
    wb.close()
    print(f"\nChanges applied to {xlsx_path}")
    print(f"  {len(new_rows)} rows added, {len(update_pairs)} rows updated")

    # Step 7: Verify with ExcelManager
    print("\nVerifying with ExcelManager...")
    try:
        with ExcelManager(xlsx_path) as mgr:
            accounts = mgr.read_accounts()
            print(f"  ExcelManager loaded {len(accounts)} accounts successfully.")
    except Exception as e:
        print(f"  WARNING: ExcelManager verification failed: {e}")
        print(f"  The backup is at {backup_path}")


if __name__ == "__main__":
    main()
